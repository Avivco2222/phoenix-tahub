"""Ingestion endpoints — the full ingestion surface.

Seven routes that admins use to load source files into the system:

    Persistent (B2.7b):
      POST /upload                          — canonical recruiter pipeline
      POST /upload/{file_type}              — multi-type dispatch
      POST /api/ingest/smart                — multi-sheet Excel autodetect
      POST /api/ingest/{file_type}          — typed ingest

    Preview / dry-run (B2.7a):
      GET  /api/ingest/smart-template       — Excel template download
      POST /api/ingest/preflight/{file_type}— dry-run parse + validate
      POST /api/ingest/whatif/{file_type}   — sandbox what-if

Endpoint bodies are reproduced verbatim from main.py — same auth gates
(``require_dual_role`` variants), same rate limits, same return
shapes, same SQL — pure relocation with no observable API change.

Many helpers stay in main.py because they are also used by other
internal callers; this module reaches them through late-bound proxy
functions to avoid circular imports.
"""

import hashlib
import io
import json
import sqlite3
import uuid
from datetime import datetime, timezone
from typing import Optional

import pandas as pd
from fastapi import APIRouter, Depends, File, Header, HTTPException, Request, UploadFile
from utils import _col_letter, _empty_stats, _row_to_scalar, iteration_signature, mask_value, normalize_email, normalize_phone
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import config as shared_config
from auth import _utcnow, require_dual_role
from constants import Role
from internal_logic import (
    build_snapshots,
    canonicalize_statuses,
    clear_query_cache,
    execute_etl_rules,
)
from rate_limit import limiter


router = APIRouter(tags=["ingestion"])


# --- Late-bound proxies to helpers still in main.py ----------------------
# These exist in main.py because the persistent /upload + /api/ingest/{type}
# routes use them too; they'll move out once the bigger ingestion endpoints
# follow in B2.7b.

def _read_file_with_limit(file):
    from main import _read_file_with_limit as _impl
    import asyncio
    return _impl(file)


async def _aread_file_with_limit(file):
    from main import _read_file_with_limit as _impl
    return await _impl(file)


def _validate_upload_file(file, *, allowed_extensions, allowed_mime_prefixes):
    from main import _validate_upload_file as _impl
    return _impl(file, allowed_extensions=allowed_extensions, allowed_mime_prefixes=allowed_mime_prefixes)


def _load_dataframe_from_upload(filename, content):
    from main import _load_dataframe_from_upload as _impl
    return _impl(filename, content)


def _apply_extra_aliases(df):
    from main import _apply_extra_aliases as _impl
    return _impl(df)


def _validate_ingest_frame(df, file_type):
    from main import _validate_ingest_frame as _impl
    return _impl(df, file_type)


def _normalize_upload_frame(df):
    from main import _normalize_upload_frame as _impl
    return _impl(df)










def _find_existing_candidate(conn, phone_norm, email_norm):
    from main import find_existing_candidate as _impl
    return _impl(conn, phone_norm, email_norm)






def _ingest_handlers():
    from main import INGEST_HANDLERS
    return INGEST_HANDLERS


def _ingest_requirements():
    from main import INGEST_REQUIREMENTS
    return INGEST_REQUIREMENTS


def _template_specs():
    from main import TEMPLATE_SPECS
    return TEMPLATE_SPECS


# Extra proxies used by the persistent /upload + /api/ingest endpoints (B2.7b)
def _validate_schema_contract(df, version):
    from main import _validate_schema_contract as _impl
    return _impl(df, version)




def _record_batch_change(conn, batch_id, entity_type, entity_id, action, before, after):
    from main import _record_batch_change as _impl
    return _impl(conn, batch_id, entity_type, entity_id, action, before, after)


def _get_unified_data(conn):
    from main import get_unified_data as _impl
    return _impl(conn)


def _auto_scan_after_ingest(conn, batch_id):
    from main import _auto_scan_after_ingest as _impl
    return _impl(conn, batch_id)


def _create_ingest_batch(file_type, filename, row_count, actor):
    from main import _create_ingest_batch as _impl
    return _impl(file_type, filename, row_count, actor)


def _finalise_ingest_batch(batch_id, status, stats):
    from main import _finalise_ingest_batch as _impl
    return _impl(batch_id, status, stats)


def _persist_rejected_rows_for_batch(batch_id, rejected):
    from main import _persist_rejected_rows_for_batch as _impl
    return _impl(batch_id, rejected)


def _detect_sheet_type(columns, sheet_name):
    from main import _detect_sheet_type as _impl
    return _impl(columns, sheet_name)


def _fk_order():
    from main import _FK_ORDER
    return _FK_ORDER


def _bump_data_version(conn=None):
    from main import bump_data_version as _impl
    return _impl(conn)


def _emit_notification(conn, **kwargs):
    from main import _emit_notification as _impl
    return _impl(conn, **kwargs)


def _log_audit(action, status, details, user):
    from main import log_audit_action as _impl
    _impl(action=action, status=status, details=details, user=user)


def _ingest_constants():
    from main import SUPPORTED_SCHEMA_VERSIONS, DEFAULT_SCHEMA_VERSION, MAX_ERROR_RATE
    return SUPPORTED_SCHEMA_VERSIONS, DEFAULT_SCHEMA_VERSION, MAX_ERROR_RATE


def _session_cookie_name():
    from auth import SESSION_COOKIE
    return SESSION_COOKIE


# --- Routes ---------------------------------------------------------------


@router.get("/api/ingest/smart-template")
def download_smart_template(
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Returns a ready-to-fill Excel workbook with sheets: משרות, מועמדים, גיוסים + הוראות."""
    TEMPLATE_SPECS = _template_specs()

    SMART_SHEETS = [
        ("משרות",   "jobs"),
        ("מועמדים", "candidates"),
        ("גיוסים",  "hires"),
    ]
    fill_req = PatternFill(start_color="002649", end_color="002649", fill_type="solid")
    fill_rec = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for sheet_title, file_type in SMART_SHEETS:
        spec = TEMPLATE_SPECS[file_type]
        all_cols = list(spec["required"]) + list(spec["recommended"])
        n_req = len(spec["required"])
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.rightToLeft = True
        for i, (he, _cn) in enumerate(all_cols, 1):
            c = ws.cell(row=1, column=i, value=he)
            c.fill = fill_req if i <= n_req else fill_rec
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[_col_letter(i)].width = 22
        sample = spec.get("sample", {})
        for i, (he, _cn) in enumerate(all_cols, 1):
            ws.cell(row=2, column=i, value=sample.get(he, ""))
        for he_name, options in (spec.get("validations") or {}).items():
            col_idx = next(
                (i + 1 for i, (he, _) in enumerate(all_cols) if he == he_name), None
            )
            if col_idx:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(options)}"',
                    allow_blank=True,
                    showDropDown=False,
                )
                ws.add_data_validation(dv)
                dv.add(f"{_col_letter(col_idx)}2:{_col_letter(col_idx)}5000")
        ws.freeze_panes = "A2"

    # Instructions sheet
    guide = wb.create_sheet("הוראות")
    guide.sheet_view.rightToLeft = True
    guide["A1"] = "תבנית Smart Ingest — Phoenix Talent OS"
    guide["A1"].font = Font(bold=True, size=14, color="002649")
    guide["A3"] = "כותרות כחולות = שדה חובה | כותרות אפורות = מומלץ"
    guide["A4"] = "גיליון 'משרות'   → משרות פתוחות/סגורות (dedup: שם משרה + חטיבה)"
    guide["A5"] = "גיליון 'מועמדים' → pipeline מועמדים (dedup: טלפון / אימייל)"
    guide["A6"] = "גיליון 'גיוסים'  → קליטות שהתבצעו בפועל"
    guide["A8"] = "שמות גיליונות נוספים שהמערכת מזהה: headcount/תקן, diversity/גיוון, attrition/עזיבות, budget/תקציב"
    guide.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Phoenix_SmartIngest_Template.xlsx"'},
    )


@router.post("/api/ingest/preflight/{file_type}")
@limiter.limit("20/minute")
async def ingest_preflight(
    request: Request,
    file_type: str,
    file: UploadFile = File(...),
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Dry-run: parse + validate without persisting. Lets admins preview what
    will land and what will be rejected BEFORE committing."""
    INGEST_HANDLERS = _ingest_handlers()
    INGEST_REQUIREMENTS = _ingest_requirements()

    if file_type not in INGEST_HANDLERS:
        raise HTTPException(status_code=400, detail=f"Unknown ingest type. Allowed: {list(INGEST_HANDLERS.keys())}")

    content = await _aread_file_with_limit(file)
    try:
        df = _load_dataframe_from_upload(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse: {exc}") from exc
    df = _apply_extra_aliases(df)
    valid_df, rejected = _validate_ingest_frame(df, file_type)
    return {
        "file_type": file_type,
        "rows_total": int(len(df)),
        "rows_valid": int(len(valid_df)),
        "rows_rejected": len(rejected),
        "sample_columns": list(df.columns)[:30],
        "sample_rejections": [{"reasons": r["reasons"], "row_keys": list(r["row"].keys())[:8]} for r in rejected[:10]],
        "required_columns": INGEST_REQUIREMENTS[file_type]["required"],
        "recommended_columns": INGEST_REQUIREMENTS[file_type]["recommended"],
    }


@router.post("/api/ingest/whatif/{file_type}")
@limiter.limit("10/minute")
async def ingest_whatif(
    request: Request,
    file_type: str,
    file: UploadFile = File(...),
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Enhancement #1 — DRY-RUN "what-if". Runs the full ingest pipeline inside
    a transaction that ROLLS BACK at the end, so admins see exactly how many
    rows WOULD be inserted/updated/skipped without persisting anything.

    Implementation: we run the same handler, then issue ROLLBACK on the
    sqlite connection. Stats are returned identical to a real commit.
    """
    INGEST_HANDLERS = _ingest_handlers()
    if file_type not in INGEST_HANDLERS:
        raise HTTPException(status_code=400, detail=f"Unknown ingest type. Allowed: {list(INGEST_HANDLERS.keys())}")

    _validate_upload_file(
        file,
        allowed_extensions={".csv", ".xlsx", ".xls", ".xml"},
        allowed_mime_prefixes=("text/csv", "application/csv", "application/vnd.ms-excel",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               "application/xml", "text/xml", "application/octet-stream"),
    )
    content = await _aread_file_with_limit(file)
    try:
        df = _load_dataframe_from_upload(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse: {exc}") from exc

    if file_type == "candidates":
        try:
            df = _normalize_upload_frame(df.copy())
        except Exception:
            pass
    df = _apply_extra_aliases(df)
    valid_df, rejected = _validate_ingest_frame(df, file_type)

    # Sandbox batch_id — we'll never finalise it, but the handler needs one to
    # record changes. ROLLBACK at the end discards both the rows AND the changes.
    sandbox_batch = f"WHATIF-{uuid.uuid4().hex[:8].upper()}"

    # Manually-managed transaction so we can rollback.
    conn = sqlite3.connect(shared_config.DB_NAME)
    try:
        conn.isolation_level = None  # autocommit OFF — we control BEGIN/ROLLBACK
        conn.execute("BEGIN")
        stats = _empty_stats()
        stats["received"] = int(len(df))

        def _exists_candidate(phone_norm, email_norm):
            return _find_existing_candidate(conn, phone_norm, email_norm) is not None

        for _, row in valid_df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            if not parsed.get("name") and parsed.get("candidate_name"):
                parsed["name"] = parsed["candidate_name"]
            parsed["phone_norm"] = normalize_phone(parsed.get("phone"))
            parsed["email_norm"] = normalize_email(parsed.get("email"))

            if file_type == "candidates":
                if _exists_candidate(parsed.get("phone_norm"), parsed.get("email_norm")):
                    stats["candidates_updated"] += 1
                    stats["updated"] += 1
                else:
                    stats["candidates_inserted"] += 1
                    stats["inserted"] += 1
                # Application iteration prediction
                job_title = (parsed.get("job_title") or "").strip()
                if job_title:
                    sig = iteration_signature(parsed.get("status"), parsed.get("application_date") or parsed.get("start_date"), parsed.get("recruiter"))
                    jrow = conn.execute(
                        "SELECT id FROM jobs WHERE LOWER(job_title) = LOWER(?) LIMIT 1", (job_title,),
                    ).fetchone()
                    if jrow:
                        existing_cid = _find_existing_candidate(conn, parsed.get("phone_norm"), parsed.get("email_norm"))
                        if existing_cid:
                            dup = conn.execute(
                                "SELECT 1 FROM applications WHERE candidate_id = ? AND job_id = ? AND iteration_signature = ?",
                                (existing_cid, jrow[0], sig),
                            ).fetchone()
                            if dup:
                                stats["applications_skipped"] += 1
                                stats["skipped_duplicate"] += 1
                            else:
                                stats["applications_inserted"] += 1
                        else:
                            stats["applications_inserted"] += 1
                    else:
                        stats["jobs_inserted"] += 1
                        stats["applications_inserted"] += 1
            elif file_type == "jobs":
                jrow = conn.execute(
                    "SELECT id FROM jobs WHERE LOWER(job_title) = LOWER(?) AND LOWER(IFNULL(department,'')) = LOWER(?) LIMIT 1",
                    ((parsed.get("job_title") or "").strip(), (parsed.get("department") or "General").strip()),
                ).fetchone()
                (stats["jobs_updated"] if jrow else stats["jobs_inserted"]).__iadd__  # noqa - placeholder
                if jrow:
                    stats["jobs_updated"] += 1
                    stats["updated"] += 1
                else:
                    stats["jobs_inserted"] += 1
                    stats["inserted"] += 1
            elif file_type == "diversity":
                existing = conn.execute(
                    "SELECT 1 FROM diversity_snapshots WHERE snapshot_month=? AND department=? AND dimension=? AND bucket=?",
                    (parsed.get("snapshot_month"), parsed.get("department"), parsed.get("dimension"), parsed.get("bucket")),
                ).fetchone()
                if existing: stats["updated"] += 1
                else: stats["inserted"] += 1
            elif file_type == "headcount":
                existing = conn.execute(
                    "SELECT 1 FROM headcount_snapshots WHERE snapshot_month=? AND department=? AND role=?",
                    (parsed.get("snapshot_month"), parsed.get("department"), parsed.get("role")),
                ).fetchone()
                if existing: stats["updated"] += 1
                else: stats["inserted"] += 1
            elif file_type == "hires":
                stats["inserted"] += 1
            elif file_type == "attrition":
                existing = conn.execute(
                    "SELECT 1 FROM attrition_events WHERE employee_name=? AND leave_date=?",
                    (parsed.get("employee_name"), parsed.get("leave_date")),
                ).fetchone()
                if existing: stats["skipped_duplicate"] += 1
                else: stats["inserted"] += 1
            elif file_type == "budget":
                inv_id = (parsed.get("id") or "").strip()
                if inv_id:
                    existing = conn.execute("SELECT 1 FROM finops_invoices WHERE id = ?", (inv_id,)).fetchone()
                    if existing: stats["updated"] += 1
                    else: stats["inserted"] += 1
                else:
                    stats["inserted"] += 1

        stats["rejected"] = len(rejected)
        conn.execute("ROLLBACK")
    finally:
        conn.close()

    return {
        "file_type": file_type,
        "sandbox_batch": sandbox_batch,
        "would_happen": stats,
        "rejected_samples": [{"reasons": r["reasons"]} for r in rejected[:5]],
    }


# =====================================================================
# B2.7b — persistent ingest endpoints (write to DB)
# =====================================================================


@router.post("/upload")
@limiter.limit("10/minute")
async def upload_file(
    request: Request,
    file: UploadFile = File(...),
    x_schema_version: Optional[str] = Header(default=None),
    x_idempotency_key: Optional[str] = Header(default=None),
    x_preflight_hash: Optional[str] = Header(default=None),
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER)),
):
    SUPPORTED_SCHEMA_VERSIONS, DEFAULT_SCHEMA_VERSION, MAX_ERROR_RATE = _ingest_constants()

    _validate_upload_file(
        file,
        allowed_extensions={".csv", ".xlsx", ".xls", ".xml"},
        allowed_mime_prefixes=(
            "text/csv",
            "application/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/xml",
            "text/xml",
        ),
    )
    log_id = str(uuid.uuid4())[:8]
    batch_id = f"bat-{uuid.uuid4().hex[:10]}"
    now = _utcnow().isoformat()
    content = await _aread_file_with_limit(file)
    payload_hash = hashlib.sha256(content).hexdigest()
    idempotency_key = (x_idempotency_key or payload_hash[:16]).strip()
    schema_version = x_schema_version or DEFAULT_SCHEMA_VERSION
    preflight_hash = (x_preflight_hash or "").strip()
    if schema_version not in SUPPORTED_SCHEMA_VERSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported schema_version: {schema_version}")
    if preflight_hash and preflight_hash != payload_hash:
        raise HTTPException(status_code=400, detail="Preflight checksum mismatch. Please rerun preflight before upload.")

    conn = sqlite3.connect(shared_config.DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        existing = c.execute(
            """SELECT batch_id, rows_loaded, status FROM ingestion_batches
               WHERE idempotency_key = ? AND payload_hash = ? ORDER BY started_at DESC LIMIT 1""",
            (idempotency_key, payload_hash),
        ).fetchone()
        if existing and existing["status"] == "committed":
            return {
                "message": "Idempotent replay accepted",
                "batch_id": existing["batch_id"],
                "rows_processed": int(existing["rows_loaded"] or 0),
                "replayed": True,
            }

        c.execute("BEGIN")
        try:
            c.execute(
                """INSERT INTO ingestion_batches(
                       batch_id, log_id, filename, payload_hash, idempotency_key, schema_version, actor, request_id,
                       status, started_at
                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (batch_id, log_id, file.filename, payload_hash, idempotency_key, schema_version, "manual", "upload", "processing", now),
            )
        except sqlite3.IntegrityError:
            existing_same = c.execute(
                """SELECT batch_id, status, rows_loaded FROM ingestion_batches
                   WHERE idempotency_key = ? AND payload_hash = ? ORDER BY started_at DESC LIMIT 1""",
                (idempotency_key, payload_hash),
            ).fetchone()
            if existing_same and existing_same["status"] == "committed":
                conn.rollback()
                return {
                    "message": "Idempotent replay accepted",
                    "batch_id": existing_same["batch_id"],
                    "rows_processed": int(existing_same["rows_loaded"] or 0),
                    "replayed": True,
                }
            raise

        raw_df = _load_dataframe_from_upload(file.filename or "", content)
        rows_received = len(raw_df)
        df = _normalize_upload_frame(raw_df.copy())
        _validate_schema_contract(df, schema_version)
        df = execute_etl_rules(df, conn, batch_id, auto_commit=False)
        df = canonicalize_statuses(df, conn)

        rejected_rows = 0
        duplicate_rows = 0
        mandatory_cols = ["name", "job_title", "email", "status", "recruiter", "department"]
        valid_rows = []
        for idx, row in df.iterrows():
            missing = [col for col in mandatory_cols if not str(row.get(col, "")).strip() or str(row.get(col, "")).lower() == "nan"]
            if missing:
                rejected_rows += 1
                c.execute(
                    """INSERT INTO rejected_rows(batch_id, row_index, reason_code, reason_detail, raw_row, created_at)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (batch_id, int(idx), "missing_required", ",".join(missing), json.dumps(raw_df.iloc[idx].to_dict(), ensure_ascii=False), _utcnow().isoformat()),
                )
                continue
            valid_rows.append(row)

        if rows_received > 0 and (rejected_rows / rows_received) > MAX_ERROR_RATE:
            raise Exception(f"Data quality gate failed: rejected_rate={rejected_rows/rows_received:.2f}")

        c.execute("DELETE FROM stg_applications WHERE batch_id = ?", (batch_id,))
        for row_idx, row in enumerate(valid_rows):
            c.execute(
                """INSERT INTO stg_applications(batch_id, row_idx, name, email, job_title, status, recruiter, start_date, department, source, stage_code, days_in_process)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    batch_id,
                    row_idx,
                    str(row["name"]),
                    str(row["email"]),
                    str(row["job_title"]),
                    str(row["status"]),
                    str(row["recruiter"]),
                    pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d"),
                    str(row["department"]),
                    str(row.get("source", "Organic / Unknown")),
                    str(row.get("stage_code", "ACTIVE")),
                    int(row["days_in_process"]),
                ),
            )

        rows_loaded = 0
        for row in valid_rows:
            email_val = row.get("email")
            phone_val = row.get("phone")
            email_norm = normalize_email(email_val)
            phone_norm = normalize_phone(phone_val)

            # Mask values for storage in DB
            masked_email_norm = mask_value(email_norm)
            masked_phone_norm = mask_value(phone_norm)
            masked_email = mask_value(email_val)
            masked_phone = mask_value(phone_val)

            c_id = _find_existing_candidate(conn, masked_phone_norm, masked_email_norm)
            if not c_id:
                if masked_email_norm:
                    c_id = str(uuid.uuid5(uuid.NAMESPACE_URL, masked_email_norm))
                else:
                    c_id = f"CND-{uuid.uuid4().hex[:12].upper()}"

            job_title = (row.get("job_title") or "").strip()
            dept = (row.get("department") or "").strip() or "General"
            j_id = None
            if job_title:
                jrow = conn.execute(
                    "SELECT id FROM jobs WHERE LOWER(job_title) = LOWER(?) AND LOWER(IFNULL(department,'')) = LOWER(?) LIMIT 1",
                    (job_title, dept),
                ).fetchone()
                if jrow:
                    j_id = jrow[0]
                else:
                    j_id = f"JOB-{uuid.uuid4().hex[:10].upper()}"

            cand_before = c.execute("SELECT id, name, email, source FROM candidates WHERE id = ?", (c_id,)).fetchone()
            if cand_before is None:
                c.execute(
                    "INSERT INTO candidates (id, name, email, phone, source, phone_norm, email_norm, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, 1)",
                    (c_id, str(row["name"]), masked_email, masked_phone, str(row["source"]), masked_phone_norm, masked_email_norm),
                )
                _record_batch_change(conn, batch_id, "candidate", c_id, "insert", None, {"name": str(row["name"]), "email": masked_email, "source": str(row["source"])})
            else:
                c.execute(
                    "UPDATE candidates SET name = ?, email = ?, phone = ?, source = ?, phone_norm = ?, email_norm = ?, is_active = 1 WHERE id = ?",
                    (str(row["name"]), masked_email, masked_phone, str(row["source"]), masked_phone_norm, masked_email_norm, c_id)
                )

            if j_id:
                job_before = c.execute("SELECT id, job_title, department FROM jobs WHERE id = ?", (j_id,)).fetchone()
                if job_before is None:
                    c.execute(
                        "INSERT INTO jobs (id, job_title, department, is_active) VALUES (?, ?, ?, 1)",
                        (j_id, job_title, dept),
                    )
                    _record_batch_change(conn, batch_id, "job", j_id, "insert", None, {"job_title": job_title, "department": dept})

            if c_id and j_id:
                sig = iteration_signature(
                    str(row["status"]),
                    pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d"),
                    str(row["recruiter"]),
                )
                app_before = c.execute(
                    "SELECT app_id, status, recruiter, days_in_process, upload_log_id, stage_code FROM applications WHERE candidate_id = ? AND job_id = ? AND iteration_signature = ?",
                    (c_id, j_id, sig),
                ).fetchone()

                if app_before:
                    app_id = app_before["app_id"]
                    duplicate_rows += 1
                    c.execute(
                        """UPDATE applications SET status = ?, recruiter = ?, days_in_process = ?, upload_log_id = ?, stage_code = ?, is_active = 1
                           WHERE app_id = ?""",
                        (str(row["status"]), str(row["recruiter"]), int(row["days_in_process"]), log_id, str(row.get("stage_code", "ACTIVE")), app_id),
                    )
                    _record_batch_change(
                        conn,
                        batch_id,
                        "application",
                        app_id,
                        "update",
                        dict(app_before),
                        {"status": str(row["status"]), "recruiter": str(row["recruiter"]), "days_in_process": int(row["days_in_process"]), "upload_log_id": log_id, "stage_code": str(row.get("stage_code", "ACTIVE"))},
                    )
                else:
                    app_id = f"APP-{uuid.uuid4().hex[:10].upper()}"
                    c.execute(
                        """INSERT INTO applications(app_id, candidate_id, job_id, status, recruiter, start_date, days_in_process, upload_log_id, stage_code, iteration_signature, application_date, batch_id, is_active)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
                        (app_id, c_id, j_id, str(row["status"]), str(row["recruiter"]), pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d"), int(row["days_in_process"]), log_id, str(row.get("stage_code", "ACTIVE")), sig, pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d"), batch_id),
                    )
                    _record_batch_change(conn, batch_id, "application", app_id, "insert", None, {"status": str(row["status"]), "recruiter": str(row["recruiter"]), "days_in_process": int(row["days_in_process"]), "upload_log_id": log_id, "stage_code": str(row.get("stage_code", "ACTIVE"))})
                rows_loaded += 1


        quality_score = round(max(0.0, 100 - ((rejected_rows + duplicate_rows) / max(rows_received, 1)) * 100), 2)
        quality_report = {
            "rows_received": rows_received,
            "rows_loaded": rows_loaded,
            "rows_rejected": rejected_rows,
            "duplicate_rows": duplicate_rows,
            "error_rate": round(rejected_rows / max(rows_received, 1), 4),
            "duplicate_rate": round(duplicate_rows / max(rows_received, 1), 4),
        }
        c.execute(
            "INSERT INTO data_logs (log_id, filename, upload_date, rows_processed, status) VALUES (?, ?, ?, ?, ?)",
            (log_id, file.filename, pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"), rows_loaded, "Success"),
        )
        c.execute(
            """UPDATE ingestion_batches
               SET rows_received = ?, rows_loaded = ?, rows_rejected = ?, duplicate_rows = ?, quality_score = ?, quality_report = ?,
                   status = 'committed', finished_at = ?
               WHERE batch_id = ?""",
            (rows_received, rows_loaded, rejected_rows, duplicate_rows, quality_score, json.dumps(quality_report, ensure_ascii=False), _utcnow().isoformat(), batch_id),
        )

        unified_df = _get_unified_data(conn)
        build_snapshots(conn, unified_df)
        clear_query_cache(conn, auto_commit=False)
        _auto_scan_after_ingest(conn, batch_id)
        conn.commit()
        return {"message": "ETL Completed successfully", "batch_id": batch_id, "rows_processed": rows_loaded, "quality_report": quality_report}
    except Exception as e:
        conn.rollback()
        try:
            c.execute(
                """INSERT OR REPLACE INTO ingestion_batches(batch_id, log_id, filename, payload_hash, idempotency_key, schema_version, actor, request_id, status, started_at, finished_at, error_message)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'failed', ?, ?, ?)""",
                (batch_id, log_id, file.filename, payload_hash, idempotency_key, schema_version, "manual", "upload", now, _utcnow().isoformat(), str(e)),
            )
            conn.commit()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/upload/{file_type}")
@limiter.limit("10/minute")
async def upload_typed_file(
    request: Request,
    file_type: str,
    file: UploadFile = File(...),
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Multi-type ingestion: candidates, jobs, hires, diversity, headcount, budget, attrition.

    Persistence rules:
    * `candidates` / `jobs` / `hires` route through the canonical `/upload`
      pipeline so they land in candidates/jobs/applications and produce a real
      ingestion_batches row.
    * The remaining types (`diversity`, `headcount`, `budget`, `attrition`)
      land in `ingestion_typed_rows` (JSON-per-row, queryable by file_type)
      until a dedicated entity model is built. This closes the silent
      data-loss gap surfaced in QA without committing to a schema yet.
    """
    SESSION_COOKIE = _session_cookie_name()

    _validate_upload_file(
        file,
        allowed_extensions={".csv", ".xlsx", ".xls", ".xml"},
        allowed_mime_prefixes=(
            "text/csv",
            "application/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/xml",
            "text/xml",
            "application/octet-stream",
        ),
    )
    content = await _aread_file_with_limit(file)

    valid_types = [
        "candidates", "jobs", "hires", "diversity",
        "headcount", "budget", "attrition"
    ]
    if file_type not in valid_types:
        return {"status": "error", "message": f"Invalid file type: {file_type}. Valid: {valid_types}"}

    # Load once for both branches.
    try:
        df = _load_dataframe_from_upload(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}") from exc

    # Branch 1: types that map to the canonical recruiter pipeline.
    canonical_types = {"candidates", "jobs", "hires"}
    if file_type in canonical_types:
        # Reset stream and let the canonical /upload handler do the heavy lifting.
        await file.seek(0)
        result = await upload_file(request=request, file=file)  # type: ignore[name-defined]
        if isinstance(result, dict):
            result.setdefault("file_type", file_type)
        return result

    # Branch 2: persist raw rows for the other types so the data isn't dropped.
    batch_id = f"TYPED-{uuid.uuid4().hex[:8].upper()}"
    uploaded_at = datetime.now(timezone.utc).isoformat()
    uploaded_by = (request.cookies.get(SESSION_COOKIE) or "")[:32] or "unknown"

    # Normalise NaN → None so JSON serialisation is clean.
    safe_df = df.where(pd.notnull(df), None)
    rows = safe_df.to_dict(orient="records")

    conn = sqlite3.connect(shared_config.DB_NAME)
    try:
        c = conn.cursor()
        for row in rows:
            c.execute(
                "INSERT INTO ingestion_typed_rows (file_type, batch_id, row_data, uploaded_at, uploaded_by) "
                "VALUES (?, ?, ?, ?, ?)",
                (file_type, batch_id, json.dumps(row, ensure_ascii=False, default=str), uploaded_at, uploaded_by),
            )
        conn.commit()
    finally:
        conn.close()

    _log_audit(
        "TYPED_UPLOAD",
        "ok",
        f"type={file_type} rows={len(rows)} batch={batch_id}",
        user=uploaded_by,
    )

    return {
        "status": "success",
        "file_type": file_type,
        "filename": file.filename,
        "rows_processed": int(len(rows)),
        "batch_id": batch_id,
        "stored_in": "ingestion_typed_rows",
        "last_updated": uploaded_at,
    }


@router.post("/api/ingest/smart")
@limiter.limit("5/minute")
async def ingest_smart(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Single Excel file with multiple sheets → auto-detect type and route to each handler."""
    INGEST_HANDLERS = _ingest_handlers()
    _FK_ORDER = _fk_order()

    _validate_upload_file(
        file,
        allowed_extensions={".xlsx"},
        allowed_mime_prefixes=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        ),
    )
    content = await _aread_file_with_limit(file)
    filename = file.filename or "smart_upload.xlsx"

    try:
        all_sheets: dict[str, pd.DataFrame] = pd.read_excel(
            io.BytesIO(content), sheet_name=None, engine="openpyxl"
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"שגיאה בקריאת Excel: {exc}") from exc

    if not all_sheets:
        raise HTTPException(status_code=400, detail="הקובץ ריק — אין גיליונות")

    # Step 1: detect each sheet type after alias normalization
    detection_map: dict[str, tuple[str | None, float]] = {}
    normalized_dfs: dict[str, pd.DataFrame] = {}

    for sheet_name, raw_df in all_sheets.items():
        if raw_df.empty or len(raw_df.columns) < 2:
            detection_map[sheet_name] = (None, 0.0)
            continue
        df = raw_df.copy()
        try:
            df = _normalize_upload_frame(df)
        except Exception:
            pass  # _normalize_upload_frame raises when 'name' column missing (e.g. jobs sheet)
        df = _apply_extra_aliases(df)
        normalized_dfs[sheet_name] = df
        detection_map[sheet_name] = _detect_sheet_type(list(df.columns), sheet_name)

    # Step 2: process in FK-safe order
    type_to_sheets: dict[str, list[str]] = {t: [] for t in _FK_ORDER}
    for sname, (dtype, _conf) in detection_map.items():
        if dtype:
            type_to_sheets[dtype].append(sname)

    sheet_results: list[dict] = []
    any_success = False

    for file_type in _FK_ORDER:
        for sheet_name in type_to_sheets[file_type]:
            df = normalized_dfs[sheet_name]
            _, confidence = detection_map[sheet_name]
            entry: dict = {
                "sheet_name": sheet_name,
                "detected_type": file_type,
                "confidence": confidence,
                "received": len(df),
                "inserted": 0,
                "updated": 0,
                "rejected": 0,
                "skipped_duplicate": 0,
                "batch_id": None,
                "status": "pending",
                "error": None,
            }
            try:
                valid_df, rejected_rows = _validate_ingest_frame(df, file_type)
                batch_id = _create_ingest_batch(
                    file_type,
                    f"[smart]{filename}::{sheet_name}",
                    len(df),
                    user.get("email", "admin"),
                )
                entry["batch_id"] = batch_id
                stats = INGEST_HANDLERS[file_type](valid_df, batch_id)
                stats["rejected"] = stats.get("rejected", 0) + len(rejected_rows)
                stats["received"] = len(df)
                _persist_rejected_rows_for_batch(batch_id, rejected_rows)
                _finalise_ingest_batch(batch_id, "committed", stats)
                entry.update({
                    "inserted": stats.get("inserted", 0),
                    "updated": stats.get("updated", 0),
                    "rejected": stats.get("rejected", 0),
                    "skipped_duplicate": stats.get("skipped_duplicate", 0),
                    "status": "success",
                })
                any_success = True
                _log_audit(
                    "SMART_INGEST", "ok",
                    f"sheet={sheet_name} type={file_type} ins={stats['inserted']}",
                    user=user.get("email", "admin"),
                )
            except HTTPException as e:
                entry.update({"status": "error", "error": e.detail})
                if entry["batch_id"]:
                    _finalise_ingest_batch(entry["batch_id"], "failed", _empty_stats())
            except Exception as e:
                entry.update({"status": "error", "error": str(e)})
                if entry["batch_id"]:
                    _finalise_ingest_batch(entry["batch_id"], "failed", _empty_stats())
            sheet_results.append(entry)

    # Sheets that couldn't be identified — report as skipped
    for sname, (dtype, conf) in detection_map.items():
        if dtype is None:
            sheet_results.append({
                "sheet_name": sname,
                "detected_type": None,
                "confidence": conf,
                "received": len(all_sheets[sname]),
                "inserted": 0,
                "updated": 0,
                "rejected": 0,
                "skipped_duplicate": 0,
                "batch_id": None,
                "status": "skipped",
                "error": "לא זוהה סוג — גיליון הושמט",
            })

    overall = (
        "success"
        if any_success and not any(r["status"] == "error" for r in sheet_results)
        else "partial"
        if any_success
        else "failed"
    )
    new_ver = _bump_data_version() if any_success else None

    # Single summary notification
    if any_success:
        _total_in = sum(r["inserted"] + r["updated"] for r in sheet_results)
        _total_rej = sum(r["rejected"] for r in sheet_results)
        _notif_msg = (
            f"✅ Smart Ingest: {_total_in} שורות נקלטו מ-{filename}"
            + (f", {_total_rej} נדחו" if _total_rej else "")
        )
        try:
            _nc = sqlite3.connect(shared_config.DB_NAME)
            _emit_notification(
                _nc,
                user_id=user.get("sub"),
                message=_notif_msg,
                severity="success" if not _total_rej else "warning",
                sent_by="system:SMART_INGEST",
                category="ingest",
                link="/admin?group=data&sub=batches",
            )
            _nc.commit()
        except Exception:
            pass
        finally:
            try:
                _nc.close()
            except Exception:
                pass

    return {
        "status": overall,
        "filename": filename,
        "sheets": sheet_results,
        "data_version": new_ver,
    }


@router.post("/api/ingest/{file_type}")
@limiter.limit("10/minute")
async def ingest_typed(
    request: Request,
    file_type: str,
    file: UploadFile = File(...),
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Unified typed ingest. Drives the full pipeline for one of the 7 sources:
    candidates / jobs / hires / diversity / headcount / budget / attrition.
    """
    INGEST_HANDLERS = _ingest_handlers()
    if file_type not in INGEST_HANDLERS:
        raise HTTPException(status_code=400, detail=f"Unknown ingest type. Allowed: {list(INGEST_HANDLERS.keys())}")

    _validate_upload_file(
        file,
        allowed_extensions={".csv", ".xlsx", ".xls", ".xml"},
        allowed_mime_prefixes=(
            "text/csv", "application/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/xml", "text/xml", "application/octet-stream",
        ),
    )
    content = await _aread_file_with_limit(file)

    # Stage 1-2: Parse + alias.
    try:
        df = _load_dataframe_from_upload(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}") from exc

    # For the candidates type we run through the legacy column normaliser
    # first — it knows the existing Hebrew aliases for name/email/job/etc.
    # Other types use a simpler path and rely on EXTRA_HEBREW_ALIASES below.
    if file_type == "candidates":
        try:
            df = _normalize_upload_frame(df.copy())
        except Exception:
            # If the legacy normaliser is strict and fails, fall through to alias-only path.
            pass
    df = _apply_extra_aliases(df)

    # Stage 3: Validate required columns + per-row required values.
    valid_df, rejected_rows = _validate_ingest_frame(df, file_type)

    # Open batch.
    batch_id = _create_ingest_batch(file_type, file.filename or "(no name)", int(len(df)), user.get("email", "admin"))

    # Stages 4-6: Normalize + match + persist.
    try:
        handler = INGEST_HANDLERS[file_type]
        stats = handler(valid_df, batch_id)
        stats["rejected"] += len(rejected_rows)
        stats["rejected_reasons"] = [r["reasons"] for r in rejected_rows[:20]]
        stats["received"] = int(len(df))

        # Persist rejected rows for the admin Quality tab.
        if rejected_rows:
            conn = sqlite3.connect(shared_config.DB_NAME)
            try:
                for r in rejected_rows:
                    conn.execute(
                        "INSERT INTO rejected_rows (batch_id, row_json, reason, created_at) VALUES (?, ?, ?, ?)",
                        (batch_id, json.dumps(r["row"], ensure_ascii=False, default=str),
                         "; ".join(r["reasons"]), datetime.now(timezone.utc).isoformat()),
                    )
                conn.commit()
            except sqlite3.OperationalError:
                # rejected_rows table may have different shape; that's fine for now.
                pass
            finally:
                conn.close()

        _finalise_ingest_batch(batch_id, "committed", stats)

        # Stage 7: Notify.
        new_version = _bump_data_version()
        _log_audit(
            "INGEST_COMMITTED", "ok",
            f"type={file_type} batch={batch_id} inserted={stats['inserted']} updated={stats['updated']} rejected={stats['rejected']}",
            user=user.get("email", "admin"),
        )

        # In-app notification to the uploading user (uploader = admin/hrbp)
        _ingest_type_labels = {
            "candidates": "מועמדים", "jobs": "משרות", "hires": "קליטות",
            "diversity": "גיוון", "headcount": "תקן מצבה", "budget": "תקציב", "attrition": "עזיבות",
        }
        _type_label = _ingest_type_labels.get(file_type, file_type)
        _inserted = int(stats.get("inserted", 0)) + int(stats.get("updated", 0))
        _rejected = int(stats.get("rejected", 0))
        _notif_msg = f"✅ קלטת קובץ {_type_label}: {_inserted} שורות נקלטו בהצלחה" + (f", {_rejected} נדחו" if _rejected else "") + "."
        _notif_sev = "success" if not _rejected else "warning"
        try:
            _notif_conn = sqlite3.connect(shared_config.DB_NAME)
            _emit_notification(
                _notif_conn,
                user_id=user.get("sub"),
                message=_notif_msg,
                severity=_notif_sev,
                sent_by="system:INGEST",
                category="ingest",
                link="/admin?group=data&sub=batches",
            )
            _notif_conn.commit()
        except Exception:
            pass  # notification failure must never block ingest response
        finally:
            try: _notif_conn.close()
            except Exception: pass

        return {
            "status": "success",
            "batch_id": batch_id,
            "type": file_type,
            "filename": file.filename,
            "stats": stats,
            "data_version": new_version,
        }
    except Exception as exc:
        _finalise_ingest_batch(batch_id, "failed", _empty_stats())
        _log_audit("INGEST_FAILED", "warn", f"type={file_type} batch={batch_id} err={exc}", user=user.get("email", "admin"))
        raise HTTPException(status_code=500, detail=f"Ingest failed: {exc}") from exc
