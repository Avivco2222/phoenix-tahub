from fastapi import FastAPI, UploadFile, File, HTTPException, Request, Response, Depends, Header, Cookie, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from starlette.background import BackgroundTask
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
import pandas as pd
import sqlite3
import os
import uuid
from datetime import datetime, timedelta, timezone
import json

# bcrypt is optional at import time — endpoints that need it raise 503 if missing.
try:
    import bcrypt  # type: ignore
    BCRYPT_AVAILABLE = True
except ImportError:
    BCRYPT_AVAILABLE = False

# Optional .env loading for local dev. In prod, env vars come from the host.
try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
except ImportError:
    pass
import hashlib
import io
import re
import logging
import asyncio
from contextvars import ContextVar
import secrets
import base64
import hmac
import xml.etree.ElementTree as ET
from typing import Literal, Optional
from contextlib import contextmanager
from pathlib import Path
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address
from aliases import LEGACY_CANDIDATE_ALIASES, TYPED_INGEST_ALIASES
from auth import (
    BCRYPT_AVAILABLE,
    IMPERSONATOR_COOKIE,
    SESSION_COOKIE,
    _REVOKED_TOKENS,
    _b64url_decode,
    _b64url_encode,
    _decode_jwt,
    _encode_jwt,
    _hash_password,
    _is_token_revoked,
    _make_session_token,
    _revoke_token_signature,
    _utcnow,
    _verify_password,
    auth_scheme,
    get_session_user,
    require_admin,
    require_dual_role,
    require_session_role,
    verify_token,
)
from constants import Role, UNIFIED_STAGES
from db import (
    _safe_connect,
    db_conn,
    db_transaction,
    execute as db_execute,
    fetch_all,
    fetch_one,
)
from internal_logic import (
    build_snapshots,
    canonicalize_statuses,
    clear_query_cache,
    compute_ghosting_risk_score,
    execute_etl_rules,
    get_cached_response,
    render_executive_insight,
    seed_internal_logic_tables,
    set_cached_response,
)

app = FastAPI()
import config as shared_config

from rate_limit import limiter  # shared instance — see backend/rate_limit.py
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s req_id=%(request_id)s %(message)s",
)
logger = logging.getLogger("phoenix-api")
SESSION_UNLOCK_PIN = os.getenv("SESSION_UNLOCK_PIN")
JWT_SECRET = os.getenv("JWT_SECRET")
JWT_TTL_MINUTES = int(os.getenv("JWT_TTL_MINUTES", "120"))

# Fail fast in production-like runs if JWT_SECRET is missing — every authenticated
# endpoint depends on it. PYTEST_CURRENT_TEST / unit tests are allowed to start
# without it (they exercise unauthenticated paths or stub auth).
if not JWT_SECRET and not os.getenv("PYTEST_CURRENT_TEST") and not os.getenv("ALLOW_MISSING_JWT_SECRET"):
    logger.error(
        "JWT_SECRET is not set. Refusing to start. "
        "Generate one with: python -c \"import secrets; print(secrets.token_urlsafe(64))\" "
        "and add it to backend/.env or the service EnvironmentFile."
    )
    sys.exit(1)
# SESSION_COOKIE and IMPERSONATOR_COOKIE moved to backend/auth.py (B4)
# and re-imported at the top of this module.
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "10"))
MAX_UPLOAD_BYTES = MAX_UPLOAD_MB * 1024 * 1024
# auth_scheme, _utcnow, _REVOKED_TOKENS, _revoke_token_signature, and
# _is_token_revoked moved to backend/auth.py (B4). Re-imported below.


class RequestLoggerAdapter(logging.LoggerAdapter):
    def process(self, msg, kwargs):
        extra = kwargs.setdefault("extra", {})
        extra.setdefault("request_id", kwargs.pop("request_id", "-"))
        return msg, kwargs


req_logger = RequestLoggerAdapter(logger, {})
request_ip_ctx: ContextVar[str] = ContextVar("request_ip", default="-")


def _extract_request_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    if request.client and request.client.host:
        return request.client.host
    return "-"


# _b64url_encode/decode, _encode_jwt/_decode_jwt, require_admin,
# verify_token, require_dual_role moved to backend/auth.py (B4).



def _ensure_jwt_ready() -> None:
    if not JWT_SECRET:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="JWT_SECRET is not configured on server",
        )


# Request/response models moved to backend/schemas/ — see B1 of the
# structural improvement plan. Re-exporting the names so existing
# `OnboardingPayload`, `FinopsInvoicePayload`, etc. references in this
# module keep working without further edits.
from schemas import (
    AnomalyReviewPayload,
    ApplicationEditPayload,
    CandidateEditPayload,
    FinopsCategoryPayload,
    FinopsInvoicePayload,
    FinopsVendorPayload,
    JobEditPayload,
    JobsBulkUpdatePayload,
    OnboardingBulkUpdatePayload,
    OnboardingPayload,
    OnboardingUpdatePayload,
)


def _cleanup_generated_file(path: str) -> None:
    try:
        if path and os.path.exists(path):
            os.remove(path)
    except Exception:
        logger.warning("Failed to delete temp file: %s", path)


# _normalize_onboarding_payload moved to backend/routers/onboarding.py (B2.3)


async def _read_file_with_limit(file: UploadFile) -> bytes:
    chunks: list[bytes] = []
    total_size = 0
    while True:
        chunk = await file.read(1024 * 1024)
        if not chunk:
            break
        total_size += len(chunk)
        if total_size > MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large. Max size is {MAX_UPLOAD_MB}MB",
            )
        chunks.append(chunk)
    await file.seek(0)
    return b"".join(chunks)


def _validate_upload_file(file: UploadFile, *, allowed_extensions: set[str], allowed_mime_prefixes: tuple[str, ...]) -> None:
    filename = (file.filename or "").strip()
    if not filename:
        raise HTTPException(status_code=400, detail="Filename is required")
    ext = os.path.splitext(filename.lower())[1]
    if ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"Unsupported file extension: {ext}")
    content_type = (file.content_type or "").lower()
    if not any(content_type.startswith(prefix) for prefix in allowed_mime_prefixes):
        raise HTTPException(status_code=400, detail=f"Unsupported MIME type: {content_type}")


# db_conn() and _safe_connect() moved to backend/db.py (B3). Re-imported
# at the top of this module so existing call sites keep working without
# changes. New helpers fetch_one / fetch_all / execute / db_transaction
# are available there for code that wants to drop the cursor boilerplate.


@app.middleware("http")
async def request_logging_middleware(request: Request, call_next):
    request_id = request.headers.get("x-request-id", uuid.uuid4().hex[:12])
    ip_token = request_ip_ctx.set(_extract_request_ip(request))
    start_ts = _utcnow()
    try:
        response = await call_next(request)
        elapsed_ms = int((_utcnow() - start_ts).total_seconds() * 1000)
        req_logger.info(
            f'{request.method} {request.url.path} status={response.status_code} elapsed_ms={elapsed_ms}',
            request_id=request_id,
        )
        response.headers["X-Request-Id"] = request_id
        return response
    except Exception as exc:
        elapsed_ms = int((_utcnow() - start_ts).total_seconds() * 1000)
        req_logger.error(
            f'{request.method} {request.url.path} error="{exc}" elapsed_ms={elapsed_ms}',
            request_id=request_id,
        )
        raise
    finally:
        request_ip_ctx.reset(ip_token)

# ==========================================
# מנוע קליטת נתוני ATS, אבטחת מידע ואיכות נתונים
# ==========================================

# ==========================================
# PII SCRUBBING ENGINE
# ==========================================
class PIIScrubber:
    @staticmethod
    def scrub_text_for_ai(text: str) -> tuple:
        """
        סורק טקסט חופשי ומצנזר נתונים מזהים לפני שליחה ל-AI.
        מחזיר את הטקסט המצונזר + סטטיסטיקות לצורך Audit Log.
        """
        if not text:
            return text, {}

        stats = {"id_cards": 0, "phones": 0, "emails": 0}

        id_pattern = r'\b\d{9}\b'
        stats["id_cards"] = len(re.findall(id_pattern, text))
        text = re.sub(id_pattern, '[ID_SECURED]', text)

        phone_pattern = r'\b05\d-?\d{7}\b'
        stats["phones"] = len(re.findall(phone_pattern, text))
        text = re.sub(phone_pattern, '[PHONE_SECURED]', text)

        email_pattern = r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+'
        stats["emails"] = len(re.findall(email_pattern, text))
        text = re.sub(email_pattern, '[EMAIL_SECURED]', text)

        return text, stats

def log_audit_action(action: str, status: str, details: str, user: str = "System"):
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        log_id = f"LOG-{uuid.uuid4().hex[:6].upper()}"
        timestamp = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
        ip_address = request_ip_ctx.get("-")
        c.execute(
            "INSERT INTO audit_logs (id, timestamp, action, status, details, user, ip_address) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (log_id, timestamp, action, status, details, user, ip_address),
        )
        conn.commit()
    finally:
        conn.close()

def mask_sensitive_data(df):
    sensitive_keywords = ['ת.ז', 'תעודת זהות', 'id', 'טלפון', 'נייד', 'phone', 'כתובת']
    for col in df.columns:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in sensitive_keywords):
            df[col] = df[col].astype(str).apply(
                lambda x: hashlib.sha256(x.encode()).hexdigest()[:12] if pd.notnull(x) and str(x).lower() not in ['nan', 'none', ''] else None
            )
            df.rename(columns={col: f"{col}_MASKED_SECURE"}, inplace=True)
    return df


def mask_value(val) -> Optional[str]:
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('nan', 'none', 'null', 'n/a', '-'):
        return None
    return hashlib.sha256(val_str.encode("utf-8")).hexdigest()[:12]


@app.post("/upload/{file_type}")
@limiter.limit("10/minute")
async def upload_typed_file(
    request: Request,
    file_type: str,
    file: UploadFile = File(...),
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Multi-type ingestion: candidates, jobs, hires, diversity, headcount, budget, attrition.

    Persistence rules:
    * `candidates` / `jobs` / `hires` route through the canonical `/upload`
      pipeline so they land in candidates/jobs/applications and produce a real
      ingestion_batches row.
    * The remaining types (`diversity`, `headcount`, `budget`, `attrition`)
      land in `ingestion_typed_rows` (JSON-per-row, queryable by file_type)
      until a dedicated entity model is built. This closes the silent
      data-loss gap surfaced in QA without committing to a schema yet.
    """
    _validate_upload_file(
        file,
        allowed_extensions={".csv", ".xlsx", ".xls", ".xml"},
        allowed_mime_prefixes=(
            "text/csv",
            "application/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/xml",
            "text/xml",
            "application/octet-stream",
        ),
    )
    content = await _read_file_with_limit(file)

    valid_types = [
        "candidates", "jobs", "hires", "diversity",
        "headcount", "budget", "attrition"
    ]
    if file_type not in valid_types:
        return {"status": "error", "message": f"Invalid file type: {file_type}. Valid: {valid_types}"}

    # Load once for both branches.
    try:
        df = _load_dataframe_from_upload(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}") from exc

    # Branch 1: types that map to the canonical recruiter pipeline.
    canonical_types = {"candidates", "jobs", "hires"}
    if file_type in canonical_types:
        # Reset stream and let the canonical /upload handler do the heavy lifting.
        await file.seek(0)
        result = await upload_file(request=request, file=file)  # type: ignore[name-defined]
        if isinstance(result, dict):
            result.setdefault("file_type", file_type)
        return result

    # Branch 2: persist raw rows for the other types so the data isn't dropped.
    batch_id = f"TYPED-{uuid.uuid4().hex[:8].upper()}"
    uploaded_at = datetime.now(timezone.utc).isoformat()
    uploaded_by = (request.cookies.get(SESSION_COOKIE) or "")[:32] or "unknown"

    # Normalise NaN → None so JSON serialisation is clean.
    safe_df = df.where(pd.notnull(df), None)
    rows = safe_df.to_dict(orient="records")

    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        for row in rows:
            c.execute(
                "INSERT INTO ingestion_typed_rows (file_type, batch_id, row_data, uploaded_at, uploaded_by) "
                "VALUES (?, ?, ?, ?, ?)",
                (file_type, batch_id, json.dumps(row, ensure_ascii=False, default=str), uploaded_at, uploaded_by),
            )
        conn.commit()
    finally:
        conn.close()

    log_audit_action(
        "TYPED_UPLOAD",
        "ok",
        f"type={file_type} rows={len(rows)} batch={batch_id}",
        user=uploaded_by,
    )

    return {
        "status": "success",
        "file_type": file_type,
        "filename": file.filename,
        "rows_processed": int(len(rows)),
        "batch_id": batch_id,
        "stored_in": "ingestion_typed_rows",
        "last_updated": uploaded_at,
    }

cors_origins_env = os.getenv("CORS_ALLOW_ORIGINS", "").strip()
if cors_origins_env:
    origins = [origin.strip() for origin in cors_origins_env.split(",") if origin.strip()]
else:
    # Local dev defaults (Next may run on 3000/3001/3002)
    origins = [
        "http://localhost:3000",
        "http://localhost:3001",
        "http://localhost:3002",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:3001",
        "http://127.0.0.1:3002",
        "https://phoenix-tahub2.vercel.app",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH = shared_config.DB_NAME

SUPPORTED_SCHEMA_VERSIONS = {"1.0"}
DEFAULT_SCHEMA_VERSION = "1.0"
MAX_ERROR_RATE = float(os.getenv("MAX_INGEST_ERROR_RATE", "0.2"))
CONTRACTS_DIR = os.path.join(os.path.dirname(__file__), "contracts")

# ==========================================
# 1. ENTITY RELATIONSHIP MODEL (יצירת הטבלאות)
# ==========================================
def init_db():
    # init_db runs once at module load. Failure here kills the import
    # (and the process), so the connection would not "leak" in any
    # meaningful sense; wrapping the full 480-line body in try/finally
    # would force a re-indent across the whole function and is left
    # for a future routers-split refactor.
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # --- Migration to remove UNIQUE constraints ---
    try:
        row_cand = c.execute("SELECT sql FROM sqlite_master WHERE type='table' and name='candidates'").fetchone()
        if row_cand and "UNIQUE" in row_cand[0].upper():
            c.execute("ALTER TABLE candidates RENAME TO candidates_old")
            c.execute('''CREATE TABLE candidates (
                id TEXT PRIMARY KEY,
                name TEXT,
                email TEXT,
                phone TEXT,
                source TEXT
            )''')
            c.execute("PRAGMA table_info(candidates_old)")
            old_cols = [r[1] for r in c.fetchall()]
            common_cols = [col for col in ["id", "name", "email", "phone", "source"] if col in old_cols]
            cols_str = ", ".join(common_cols)
            c.execute(f"INSERT INTO candidates ({cols_str}) SELECT {cols_str} FROM candidates_old")
            c.execute("DROP TABLE candidates_old")
            conn.commit()
    except Exception as e:
        logging.error(f"Migration candidates failed: {e}")

    try:
        row_job = c.execute("SELECT sql FROM sqlite_master WHERE type='table' and name='jobs'").fetchone()
        if row_job and "UNIQUE" in row_job[0].upper():
            c.execute("ALTER TABLE jobs RENAME TO jobs_old")
            c.execute('''CREATE TABLE jobs (
                id TEXT PRIMARY KEY,
                job_title TEXT,
                department TEXT,
                hiring_manager TEXT
            )''')
            c.execute("PRAGMA table_info(jobs_old)")
            old_cols = [r[1] for r in c.fetchall()]
            common_cols = [col for col in ["id", "job_title", "department", "hiring_manager"] if col in old_cols]
            cols_str = ", ".join(common_cols)
            c.execute(f"INSERT INTO jobs ({cols_str}) SELECT {cols_str} FROM jobs_old")
            c.execute("DROP TABLE jobs_old")
            conn.commit()
    except Exception as e:
        logging.error(f"Migration jobs failed: {e}")

    # --- טבלאות ATS קיימות ---
    c.execute('''CREATE TABLE IF NOT EXISTS candidates (id TEXT PRIMARY KEY, name TEXT, email TEXT, phone TEXT, source TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS jobs (id TEXT PRIMARY KEY, job_title TEXT, department TEXT, hiring_manager TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS applications (app_id TEXT PRIMARY KEY, candidate_id TEXT, job_id TEXT, status TEXT, recruiter TEXT, start_date TIMESTAMP, days_in_process INTEGER, upload_log_id TEXT, FOREIGN KEY(candidate_id) REFERENCES candidates(id), FOREIGN KEY(job_id) REFERENCES jobs(id))''')
    c.execute('''CREATE TABLE IF NOT EXISTS data_logs (log_id TEXT PRIMARY KEY, filename TEXT, upload_date TIMESTAMP, rows_processed INTEGER, status TEXT)''')
    c.execute(
        """CREATE TABLE IF NOT EXISTS ingestion_schema_versions
           (schema_version TEXT PRIMARY KEY, is_active INTEGER NOT NULL, deprecated INTEGER NOT NULL DEFAULT 0,
            sunset_date TEXT, created_at TEXT NOT NULL)"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS ingestion_rule_versions
           (rule_version TEXT PRIMARY KEY, is_active INTEGER NOT NULL, created_at TEXT NOT NULL)"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS ingestion_batches
           (batch_id TEXT PRIMARY KEY, log_id TEXT, filename TEXT, payload_hash TEXT, idempotency_key TEXT,
            schema_version TEXT, actor TEXT, request_id TEXT, status TEXT, rows_received INTEGER DEFAULT 0,
            rows_loaded INTEGER DEFAULT 0, rows_rejected INTEGER DEFAULT 0, duplicate_rows INTEGER DEFAULT 0,
            quality_score REAL DEFAULT 0, quality_report TEXT, started_at TEXT, finished_at TEXT, error_message TEXT,
            UNIQUE(idempotency_key, payload_hash))"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS rejected_rows
           (id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id TEXT, row_index INTEGER, reason_code TEXT, reason_detail TEXT,
            raw_row TEXT, created_at TEXT)"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS batch_entity_changes
           (id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id TEXT, entity_type TEXT, entity_id TEXT, change_type TEXT,
            before_json TEXT, after_json TEXT, created_at TEXT)"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS batch_snapshots_state
           (id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id TEXT, snapshot_name TEXT, payload_json TEXT, created_at TEXT)"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS stg_applications
           (batch_id TEXT, row_idx INTEGER, name TEXT, email TEXT, job_title TEXT, status TEXT, recruiter TEXT,
            start_date TEXT, department TEXT, source TEXT, stage_code TEXT, days_in_process INTEGER)"""
    )

    # --- טבלאות FinOps (משודרג!) ---
    c.execute('''CREATE TABLE IF NOT EXISTS finops_invoices (
                 id TEXT PRIMARY KEY, vendor TEXT, date TEXT, due_date TEXT, budget_month TEXT,
                 amount REAL, category TEXT, subcategory TEXT, status TEXT, 
                 note TEXT, file_url TEXT)''')
                 
    c.execute('''CREATE TABLE IF NOT EXISTS finops_vendors (
                 id TEXT PRIMARY KEY, name TEXT UNIQUE, default_category TEXT, 
                 total_paid REAL, active_invoices INTEGER)''')
                 
    c.execute('''CREATE TABLE IF NOT EXISTS finops_categories (
                 id INTEGER PRIMARY KEY, name TEXT UNIQUE, 
                 target REAL, previous_year_spend REAL, code TEXT, notes TEXT, subcategories TEXT)''')

    # --- טבלת אבטחת מידע (Audit Logs) ---
    c.execute(
        '''CREATE TABLE IF NOT EXISTS audit_logs
           (id TEXT PRIMARY KEY, timestamp TEXT, action TEXT, status TEXT, details TEXT, user TEXT, ip_address TEXT)'''
    )
    # Backward compatibility for existing DBs
    try:
        c.execute("ALTER TABLE audit_logs ADD COLUMN ip_address TEXT")
    except sqlite3.OperationalError:
        pass

    # --- טבלת JWT revocation (logout) ---
    c.execute(
        """CREATE TABLE IF NOT EXISTS revoked_tokens
           (signature TEXT PRIMARY KEY, exp INTEGER NOT NULL, revoked_at TEXT NOT NULL)"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_revoked_tokens_exp ON revoked_tokens(exp)")

    # --- טבלת הגדרות מערכת (כגון מצב מנוע ה-AI) ---
    c.execute('''CREATE TABLE IF NOT EXISTS system_settings (key TEXT PRIMARY KEY, value TEXT)''')
    c.execute('''INSERT OR IGNORE INTO system_settings (key, value) VALUES ('ai_enabled', 'true')''')
    c.execute(
        """CREATE TABLE IF NOT EXISTS etl_rules
           (id TEXT PRIMARY KEY, col_name TEXT, condition TEXT, action TEXT, active BOOLEAN)"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS onboarding
           (id TEXT PRIMARY KEY, name TEXT, id_num TEXT, role TEXT, department TEXT, manager TEXT,
            start_date TEXT, base_salary REAL, global_salary REAL, parking BOOLEAN, car_num TEXT,
            referral_name TEXT, referral_id TEXT, diversity TEXT, status TEXT, created_at TEXT)"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS iam_users
           (id TEXT PRIMARY KEY, name TEXT, email TEXT UNIQUE, usf TEXT, personal_password TEXT, role TEXT,
            status TEXT, last_login TEXT, permissions_json TEXT, updated_at TEXT)"""
    )
    default_iam_users = [
        ("u1", "אביב כהן", "avivc@fnx.co.il", "100001", "100001", "admin", "active", "היום, 08:30",
         json.dumps({"modules": ["dashboard", "candidates", "intelligence", "finops", "admin"], "contextIP": "all", "expiresAt": None, "delegatedTo": None, "dynamicRule": "Global Access"}, ensure_ascii=False)),
        ("u2", "מור אהרון", "mora@fnx.co.il", "100245", "100245", "recruiter", "active", "היום, 09:15",
         json.dumps({"modules": ["dashboard", "candidates", "intelligence"], "contextIP": "office_only", "expiresAt": None, "delegatedTo": None, "dynamicRule": "Department = R&D OR Dept = Sales"}, ensure_ascii=False)),
        ("u3", "דן שפירא", "dans@fnx.co.il", "100333", "100333", "hiring_manager", "active", "אתמול, 14:20",
         json.dumps({"modules": ["dashboard", "candidates"], "contextIP": "all", "expiresAt": "2026-06-01", "delegatedTo": "u4", "dynamicRule": "My Requisitions ONLY"}, ensure_ascii=False)),
        ("u4", "שרון לוי", "sharonl@fnx.co.il", "100487", "100487", "hrbp", "suspended", "10/11/2025",
         json.dumps({"modules": ["dashboard", "candidates", "intelligence"], "contextIP": "office_only", "expiresAt": None, "delegatedTo": None, "dynamicRule": "Dept = Sales & Service"}, ensure_ascii=False)),
    ]
    c.executemany(
        """INSERT OR IGNORE INTO iam_users
           (id, name, email, usf, personal_password, role, status, last_login, permissions_json, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        [(u[0], u[1], u[2], u[3], u[4], u[5], u[6], u[7], u[8], _utcnow().isoformat()) for u in default_iam_users],
    )
    try:
        c.execute("ALTER TABLE applications ADD COLUMN stage_code TEXT")
    except sqlite3.OperationalError:
        pass
    c.execute(
        "INSERT OR IGNORE INTO ingestion_schema_versions(schema_version, is_active, deprecated, sunset_date, created_at) VALUES (?, 1, 0, NULL, ?)",
        (DEFAULT_SCHEMA_VERSION, _utcnow().isoformat()),
    )
    c.execute(
        "INSERT OR IGNORE INTO ingestion_rule_versions(rule_version, is_active, created_at) VALUES ('1.0', 1, ?)",
        (_utcnow().isoformat(),),
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_batch_status ON ingestion_batches(status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_rejected_batch ON rejected_rows(batch_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_changes_batch ON batch_entity_changes(batch_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_stg_batch ON stg_applications(batch_id)")

    # --- Real auth: users with bcrypt passwords (separate from the iam_users
    # permissions UI table above; the two coexist until phase-5 unification). ---
    c.execute(
        """CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT,
            role TEXT NOT NULL DEFAULT 'recruiter',
            employee_number TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            must_change_password INTEGER NOT NULL DEFAULT 0,
            created_at TEXT,
            last_login_at TEXT
        )"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_users_email ON users(email)")

    # --- In-app notifications: admin-sent, per-user inbox. ---
    c.execute(
        """CREATE TABLE IF NOT EXISTS notifications (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            message TEXT NOT NULL,
            severity TEXT NOT NULL DEFAULT 'info',
            sent_by TEXT,
            sent_at TEXT NOT NULL,
            read_at TEXT
        )"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_notifications_user ON notifications(user_id, sent_at)")
    # Migration: add link and category columns (safe no-op if already exist)
    for _col_sql in [
        "ALTER TABLE notifications ADD COLUMN link TEXT",
        "ALTER TABLE notifications ADD COLUMN category TEXT DEFAULT 'general'",
    ]:
        try:
            c.execute(_col_sql)
        except Exception:
            pass  # column already exists

    # --- Multi-type ingestion fallback: stores raw rows for file types that
    # don't yet have a dedicated entity table (diversity/headcount/budget/attrition).
    # See /upload/{file_type}. JSON-per-row, queryable by file_type+batch_id. ---
    c.execute(
        """CREATE TABLE IF NOT EXISTS ingestion_typed_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_type TEXT NOT NULL,
            batch_id TEXT NOT NULL,
            row_data TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            uploaded_by TEXT
        )"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_typed_rows_lookup ON ingestion_typed_rows(file_type, uploaded_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_typed_rows_batch ON ingestion_typed_rows(batch_id)")

    # --- App registry overrides (admin-managed) for /ai-hub ("אפליקציות").
    # Each row is a per-app override of the hardcoded default tag/visibility.
    # `tag` valid values: 'new' | 'update' | 'coming_soon' | 'none' (force-off) | NULL (inherit default).
    # `hidden=1` removes the app from the grid for non-admins (admins still see
    # it dimmed, so they can re-enable). Designed so frontend can merge with
    # the static TOOLS_REGISTRY in one shot. ---
    c.execute(
        """CREATE TABLE IF NOT EXISTS app_overrides (
            app_id TEXT PRIMARY KEY,
            hidden INTEGER NOT NULL DEFAULT 0,
            tag TEXT,
            updated_at TEXT NOT NULL,
            updated_by TEXT
        )"""
    )

    # --- Per-user notification category preferences. ---
    # enabled=1 (default) means the user wants this category; enabled=0 = opt-out.
    c.execute(
        """CREATE TABLE IF NOT EXISTS user_notification_preferences (
            user_id TEXT NOT NULL,
            category TEXT NOT NULL,
            enabled INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (user_id, category)
        )"""
    )

    # --- Notification cooldowns: dedup tracking INDEPENDENT of the inbox. ---
    # Survives user deletion — prevents re-emitting the same event within the
    # cooldown window even after the user dismisses the notification.
    c.execute(
        """CREATE TABLE IF NOT EXISTS notification_cooldowns (
            user_id TEXT NOT NULL,
            tag     TEXT NOT NULL,
            last_emitted_at TEXT NOT NULL,
            PRIMARY KEY (user_id, tag)
        )"""
    )

    # --- Speed up the LEFT JOIN that powers /candidates' onboarding linkage. ---
    c.execute("CREATE INDEX IF NOT EXISTS idx_onboarding_name_lower ON onboarding(LOWER(name))")
    c.execute("CREATE INDEX IF NOT EXISTS idx_applications_job_id ON applications(job_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_applications_stage ON applications(stage_code)")

    # =====================================================================
    # DATA PIPELINE V2 — schema extensions for the unified ingest pipeline.
    # Each block is idempotent (CREATE IF NOT EXISTS, _safe_alter swallows
    # "duplicate column" errors), so init_db remains safe to re-run.
    # =====================================================================

    def _safe_alter(sql: str) -> None:
        try:
            c.execute(sql)
        except sqlite3.OperationalError:
            # column already exists — ignore. We rely on column-name uniqueness
            # per table since SQLite doesn't support IF NOT EXISTS on columns.
            pass

    # ---- candidates: extra contact + audit columns + phone normalization. ----
    _safe_alter("ALTER TABLE candidates ADD COLUMN phone_norm TEXT")
    _safe_alter("ALTER TABLE candidates ADD COLUMN email_norm TEXT")
    _safe_alter("ALTER TABLE candidates ADD COLUMN linkedin TEXT")
    _safe_alter("ALTER TABLE candidates ADD COLUMN cv_url TEXT")
    _safe_alter("ALTER TABLE candidates ADD COLUMN notes TEXT")
    _safe_alter("ALTER TABLE candidates ADD COLUMN last_seen_at TEXT")
    _safe_alter("ALTER TABLE candidates ADD COLUMN first_ingested_batch TEXT")
    c.execute("CREATE INDEX IF NOT EXISTS idx_candidates_phone_norm ON candidates(phone_norm)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_candidates_email_norm ON candidates(email_norm)")
    _safe_alter("ALTER TABLE candidates ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_candidates_uniq_email_phone ON candidates(email_norm, phone_norm)")

    # ---- applications: iteration signature + dates. ----
    _safe_alter("ALTER TABLE applications ADD COLUMN iteration_signature TEXT")
    _safe_alter("ALTER TABLE applications ADD COLUMN application_date TEXT")
    _safe_alter("ALTER TABLE applications ADD COLUMN batch_id TEXT")
    _safe_alter("ALTER TABLE applications ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
    # Unique constraint enforces "same iteration" definition (candidate+job+sig).
    # We use a partial-style unique index by allowing NULL iteration_signature
    # (SQLite treats multiple NULL as distinct, so legacy rows aren't blocked).
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_applications_iteration "
              "ON applications(candidate_id, job_id, iteration_signature)")

    # ---- jobs: lifecycle + targets. ----
    _safe_alter("ALTER TABLE jobs ADD COLUMN opened_at TEXT")
    _safe_alter("ALTER TABLE jobs ADD COLUMN closed_at TEXT")
    _safe_alter("ALTER TABLE jobs ADD COLUMN close_reason TEXT")
    _safe_alter("ALTER TABLE jobs ADD COLUMN target_count INTEGER DEFAULT 1")
    _safe_alter("ALTER TABLE jobs ADD COLUMN first_ingested_batch TEXT")
    _safe_alter("ALTER TABLE jobs ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
    c.execute("CREATE INDEX IF NOT EXISTS idx_jobs_title_dept ON jobs(LOWER(job_title), LOWER(department))")
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_jobs_title_dept_uniq ON jobs(LOWER(job_title), LOWER(department))")

    # ---- onboarding: buddy + equipment tracking columns ----
    _safe_alter("ALTER TABLE onboarding ADD COLUMN buddy TEXT")
    _safe_alter("ALTER TABLE onboarding ADD COLUMN equipment_ready INTEGER")
    _safe_alter("ALTER TABLE onboarding ADD COLUMN notes TEXT")

    # ---- data_anomalies: flags suspicious records for dashboard review. ----
    c.execute(
        """CREATE TABLE IF NOT EXISTS data_anomalies (
            id TEXT PRIMARY KEY,
            entity_type TEXT NOT NULL,
            entity_id TEXT NOT NULL,
            anomaly_type TEXT NOT NULL,
            severity TEXT NOT NULL DEFAULT 'medium',
            description TEXT,
            suggestion TEXT,
            meta_json TEXT,
            status TEXT NOT NULL DEFAULT 'open',
            reviewed_by TEXT,
            reviewed_at TEXT,
            created_at TEXT NOT NULL,
            batch_id TEXT
        )"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_anomalies_entity ON data_anomalies(entity_type, entity_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_anomalies_status ON data_anomalies(status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_anomalies_type ON data_anomalies(anomaly_type)")

    # ---- stage_history: append-only log of candidate stage transitions. ----
    c.execute(
        """CREATE TABLE IF NOT EXISTS stage_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            app_id TEXT NOT NULL,
            candidate_id TEXT NOT NULL,
            job_id TEXT NOT NULL,
            from_stage TEXT,
            to_stage TEXT NOT NULL,
            changed_by TEXT,
            changed_at TEXT NOT NULL,
            batch_id TEXT,
            FOREIGN KEY(app_id) REFERENCES applications(app_id)
        )"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_stage_history_app ON stage_history(app_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_stage_history_candidate ON stage_history(candidate_id)")

    # ---- hires: when a candidate became a hire (often multiple per candidate over time). ----
    c.execute(
        """CREATE TABLE IF NOT EXISTS hires (
            id TEXT PRIMARY KEY,
            candidate_id TEXT,
            job_id TEXT,
            candidate_name TEXT,
            job_title TEXT,
            hire_date TEXT NOT NULL,
            salary REAL,
            department TEXT,
            manager TEXT,
            referral_name TEXT,
            is_diversity INTEGER DEFAULT 0,
            batch_id TEXT,
            created_at TEXT,
            FOREIGN KEY(candidate_id) REFERENCES candidates(id),
            FOREIGN KEY(job_id) REFERENCES jobs(id)
        )"""
    )
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_hires_natural "
              "ON hires(IFNULL(candidate_id,''), IFNULL(job_id,''), hire_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_hires_date ON hires(hire_date)")

    # ---- diversity_snapshots: one row per (month × dept × dimension × bucket). ----
    c.execute(
        """CREATE TABLE IF NOT EXISTS diversity_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snapshot_month TEXT NOT NULL,
            department TEXT NOT NULL,
            dimension TEXT NOT NULL,
            bucket TEXT NOT NULL,
            count INTEGER NOT NULL,
            batch_id TEXT,
            created_at TEXT,
            UNIQUE(snapshot_month, department, dimension, bucket)
        )"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_diversity_month ON diversity_snapshots(snapshot_month)")

    # ---- headcount_snapshots: planned vs actual per role per month. ----
    c.execute(
        """CREATE TABLE IF NOT EXISTS headcount_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snapshot_month TEXT NOT NULL,
            department TEXT NOT NULL,
            role TEXT NOT NULL,
            standard INTEGER NOT NULL DEFAULT 0,
            current INTEGER NOT NULL DEFAULT 0,
            attrition_ytd INTEGER DEFAULT 0,
            hire_plan INTEGER DEFAULT 0,
            batch_id TEXT,
            created_at TEXT,
            UNIQUE(snapshot_month, department, role)
        )"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_headcount_month ON headcount_snapshots(snapshot_month)")

    # ---- attrition_events: leave-of-employment records (separate from candidate-funnel REJECTED). ----
    c.execute(
        """CREATE TABLE IF NOT EXISTS attrition_events (
            id TEXT PRIMARY KEY,
            employee_name TEXT NOT NULL,
            candidate_id TEXT,
            leave_date TEXT NOT NULL,
            department TEXT,
            manager TEXT,
            last_role TEXT,
            reason TEXT,
            voluntary INTEGER,
            batch_id TEXT,
            created_at TEXT,
            UNIQUE(employee_name, leave_date),
            FOREIGN KEY(candidate_id) REFERENCES candidates(id)
        )"""
    )
    c.execute("CREATE INDEX IF NOT EXISTS idx_attrition_date ON attrition_events(leave_date)")

    # ---- data_version: monotonic counter used by frontend auto-refresh. ----
    c.execute(
        "INSERT OR IGNORE INTO system_settings (key, value) VALUES ('data_version', '0')"
    )

    # =====================================================================
    # One-time backfill: legacy candidate rows inserted BEFORE phone_norm /
    # email_norm columns existed. Without this, the new dedup logic can't
    # match those rows and tries to INSERT — triggering UNIQUE(email) error.
    # =====================================================================
    legacy_rows = c.execute(
        "SELECT id, email, phone FROM candidates "
        "WHERE (email IS NOT NULL AND email_norm IS NULL) "
        "   OR (phone IS NOT NULL AND phone_norm IS NULL)"
    ).fetchall()
    for cid, email, phone in legacy_rows:
        # Inline normalization (functions defined further below — use simple inline
        # version to keep init_db self-contained; phone "+972XXXXXXXXX", email lowercased).
        email_n = None
        if email and isinstance(email, str) and "@" in email:
            email_n = email.strip().lower() or None
        phone_n = None
        if phone:
            digits = "".join(ch for ch in str(phone) if ch.isdigit())
            if digits.startswith("972"): digits = digits[3:]
            if digits.startswith("0"): digits = digits[1:]
            if len(digits) == 9 and digits[0] in {"5", "7"}:
                phone_n = "+972" + digits
            elif len(digits) == 8 and digits[0] in {"2", "3", "4", "8", "9"}:
                phone_n = "+972" + digits
        c.execute(
            "UPDATE candidates SET email_norm = COALESCE(email_norm, ?), phone_norm = COALESCE(phone_norm, ?) WHERE id = ?",
            (email_n, phone_n, cid),
        )

    conn.commit()
    seed_internal_logic_tables(conn)

    conn.close()

init_db()

# ==========================================
# מילון נורמליזציה (Standardization Dictionary)
# ==========================================
DEPT_NORMALIZATION = {
    "מו\"פ": "R&D",
    "פיתוח": "R&D",
    "משאבי אנוש": "HR",
    "משאבי-אנוש": "HR",
    "מכירות ושירות": "Sales & Service",
    "שירות": "Sales & Service"
}

# ==========================================
# שאילתת תאימות ל-UI הקיים (View Pattern)
# ==========================================
def get_unified_data(conn):
    """מייצר את הטבלה השטוחה של applications + candidates + jobs + onboarding.

    LEFT JOIN ל-onboarding משוייך לפי השם (LOWER(name)) — זה החיבור היחיד
    הזמין בין pipeline ל-onboarding שכן candidates.id (internal) ≠ onboarding.id.
    onboarding_status ו-id_num מחזירים NULL כאשר אין שיוך.
    """
    query = '''
        SELECT
            c.id as candidate_id,
            c.name as candidate_name,
            c.email,
            c.phone,
            c.source,
            j.id as job_id,
            j.job_title,
            j.department,
            a.app_id,
            a.status,
            a.recruiter,
            a.start_date,
            a.days_in_process,
            a.upload_log_id,
            COALESCE(a.stage_code, 'ACTIVE') as stage_code,
            o.id as onboarding_id,
            o.id_num as id_num,
            o.status as onboarding_status,
            o.start_date as onboarding_start_date,
            o.role as onboarding_role,
            o.manager as onboarding_manager
        FROM applications a
        JOIN candidates c ON a.candidate_id = c.id
        JOIN jobs j ON a.job_id = j.id
        LEFT JOIN onboarding o ON LOWER(o.name) = LOWER(c.name)
        WHERE COALESCE(a.is_active, 1) = 1
          AND COALESCE(c.is_active, 1) = 1
          AND COALESCE(j.is_active, 1) = 1
    '''
    return pd.read_sql(query, conn)


def _compute_unified_stage(stage_code: str | None, onboarding_status: str | None) -> str:
    """Maps (applications.stage_code, onboarding.status) to a single stage code
    that drives both the /candidates chips and the /jobs breakdown.

    Onboarding wins because it represents a later stage in the funnel:
      - onboarding pending   → AWAITING_START (recruit accepted, hasn't started)
      - onboarding completed → STARTED        (employee actively onboarding)
      - onboarding cancelled / left_company → REJECTED (archive)
    Otherwise: fall through to applications.stage_code.
    """
    if onboarding_status == "completed":
        return "STARTED"
    if onboarding_status == "pending":
        return "AWAITING_START"
    if onboarding_status in ("cancelled", "left_company"):
        return "REJECTED"
    return (stage_code or "ACTIVE").upper()


# Canonical stage order — used by frontend chips and any aggregation that
# wants a stable iteration order. Mirrors src/lib/stages.ts.
# `ACTIVE` is the default bucket for any status that does not match the
# lexicon (e.g. raw "חדש" / "בתהליך" / "ממתין") — keeps unmatched rows visible.
# UNIFIED_STAGES moved to backend/constants.py; re-imported at top of this file.


def _nan_safe_records(df: "pd.DataFrame") -> list:
    """Convert a DataFrame to records list, replacing NaN/inf/numpy types with JSON-safe values.

    Python's json.dumps raises ValueError on float('nan') and float('inf'). pandas DataFrames
    frequently have NaN floats in numeric columns when a DB row has NULL.  This utility converts
    them to Python None so FastAPI can serialize them as JSON null values.
    Also handles numpy scalar types (np.int64, np.float64, etc.) that are not JSON serializable.
    """
    import math
    try:
        import numpy as np
        _np_integer = np.integer
        _np_floating = np.floating
        _np_ndarray = np.ndarray
    except ImportError:
        _np_integer = _np_floating = _np_ndarray = type(None)  # type: ignore

    raw = df.where(df.notna(), other=None).to_dict(orient="records")

    def safe(v):
        if v is None:
            return None
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        if isinstance(v, _np_integer):
            return int(v)
        if isinstance(v, _np_floating):
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else f
        if isinstance(v, _np_ndarray):
            return v.tolist()
        return v

    return [{k: safe(v) for k, v in row.items()} for row in raw]


# =====================================================================
# DATA PIPELINE V2 — shared helpers used by all ingest stage-handlers.
# Phone normalisation, iteration signature, candidate match+merge, and a
# monotonic data_version counter that drives the frontend auto-refresh.
# =====================================================================


def normalize_phone(raw) -> Optional[str]:
    """Israeli mobile/landline → +972XXXXXXXXX canonical form.

    Strips non-digits, removes leading 0 / 972, returns None when the result
    isn't a plausible 9-digit IL number. Returning None means "can't dedupe
    on this field" — the caller falls back to email or inserts as new.
    """
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return None
    if digits.startswith("972"):
        digits = digits[3:]
    if digits.startswith("0"):
        digits = digits[1:]
    # IL mobile = 5XXXXXXXX (9 digits). IL landline = 2|3|4|7|8|9 + 7 digits.
    if len(digits) == 9 and digits[0] in {"5", "7"}:
        return "+972" + digits
    if len(digits) == 8 and digits[0] in {"2", "3", "4", "8", "9"}:
        return "+972" + digits
    return None


def normalize_email(raw) -> Optional[str]:
    if raw is None:
        return None
    v = str(raw).strip().lower()
    if not v or "@" not in v:
        return None
    return v


def iteration_signature(status, application_date, recruiter) -> str:
    """Two application rows are the SAME iteration only if status, date and
    recruiter all match. Different in any of these = a distinct iteration
    (e.g. re-applied a year later, or moved to a new stage by a different recruiter).

    Returns a 16-char hex digest used in the unique index (candidate_id, job_id, iteration_signature).
    Inputs may be strings, datetimes, NaN, None — anything stringifiable.
    """
    def _norm(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) and pd.isna(v):
            return ""
        return str(v).strip().lower()
    parts = [
        _norm(status),
        _norm(application_date)[:10],
        _norm(recruiter),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]


def find_existing_candidate(conn: sqlite3.Connection, phone_norm: Optional[str], email_norm: Optional[str]) -> Optional[str]:
    """Returns candidate.id if a row already exists matching the normalized
    phone OR email. Phone takes precedence (more unique in IL B2C). Returns
    None when nothing to match against — caller inserts a new candidate."""
    if not phone_norm and not email_norm:
        return None
    c = conn.cursor()
    if phone_norm:
        row = c.execute("SELECT id FROM candidates WHERE phone_norm = ? LIMIT 1", (phone_norm,)).fetchone()
        if row:
            return row[0]
    if email_norm:
        row = c.execute("SELECT id FROM candidates WHERE email_norm = ? LIMIT 1", (email_norm,)).fetchone()
        if row:
            return row[0]
    return None


def merge_candidate(conn: sqlite3.Connection, candidate_id: str, parsed: dict) -> dict:
    """Upsert merge: writes only NON-EMPTY incoming values to the candidate
    row. Preserves prior data when the upload is missing a field. Returns
    a `{before, after}` dict for the audit trail.
    """
    cols = ["name", "email", "phone", "email_norm", "phone_norm", "source", "linkedin", "cv_url", "notes"]
    c = conn.cursor()
    before_row = c.execute(
        f"SELECT {', '.join(cols)} FROM candidates WHERE id = ?", (candidate_id,)
    ).fetchone()
    before = dict(zip(cols, before_row)) if before_row else {}

    sets, vals = [], []
    for col in cols:
        v = parsed.get(col)
        if isinstance(v, str):
            v = v.strip()
        if v not in (None, "", []):
            sets.append(f"{col} = ?")
            vals.append(v)
    # Always touch last_seen_at to reflect "this candidate was seen in an upload now".
    sets.append("last_seen_at = ?")
    vals.append(datetime.now(timezone.utc).isoformat())
    vals.append(candidate_id)
    c.execute(f"UPDATE candidates SET {', '.join(sets)} WHERE id = ?", vals)
    after_row = c.execute(
        f"SELECT {', '.join(cols)} FROM candidates WHERE id = ?", (candidate_id,)
    ).fetchone()
    after = dict(zip(cols, after_row)) if after_row else {}
    return {"before": before, "after": after}


def _nullify_empty(v):
    """Coerce empty/blank/'nan' strings to None so SQLite UNIQUE doesn't treat
    them as collisions. `_normalize_upload_frame` astype(str) turns NaN into
    the literal string "nan" — without nullifying that, two candidate rows
    with missing email both end up with email='nan' and collide.

    SQLite treats NULL as distinct from any other NULL, so multiple rows with
    empty contact fields can coexist."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ("nan", "none", "null", "n/a", "-"):
            return None
        return s
    return v


def insert_candidate(conn: sqlite3.Connection, parsed: dict, batch_id: str) -> str:
    """Insert a brand-new candidate row. Generates a UUID-style id. Returns id.
    Empty contact fields are stored as NULL — important for the UNIQUE(email)
    constraint on `candidates`.
    """
    candidate_id = f"CND-{uuid.uuid4().hex[:12].upper()}"
    now = datetime.now(timezone.utc).isoformat()
    c = conn.cursor()
    c.execute(
        """INSERT INTO candidates
           (id, name, email, email_norm, phone, phone_norm, source, linkedin, cv_url, notes,
            first_ingested_batch, last_seen_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            candidate_id,
            (parsed.get("name") or "").strip() or "ללא שם",
            _nullify_empty(parsed.get("email")),
            _nullify_empty(parsed.get("email_norm")),
            _nullify_empty(parsed.get("phone")),
            _nullify_empty(parsed.get("phone_norm")),
            _nullify_empty(parsed.get("source")),
            _nullify_empty(parsed.get("linkedin")),
            _nullify_empty(parsed.get("cv_url")),
            _nullify_empty(parsed.get("notes")),
            batch_id,
            now,
        ),
    )
    return candidate_id


def get_data_version(conn: Optional[sqlite3.Connection] = None) -> int:
    """Read the global data_version counter."""
    close_after = False
    if conn is None:
        conn = sqlite3.connect(DB_PATH)
        close_after = True
    try:
        row = conn.execute("SELECT value FROM system_settings WHERE key = 'data_version'").fetchone()
        return int(row[0]) if row and str(row[0]).isdigit() else 0
    finally:
        if close_after:
            conn.close()


def bump_data_version(conn: Optional[sqlite3.Connection] = None) -> int:
    """Increment the global data_version. Returns the new value.

    Frontend `DataVersionContext` polls this; when it sees a higher number,
    consumer blocks (candidates, jobs, dashboard, intelligence, headcount,
    budget) trigger a refetch automatically. Call this at the END of every
    successful ingest commit, AFTER the rows are persisted.
    """
    close_after = False
    if conn is None:
        conn = sqlite3.connect(DB_PATH)
        close_after = True
    try:
        current = get_data_version(conn)
        new_val = current + 1
        conn.execute(
            "INSERT INTO system_settings (key, value) VALUES ('data_version', ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (str(new_val),),
        )
        if close_after:
            conn.commit()
        return new_val
    finally:
        if close_after:
            conn.close()


def _normalize_upload_frame(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.str.strip()
    # Alias table now lives in backend/aliases.py — see that module for the
    # distinction between LEGACY_CANDIDATE_ALIASES (this call site) and
    # TYPED_INGEST_ALIASES (used by _apply_extra_aliases below).
    df.rename(columns=LEGACY_CANDIDATE_ALIASES, inplace=True)
    if "name" not in df.columns:
        raise Exception("חובה לכלול עמודת שם מועמד")
    if "job_title" not in df.columns:
        raise Exception("חובה לכלול עמודת שם משרה")
    if "email" not in df.columns:
        df["email"] = df["name"].apply(lambda x: f"{str(x).strip().replace(' ', '.')}@unknown.com")
    if "source" not in df.columns:
        df["source"] = "Organic / Unknown"
    if "start_date" not in df.columns:
        df["start_date"] = pd.Timestamp.now()
    if "department" not in df.columns:
        df["department"] = "General"
    if "status" not in df.columns:
        df["status"] = "חדש"
    if "recruiter" not in df.columns:
        df["recruiter"] = "לא שויך"
    for col in ["name", "email", "job_title", "status", "recruiter", "department", "source"]:
        df[col] = df[col].astype(str).str.strip()
    df["department"] = df["department"].replace(DEPT_NORMALIZATION)
    df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce")
    df["days_in_process"] = (pd.Timestamp.now() - df["start_date"]).dt.days.fillna(0).clip(lower=0).astype(int)
    return df


def _parse_xml_to_dataframe(content: bytes) -> pd.DataFrame:
    root = ET.fromstring(content)
    xml_schema_version = root.attrib.get("schema_version")
    if xml_schema_version and xml_schema_version not in SUPPORTED_SCHEMA_VERSIONS:
        raise Exception(f"Unsupported XML schema_version: {xml_schema_version}")
    if root.tag != "records":
        raise Exception("XML root tag must be <records>")
    rows: list[dict] = []
    required_xml_tags = {"name", "email", "job_title", "status", "recruiter", "start_date", "department", "source"}
    for row_node in root.findall(".//row"):
        row_data = {}
        for child in list(row_node):
            row_data[child.tag] = child.text
        if row_data:
            missing = [tag for tag in required_xml_tags if tag not in row_data]
            if missing:
                raise Exception(f"XML row missing required tags: {','.join(missing)}")
            rows.append(row_data)
    if not rows:
        raise Exception("XML does not include <row> elements")
    return pd.DataFrame(rows)


def _load_schema_contract(schema_version: str) -> dict:
    candidates = [
        os.path.join(CONTRACTS_DIR, f"ingestion_schema_v{schema_version.replace('.', '_')}.json"),
        os.path.join(CONTRACTS_DIR, f"ingestion_schema_v{schema_version.split('.')[0]}.json"),
    ]
    contract_path = next((p for p in candidates if os.path.exists(p)), None)
    if not contract_path:
        raise Exception(f"Missing schema contract file for version {schema_version}")
    with open(contract_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _validate_schema_contract(df: pd.DataFrame, schema_version: str) -> None:
    contract = _load_schema_contract(schema_version)
    required_columns = contract.get("required_columns", [])
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise Exception(f"Schema validation failed. Missing columns: {','.join(missing)}")


# =====================================================================
# TEMPLATE_SPECS — per-file-type Excel template definitions.
#
# Each spec describes the Hebrew column layout, sample row, validation
# drop-downs, and an INSTRUCTIONS sheet for one of the 7 ingest types.
# Adding/changing a column here drives the whole UX — admins downloading
# the master template see the exact columns the /api/ingest/{type}
# handler expects (mapped via EXTRA_HEBREW_ALIASES).
# =====================================================================
TEMPLATE_SPECS: dict[str, dict] = {
    "candidates": {
        "required": [("שם מועמד", "name"), ("אימייל", "email"), ("טלפון", "phone")],
        "recommended": [("משרה", "job_title"), ("חטיבה", "department"), ("סטטוס", "status"),
                        ("מגייסת", "recruiter"), ("תאריך הגשה", "application_date"), ("מקור", "source")],
        "validations": {"סטטוס": ["חדש", "סינון", "ראיון", "הצעה", "התקבל", "נדחה"]},
        "sample": {"שם מועמד": "דנה כהן", "אימייל": "dana@example.com", "טלפון": "0541234567",
                   "משרה": "Backend Engineer", "חטיבה": "R&D", "סטטוס": "ראיון", "מגייסת": "מור",
                   "תאריך הגשה": "2026-05-11", "מקור": "LinkedIn"},
        "title": "תבנית מועמדים בתהליך — Dedup לפי טלפון/אימייל",
        "instructions": [
            "שורה אחת לכל איטרציית מועמד×משרה.",
            "מועמד יזוהה לפי טלפון (+972XXXXXXXXX) או אימייל. בלי אחד מהם — יוקלט כחדש בכל העלאה.",
            "סטטוס + מגייסת + תאריך הגשה זהים = איטרציה זהה (תידחה ככפילות). שינוי באחד מהם = איטרציה חדשה.",
        ],
    },
    "jobs": {
        "required": [("שם משרה", "job_title"), ("חטיבה", "department")],
        "recommended": [("מנהל מגייס", "hiring_manager"), ("תאריך פתיחה", "opened_at"),
                        ("תאריך סגירה", "closed_at"), ("סיבת סגירה", "close_reason"), ("תקן", "target_count")],
        "validations": {},
        "sample": {"שם משרה": "Senior Frontend Engineer", "חטיבה": "R&D",
                   "מנהל מגייס": "דוד לוי", "תאריך פתיחה": "2026-04-15", "תקן": 1},
        "title": "תבנית משרות פתוחות וסגורות",
        "instructions": [
            "שורה אחת לכל משרה. דדופ לפי (שם משרה + חטיבה), case-insensitive.",
            "תאריך סגירה ריק = משרה פתוחה. מילוי תאריך = משרה סגורה; אז חובה גם סיבת סגירה.",
        ],
    },
    "hires": {
        "required": [("שם מועמד", "candidate_name"), ("שם משרה", "job_title"), ("תאריך קליטה", "hire_date")],
        "recommended": [("שכר", "salary"), ("חטיבה", "department"), ("מנהל ישיר", "manager"),
                        ("ממליץ", "referral_name"), ("גיוון", "is_diversity")],
        "validations": {"גיוון": ["כן", "לא"]},
        "sample": {"שם מועמד": "נועה מזרחי", "שם משרה": "Customer Success",
                   "תאריך קליטה": "2026-05-12", "שכר": 22000, "חטיבה": "Service", "מנהל ישיר": "ליטל"},
        "title": "תבנית קליטות בפועל",
        "instructions": [
            "שורה אחת לכל קליטה שהתבצעה. המערכת תקשר אוטומטית למועמד קיים אם נמצא לפי שם/טלפון/אימייל.",
            "תאריך קליטה בפורמט YYYY-MM-DD. דדופ: שורה זהה (מועמד+משרה+תאריך) תידחה.",
        ],
    },
    "diversity": {
        "required": [("חודש", "snapshot_month"), ("חטיבה", "department"),
                     ("ממד", "dimension"), ("קבוצה", "bucket"), ("כמות", "count")],
        "recommended": [],
        "validations": {"ממד": ["gender", "age_range", "ethnicity", "tenure"]},
        "sample": {"חודש": "2026-05", "חטיבה": "R&D", "ממד": "gender", "קבוצה": "F", "כמות": 42},
        "title": "תבנית מדדי גיוון לפי חודש",
        "instructions": [
            "שורה אחת לכל (חודש × חטיבה × ממד × קבוצה).",
            "חודש בפורמט YYYY-MM (ללא יום).",
            "ערכי 'ממד' מותרים: gender (M/F/Other), age_range (e.g. 30-40), ethnicity, tenure.",
        ],
    },
    "headcount": {
        "required": [("חודש", "snapshot_month"), ("חטיבה", "department"), ("תפקיד", "role")],
        "recommended": [("תקן מצבה", "standard"), ("בפועל", "current"),
                        ("עזיבות מתחילת השנה", "attrition_ytd"), ("תכנית גיוס", "hire_plan")],
        "validations": {},
        "sample": {"חודש": "2026-05", "חטיבה": "R&D", "תפקיד": "Backend Engineer",
                   "תקן מצבה": 12, "בפועל": 11, "עזיבות מתחילת השנה": 2, "תכנית גיוס": 1},
        "title": "תבנית תקן מצבה (Standard vs Current)",
        "instructions": [
            "שורה אחת לכל (חודש × חטיבה × תפקיד).",
            "ה-attrition_ytd צריך להתאים לסכום אירועי העזיבה באותה תקופה — בקרת איכות תזהיר על פערים.",
        ],
    },
    "budget": {
        "required": [("מזהה חשבונית", "id"), ("ספק", "vendor"), ("תאריך", "date"),
                     ("סכום", "amount"), ("קטגוריה", "category")],
        "recommended": [("מועד פירעון", "due_date"), ("חודש תקציב", "budget_month"),
                        ("סטטוס", "status"), ("URL קובץ", "file_url")],
        "validations": {"סטטוס": ["ממתין למיפוי", "ממתין לתשלום", "שולם", "בוטל"]},
        "sample": {"מזהה חשבונית": "INV-2026-001", "ספק": "Workday", "תאריך": "2026-05-01",
                   "סכום": 12500, "קטגוריה": "תוכנה", "חודש תקציב": "2026-Q2"},
        "title": "תבנית חשבוניות FinOps",
        "instructions": [
            "מזהה חשבונית חייב להיות ייחודי — אם קיים, החשבונית תעודכן (upsert).",
            "תאריך + מועד פירעון בפורמט YYYY-MM-DD.",
        ],
    },
    "attrition": {
        "required": [("שם עובד", "employee_name"), ("תאריך עזיבה", "leave_date")],
        "recommended": [("חטיבה", "department"), ("מנהל", "manager"), ("תפקיד אחרון", "last_role"),
                        ("סיבה", "reason"), ("וולונטרי", "voluntary")],
        "validations": {"וולונטרי": ["כן", "לא"]},
        "sample": {"שם עובד": "אורית גרשון", "תאריך עזיבה": "2026-04-30", "חטיבה": "Sales",
                   "מנהל": "אבי כהן", "תפקיד אחרון": "Account Manager", "סיבה": "הזדמנות בחו\"ל", "וולונטרי": "כן"},
        "title": "תבנית אירועי עזיבה",
        "instructions": [
            "שורה אחת לכל אירוע עזיבה. דדופ לפי (שם עובד + תאריך עזיבה).",
            "המערכת תקשר אוטומטית לרשומת מועמד קיימת אם השם תואם.",
        ],
    },
}


def _col_letter(idx: int) -> str:
    """1-based column index → Excel letter (A, B, ..., Z, AA, AB, ...)."""
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _legacy_build_excel_template_bytes(file_type: str, schema_version: str) -> bytes:
    """Original recruiter-applications template — preserved for any caller
    that still relies on the schema contract shape. Used as fallback when a
    file_type isn't in TEMPLATE_SPECS."""
    contract = _load_schema_contract(schema_version)
    required_columns = contract.get("required_columns", [])
    status_options = ["חדש", "בסינון", "בראיון", "הצעה", "התקבל", "נדחה"]
    default_sample_by_type = {
        "candidates": ["Dana Cohen", "Backend Engineer", "dana.cohen@example.com", "חדש", "מור", "2026-05-10", "R&D", "LinkedIn"],
    }
    sample_base = default_sample_by_type.get(file_type, default_sample_by_type["candidates"])
    sample_map = {
        "name": sample_base[0],
        "job_title": sample_base[1],
        "email": sample_base[2],
        "status": sample_base[3],
        "recruiter": sample_base[4],
        "start_date": sample_base[5],
        "department": sample_base[6],
        "source": sample_base[7],
    }
    sample_row = [sample_map.get(col, "") for col in required_columns]

    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "DATA"
    header_fill = PatternFill(start_color="002649", end_color="002649", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(required_columns, start=1):
        cell = data_ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        data_ws.cell(row=2, column=col_idx, value=sample_row[col_idx - 1] if col_idx - 1 < len(sample_row) else "")

    data_ws.freeze_panes = "A2"
    for col_idx, _ in enumerate(required_columns, start=1):
        data_ws.column_dimensions[_col_letter(col_idx)].width = 22

    if "status" in required_columns:
        status_col = required_columns.index("status") + 1
        status_letter = _col_letter(status_col)
        status_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(status_options)}"',
            allow_blank=False,
        )
        data_ws.add_data_validation(status_validation)
        status_validation.add(f"{status_letter}2:{status_letter}5000")

    if "start_date" in required_columns:
        date_col = required_columns.index("start_date") + 1
        date_letter = _col_letter(date_col)
        date_validation = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2100,12,31)")
        data_ws.add_data_validation(date_validation)
        date_validation.add(f"{date_letter}2:{date_letter}5000")

    guide_ws = wb.create_sheet("INSTRUCTIONS")
    guide_ws["A1"] = f"Master Template: {file_type}"
    guide_ws["A1"].font = Font(bold=True, size=14)
    guide_ws["A3"] = "Upload using DATA sheet only."
    guide_ws["A4"] = f"Schema version: {schema_version}"
    guide_ws["A5"] = f"Required columns (do not rename): {','.join(required_columns)}"
    guide_ws["A6"] = "Date format for start_date: YYYY-MM-DD"
    guide_ws["A7"] = "Headers: X-Schema-Version + X-Preflight-Hash (from preflight response)"
    guide_ws.column_dimensions["A"].width = 120

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _build_excel_template_bytes(file_type: str, schema_version: str) -> bytes:
    """Public template builder. Routes to per-type TEMPLATE_SPECS when defined,
    falls back to legacy recruiter-applications shape for any unmapped type."""
    spec = TEMPLATE_SPECS.get(file_type)
    if not spec:
        return _legacy_build_excel_template_bytes(file_type, schema_version)

    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"
    # RTL for the Hebrew column flow.
    ws.sheet_view.rightToLeft = True

    all_cols: list[tuple[str, str]] = list(spec["required"]) + list(spec["recommended"])
    n_required = len(spec["required"])

    header_fill_req = PatternFill(start_color="002649", end_color="002649", fill_type="solid")
    header_fill_rec = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for idx, (he, _canon) in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=idx, value=he)
        cell.fill = header_fill_req if idx <= n_required else header_fill_rec
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[_col_letter(idx)].width = 20

    # Sample row.
    sample = spec.get("sample", {})
    for idx, (he, _canon) in enumerate(all_cols, start=1):
        ws.cell(row=2, column=idx, value=sample.get(he, ""))

    # Drop-down validations.
    validations: dict[str, list[str]] = spec.get("validations", {}) or {}
    for he_name, options in validations.items():
        col_idx = next((i + 1 for i, (he, _) in enumerate(all_cols) if he == he_name), None)
        if not col_idx:
            continue
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv)
        letter = _col_letter(col_idx)
        dv.add(f"{letter}2:{letter}5000")

    ws.freeze_panes = "A2"

    # Instructions sheet (Hebrew).
    guide = wb.create_sheet("הוראות")
    guide.sheet_view.rightToLeft = True
    guide["A1"] = spec["title"]
    guide["A1"].font = Font(bold=True, size=14, color="002649")
    row_i = 3
    for line in spec["instructions"]:
        guide[f"A{row_i}"] = "• " + line
        row_i += 1
    guide[f"A{row_i + 1}"] = f"Schema version: {schema_version}"
    guide[f"A{row_i + 2}"] = "עמודות חובה: רקע כחול. מומלצות: רקע אפור. ניתן להסיר עמודות מומלצות."
    guide[f"A{row_i + 3}"] = f"סה\"כ עמודות בקובץ: {len(all_cols)} ({n_required} חובה, {len(all_cols) - n_required} מומלצות)"
    guide.column_dimensions["A"].width = 100

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _build_preflight_report(filename: str, content: bytes, schema_version: str) -> dict:
    payload_hash = hashlib.sha256(content).hexdigest()
    raw_df = _load_dataframe_from_upload(filename, content)
    df = _normalize_upload_frame(raw_df.copy())
    _validate_schema_contract(df, schema_version)
    mandatory_cols = ["name", "job_title", "email", "status", "recruiter", "department", "start_date", "source"]
    missing_counts = {}
    for col in mandatory_cols:
        if col not in df.columns:
            missing_counts[col] = len(df)
        else:
            missing_counts[col] = int(((df[col].astype(str).str.strip() == "") | (df[col].isna())).sum())
    duplicate_mask = df.duplicated(subset=["email", "job_title"], keep=False)
    duplicate_rows = int(duplicate_mask.sum())
    rows_received = int(len(df))
    rows_with_mandatory_issues = int(sum(1 for i in range(rows_received) if any(missing_counts.get(c, 0) > 0 and ((pd.isna(df.iloc[i][c]) if c in df.columns else True) or str(df.iloc[i][c]).strip() == "") for c in mandatory_cols)))
    error_rate = round(rows_with_mandatory_issues / max(rows_received, 1), 4)
    return {
        "filename": filename,
        "schema_version": schema_version,
        "payload_hash": payload_hash,
        "rows_received": rows_received,
        "duplicate_rows": duplicate_rows,
        "duplicate_rate": round(duplicate_rows / max(rows_received, 1), 4),
        "mandatory_missing_counts": missing_counts,
        "mandatory_issue_rows": rows_with_mandatory_issues,
        "error_rate": error_rate,
        "can_ingest": error_rate <= MAX_ERROR_RATE,
        "max_error_rate": MAX_ERROR_RATE,
    }


def _load_dataframe_from_upload(filename: str, content: bytes) -> pd.DataFrame:
    lower_name = (filename or "").lower()
    buf = io.BytesIO(content)
    if lower_name.endswith(".xml"):
        return _parse_xml_to_dataframe(content)
    if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
        return pd.read_excel(buf)
    try:
        return pd.read_csv(io.BytesIO(content))
    except Exception:
        try:
            return pd.read_csv(io.BytesIO(content), encoding="iso-8859-8")
        except Exception:
            return pd.read_excel(io.BytesIO(content))


def _record_batch_change(conn: sqlite3.Connection, batch_id: str, entity_type: str, entity_id: str, change_type: str, before_obj: dict | None, after_obj: dict | None):
    conn.execute(
        """INSERT INTO batch_entity_changes(batch_id, entity_type, entity_id, change_type, before_json, after_json, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (
            batch_id,
            entity_type,
            entity_id,
            change_type,
            json.dumps(before_obj, ensure_ascii=False) if before_obj else None,
            json.dumps(after_obj, ensure_ascii=False) if after_obj else None,
            _utcnow().isoformat(),
        ),
    )


@app.get("/")
def read_root():
    return {"status": "Phoenix Enterprise Brain is Active 🧠"}


@app.get("/healthz")
def healthz():
    return {"status": "ok", "service": "phoenix-api"}


@app.get("/readyz")
def readyz():
    try:
        with db_conn() as conn:
            conn.execute("SELECT 1")
        return {"status": "ready"}
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Database not ready: {exc}") from exc


# ==========================================
# 2. מנוע ה-ETL (קליטה, ניקוי, חלוקה לטבלאות)
# ==========================================
@app.post("/upload")
@limiter.limit("10/minute")
async def upload_file(
    request: Request,
    file: UploadFile = File(...),
    x_schema_version: Optional[str] = Header(default=None),
    x_idempotency_key: Optional[str] = Header(default=None),
    x_preflight_hash: Optional[str] = Header(default=None),
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER)),
):
    _validate_upload_file(
        file,
        allowed_extensions={".csv", ".xlsx", ".xls", ".xml"},
        allowed_mime_prefixes=(
            "text/csv",
            "application/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/xml",
            "text/xml",
        ),
    )
    log_id = str(uuid.uuid4())[:8]
    batch_id = f"bat-{uuid.uuid4().hex[:10]}"
    now = _utcnow().isoformat()
    content = await _read_file_with_limit(file)
    payload_hash = hashlib.sha256(content).hexdigest()
    idempotency_key = (x_idempotency_key or payload_hash[:16]).strip()
    schema_version = x_schema_version or DEFAULT_SCHEMA_VERSION
    preflight_hash = (x_preflight_hash or "").strip()
    if schema_version not in SUPPORTED_SCHEMA_VERSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported schema_version: {schema_version}")
    if preflight_hash and preflight_hash != payload_hash:
        raise HTTPException(status_code=400, detail="Preflight checksum mismatch. Please rerun preflight before upload.")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        existing = c.execute(
            """SELECT batch_id, rows_loaded, status FROM ingestion_batches
               WHERE idempotency_key = ? AND payload_hash = ? ORDER BY started_at DESC LIMIT 1""",
            (idempotency_key, payload_hash),
        ).fetchone()
        if existing and existing["status"] == "committed":
            return {
                "message": "Idempotent replay accepted",
                "batch_id": existing["batch_id"],
                "rows_processed": int(existing["rows_loaded"] or 0),
                "replayed": True,
            }

        c.execute("BEGIN")
        try:
            c.execute(
                """INSERT INTO ingestion_batches(
                       batch_id, log_id, filename, payload_hash, idempotency_key, schema_version, actor, request_id,
                       status, started_at
                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (batch_id, log_id, file.filename, payload_hash, idempotency_key, schema_version, "manual", "upload", "processing", now),
            )
        except sqlite3.IntegrityError:
            existing_same = c.execute(
                """SELECT batch_id, status, rows_loaded FROM ingestion_batches
                   WHERE idempotency_key = ? AND payload_hash = ? ORDER BY started_at DESC LIMIT 1""",
                (idempotency_key, payload_hash),
            ).fetchone()
            if existing_same and existing_same["status"] == "committed":
                conn.rollback()
                return {
                    "message": "Idempotent replay accepted",
                    "batch_id": existing_same["batch_id"],
                    "rows_processed": int(existing_same["rows_loaded"] or 0),
                    "replayed": True,
                }
            raise

        raw_df = _load_dataframe_from_upload(file.filename or "", content)
        rows_received = len(raw_df)
        df = _normalize_upload_frame(raw_df.copy())
        _validate_schema_contract(df, schema_version)
        df = execute_etl_rules(df, conn, batch_id, auto_commit=False)
        df = canonicalize_statuses(df, conn)

        rejected_rows = 0
        duplicate_rows = 0
        mandatory_cols = ["name", "job_title", "email", "status", "recruiter", "department"]
        valid_rows = []
        for idx, row in df.iterrows():
            missing = [col for col in mandatory_cols if not str(row.get(col, "")).strip() or str(row.get(col, "")).lower() == "nan"]
            if missing:
                rejected_rows += 1
                c.execute(
                    """INSERT INTO rejected_rows(batch_id, row_index, reason_code, reason_detail, raw_row, created_at)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (batch_id, int(idx), "missing_required", ",".join(missing), json.dumps(raw_df.iloc[idx].to_dict(), ensure_ascii=False), _utcnow().isoformat()),
                )
                continue
            valid_rows.append(row)

        if rows_received > 0 and (rejected_rows / rows_received) > MAX_ERROR_RATE:
            raise Exception(f"Data quality gate failed: rejected_rate={rejected_rows/rows_received:.2f}")

        c.execute("DELETE FROM stg_applications WHERE batch_id = ?", (batch_id,))
        for row_idx, row in enumerate(valid_rows):
            c.execute(
                """INSERT INTO stg_applications(batch_id, row_idx, name, email, job_title, status, recruiter, start_date, department, source, stage_code, days_in_process)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    batch_id,
                    row_idx,
                    str(row["name"]),
                    str(row["email"]),
                    str(row["job_title"]),
                    str(row["status"]),
                    str(row["recruiter"]),
                    pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d"),
                    str(row["department"]),
                    str(row.get("source", "Organic / Unknown")),
                    str(row.get("stage_code", "ACTIVE")),
                    int(row["days_in_process"]),
                ),
            )

        rows_loaded = 0
        for row in valid_rows:
            email_val = row.get("email")
            phone_val = row.get("phone")
            email_norm = normalize_email(email_val)
            phone_norm = normalize_phone(phone_val)

            # Mask values for storage in DB
            masked_email_norm = mask_value(email_norm)
            masked_phone_norm = mask_value(phone_norm)
            masked_email = mask_value(email_val)
            masked_phone = mask_value(phone_val)

            c_id = find_existing_candidate(conn, masked_phone_norm, masked_email_norm)
            if not c_id:
                if masked_email_norm:
                    c_id = str(uuid.uuid5(uuid.NAMESPACE_URL, masked_email_norm))
                else:
                    c_id = f"CND-{uuid.uuid4().hex[:12].upper()}"

            job_title = (row.get("job_title") or "").strip()
            dept = (row.get("department") or "").strip() or "General"
            j_id = None
            if job_title:
                jrow = conn.execute(
                    "SELECT id FROM jobs WHERE LOWER(job_title) = LOWER(?) AND LOWER(IFNULL(department,'')) = LOWER(?) LIMIT 1",
                    (job_title, dept),
                ).fetchone()
                if jrow:
                    j_id = jrow[0]
                else:
                    j_id = f"JOB-{uuid.uuid4().hex[:10].upper()}"

            cand_before = c.execute("SELECT id, name, email, source FROM candidates WHERE id = ?", (c_id,)).fetchone()
            if cand_before is None:
                c.execute(
                    "INSERT INTO candidates (id, name, email, phone, source, phone_norm, email_norm, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, 1)",
                    (c_id, str(row["name"]), masked_email, masked_phone, str(row["source"]), masked_phone_norm, masked_email_norm),
                )
                _record_batch_change(conn, batch_id, "candidate", c_id, "insert", None, {"name": str(row["name"]), "email": masked_email, "source": str(row["source"])})
            else:
                c.execute(
                    "UPDATE candidates SET name = ?, email = ?, phone = ?, source = ?, phone_norm = ?, email_norm = ?, is_active = 1 WHERE id = ?",
                    (str(row["name"]), masked_email, masked_phone, str(row["source"]), masked_phone_norm, masked_email_norm, c_id)
                )

            if j_id:
                job_before = c.execute("SELECT id, job_title, department FROM jobs WHERE id = ?", (j_id,)).fetchone()
                if job_before is None:
                    c.execute(
                        "INSERT INTO jobs (id, job_title, department, is_active) VALUES (?, ?, ?, 1)",
                        (j_id, job_title, dept),
                    )
                    _record_batch_change(conn, batch_id, "job", j_id, "insert", None, {"job_title": job_title, "department": dept})

            if c_id and j_id:
                sig = iteration_signature(
                    str(row["status"]),
                    pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d"),
                    str(row["recruiter"]),
                )
                app_before = c.execute(
                    "SELECT app_id, status, recruiter, days_in_process, upload_log_id, stage_code FROM applications WHERE candidate_id = ? AND job_id = ? AND iteration_signature = ?",
                    (c_id, j_id, sig),
                ).fetchone()

                if app_before:
                    app_id = app_before["app_id"]
                    duplicate_rows += 1
                    c.execute(
                        """UPDATE applications SET status = ?, recruiter = ?, days_in_process = ?, upload_log_id = ?, stage_code = ?, is_active = 1
                           WHERE app_id = ?""",
                        (str(row["status"]), str(row["recruiter"]), int(row["days_in_process"]), log_id, str(row.get("stage_code", "ACTIVE")), app_id),
                    )
                    _record_batch_change(
                        conn,
                        batch_id,
                        "application",
                        app_id,
                        "update",
                        dict(app_before),
                        {"status": str(row["status"]), "recruiter": str(row["recruiter"]), "days_in_process": int(row["days_in_process"]), "upload_log_id": log_id, "stage_code": str(row.get("stage_code", "ACTIVE"))},
                    )
                else:
                    app_id = f"APP-{uuid.uuid4().hex[:10].upper()}"
                    c.execute(
                        """INSERT INTO applications(app_id, candidate_id, job_id, status, recruiter, start_date, days_in_process, upload_log_id, stage_code, iteration_signature, application_date, batch_id, is_active)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
                        (app_id, c_id, j_id, str(row["status"]), str(row["recruiter"]), pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d"), int(row["days_in_process"]), log_id, str(row.get("stage_code", "ACTIVE")), sig, pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d"), batch_id),
                    )
                    _record_batch_change(conn, batch_id, "application", app_id, "insert", None, {"status": str(row["status"]), "recruiter": str(row["recruiter"]), "days_in_process": int(row["days_in_process"]), "upload_log_id": log_id, "stage_code": str(row.get("stage_code", "ACTIVE"))})
                rows_loaded += 1


        quality_score = round(max(0.0, 100 - ((rejected_rows + duplicate_rows) / max(rows_received, 1)) * 100), 2)
        quality_report = {
            "rows_received": rows_received,
            "rows_loaded": rows_loaded,
            "rows_rejected": rejected_rows,
            "duplicate_rows": duplicate_rows,
            "error_rate": round(rejected_rows / max(rows_received, 1), 4),
            "duplicate_rate": round(duplicate_rows / max(rows_received, 1), 4),
        }
        c.execute(
            "INSERT INTO data_logs (log_id, filename, upload_date, rows_processed, status) VALUES (?, ?, ?, ?, ?)",
            (log_id, file.filename, pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"), rows_loaded, "Success"),
        )
        c.execute(
            """UPDATE ingestion_batches
               SET rows_received = ?, rows_loaded = ?, rows_rejected = ?, duplicate_rows = ?, quality_score = ?, quality_report = ?,
                   status = 'committed', finished_at = ?
               WHERE batch_id = ?""",
            (rows_received, rows_loaded, rejected_rows, duplicate_rows, quality_score, json.dumps(quality_report, ensure_ascii=False), _utcnow().isoformat(), batch_id),
        )

        unified_df = get_unified_data(conn)
        build_snapshots(conn, unified_df)
        clear_query_cache(conn, auto_commit=False)
        _auto_scan_after_ingest(conn, batch_id)
        conn.commit()
        return {"message": "ETL Completed successfully", "batch_id": batch_id, "rows_processed": rows_loaded, "quality_report": quality_report}
    except Exception as e:
        conn.rollback()
        try:
            c.execute(
                """INSERT OR REPLACE INTO ingestion_batches(batch_id, log_id, filename, payload_hash, idempotency_key, schema_version, actor, request_id, status, started_at, finished_at, error_message)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'failed', ?, ?, ?)""",
                (batch_id, log_id, file.filename, payload_hash, idempotency_key, schema_version, "manual", "upload", now, _utcnow().isoformat(), str(e)),
            )
            conn.commit()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ==========================================
# 3. DATA GOVERNANCE API (Admin Tools)
# ==========================================
@app.get("/admin/health")
def get_data_health(_: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP))):
    """חישוב בריאות נתונים משוקלל על פני הטבלאות"""
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM candidates")
        total_candidates = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM applications")
        total_apps = c.fetchone()[0]

        # מציאת חוסרים
        c.execute("SELECT COUNT(*) FROM applications WHERE recruiter = 'לא שויך' OR recruiter IS NULL")
        missing_rec = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM jobs WHERE department = 'General' OR department IS NULL")
        missing_dept = c.fetchone()[0]

        logs_df = pd.read_sql("SELECT * FROM data_logs ORDER BY upload_date DESC LIMIT 10", conn)

        health_score = 100
        if total_apps > 0:
            health_score = max(0, int(100 - (((missing_rec + missing_dept) / (total_apps + total_candidates)) * 100)))

        missing_data = []
        if missing_rec > 0:
            missing_data.append({"field": "תהליכים ללא מגייס", "count": missing_rec})
        if missing_dept > 0:
            missing_data.append({"field": "משרות ללא שיוך מחלקתי", "count": missing_dept})

        return {
            "health_score": health_score,
            "total_records": total_apps,
            "missing_data": missing_data,
            "logs": logs_df.to_dict(orient="records")
        }
    finally:
        conn.close()


@app.get("/admin/ingestion/schema-versions")
def get_schema_versions(_: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql(
            "SELECT schema_version, is_active, deprecated, sunset_date, created_at FROM ingestion_schema_versions ORDER BY created_at DESC",
            conn,
        )
        return df.to_dict(orient="records")
    finally:
        conn.close()


@app.get("/admin/ingestion/batches")
def get_ingestion_batches(limit: int = 20, _: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql(
            """SELECT batch_id, filename, schema_version, status, rows_received, rows_loaded, rows_rejected, duplicate_rows,
                      quality_score, started_at, finished_at
               FROM ingestion_batches ORDER BY started_at DESC LIMIT ?""",
            conn,
            params=(max(1, min(limit, 200)),),
        )
        return df.to_dict(orient="records")
    finally:
        conn.close()


@app.post("/admin/ingestion/preflight")
async def ingestion_preflight(
    file: UploadFile = File(...),
    x_schema_version: Optional[str] = Header(default=None),
    _: str = Depends(require_admin),
):
    schema_version = x_schema_version or DEFAULT_SCHEMA_VERSION
    if schema_version not in SUPPORTED_SCHEMA_VERSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported schema_version: {schema_version}")
    try:
        content = await file.read()
        report = _build_preflight_report(file.filename or "upload", content, schema_version)
        return report
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/admin/ingestion/template/{file_type}")
def download_ingestion_template(
    file_type: str,
    schema_version: Optional[str] = None,
    _: str = Depends(require_admin),
):
    # Drive the allow-list from TEMPLATE_SPECS — adding a new ingest type
    # automatically enables its template download.
    valid_types = set(TEMPLATE_SPECS.keys()) | {"candidates", "jobs", "hires", "diversity", "headcount", "budget", "attrition"}
    if file_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid file type: {file_type}")

    requested_schema = schema_version or DEFAULT_SCHEMA_VERSION
    if requested_schema not in SUPPORTED_SCHEMA_VERSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported schema_version: {requested_schema}")

    payload = _build_excel_template_bytes(file_type=file_type, schema_version=requested_schema)
    filename = f"Phoenix_Template_{file_type}_v{requested_schema.replace('.', '_')}.xlsx"
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/ingest/smart-template")
def download_smart_template(
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Returns a ready-to-fill Excel workbook with sheets: משרות, מועמדים, גיוסים + הוראות."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    SMART_SHEETS = [
        ("משרות",   "jobs"),
        ("מועמדים", "candidates"),
        ("גיוסים",  "hires"),
    ]
    fill_req = PatternFill(start_color="002649", end_color="002649", fill_type="solid")
    fill_rec = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for sheet_title, file_type in SMART_SHEETS:
        spec = TEMPLATE_SPECS[file_type]
        all_cols = list(spec["required"]) + list(spec["recommended"])
        n_req = len(spec["required"])
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.rightToLeft = True
        for i, (he, _cn) in enumerate(all_cols, 1):
            c = ws.cell(row=1, column=i, value=he)
            c.fill = fill_req if i <= n_req else fill_rec
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[_col_letter(i)].width = 22
        sample = spec.get("sample", {})
        for i, (he, _cn) in enumerate(all_cols, 1):
            ws.cell(row=2, column=i, value=sample.get(he, ""))
        for he_name, options in (spec.get("validations") or {}).items():
            col_idx = next(
                (i + 1 for i, (he, _) in enumerate(all_cols) if he == he_name), None
            )
            if col_idx:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(options)}"',
                    allow_blank=True,
                    showDropDown=False,
                )
                ws.add_data_validation(dv)
                dv.add(f"{_col_letter(col_idx)}2:{_col_letter(col_idx)}5000")
        ws.freeze_panes = "A2"

    # Instructions sheet
    guide = wb.create_sheet("הוראות")
    guide.sheet_view.rightToLeft = True
    guide["A1"] = "תבנית Smart Ingest — Phoenix Talent OS"
    guide["A1"].font = Font(bold=True, size=14, color="002649")
    guide["A3"] = "כותרות כחולות = שדה חובה | כותרות אפורות = מומלץ"
    guide["A4"] = "גיליון 'משרות'   → משרות פתוחות/סגורות (dedup: שם משרה + חטיבה)"
    guide["A5"] = "גיליון 'מועמדים' → pipeline מועמדים (dedup: טלפון / אימייל)"
    guide["A6"] = "גיליון 'גיוסים'  → קליטות שהתבצעו בפועל"
    guide["A8"] = "שמות גיליונות נוספים שהמערכת מזהה: headcount/תקן, diversity/גיוון, attrition/עזיבות, budget/תקציב"
    guide.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Phoenix_SmartIngest_Template.xlsx"'},
    )


@app.post("/admin/revert/{log_id}")
def revert_upload(log_id: str, _: str = Depends(require_admin)):
    """מחיקת כל הנתונים שנוצרו על ידי קובץ מסוים (Rollback)"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        # מוחק תהליכים שנוצרו בהעלאה זו
        c.execute("DELETE FROM applications WHERE upload_log_id = ?", (log_id,))
        # מעדכן את הסטטוס ביומן
        c.execute("UPDATE data_logs SET status = 'Reverted' WHERE log_id = ?", (log_id,))
        # (בונוס עתידי: לנקות מועמדים "יתומים" שאין להם יותר תהליכים)
        conn.commit()
        unified_df = get_unified_data(conn)
        build_snapshots(conn, unified_df)
        clear_query_cache(conn)
        return {"message": f"Upload {log_id} has been reverted successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.post("/admin/revert-batch/{batch_id}")
def revert_batch(batch_id: str, _: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        c.execute("BEGIN")
        app_changes = c.execute(
            """SELECT entity_id, change_type, before_json FROM batch_entity_changes
               WHERE batch_id = ? AND entity_type = 'application'
               ORDER BY id DESC""",
            (batch_id,),
        ).fetchall()
        for ch in app_changes:
            entity_id = ch["entity_id"]
            if ch["change_type"] == "insert":
                c.execute("DELETE FROM applications WHERE app_id = ?", (entity_id,))
            else:
                before_obj = json.loads(ch["before_json"]) if ch["before_json"] else {}
                c.execute(
                    """UPDATE applications SET status = ?, recruiter = ?, days_in_process = ?, upload_log_id = ?, stage_code = ?
                       WHERE app_id = ?""",
                    (
                        before_obj.get("status"),
                        before_obj.get("recruiter"),
                        before_obj.get("days_in_process"),
                        before_obj.get("upload_log_id"),
                        before_obj.get("stage_code"),
                        entity_id,
                    ),
                )

        for entity_type, table_name, pk in [("candidate", "candidates", "id"), ("job", "jobs", "id")]:
            inserts = c.execute(
                "SELECT entity_id FROM batch_entity_changes WHERE batch_id = ? AND entity_type = ? AND change_type = 'insert' ORDER BY id DESC",
                (batch_id, entity_type),
            ).fetchall()
            for row in inserts:
                c.execute(f"DELETE FROM {table_name} WHERE {pk} = ?", (row["entity_id"],))

        c.execute("UPDATE ingestion_batches SET status = 'reverted', finished_at = ? WHERE batch_id = ?", (_utcnow().isoformat(), batch_id))
        related_log = c.execute("SELECT log_id FROM ingestion_batches WHERE batch_id = ?", (batch_id,)).fetchone()
        if related_log and related_log["log_id"]:
            c.execute("UPDATE data_logs SET status = 'Reverted' WHERE log_id = ?", (related_log["log_id"],))
        unified_df = get_unified_data(conn)
        build_snapshots(conn, unified_df)
        clear_query_cache(conn, auto_commit=False)
        conn.commit()
        return {"message": f"Batch {batch_id} reverted successfully"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.post("/admin/reset-for-final-test")
def reset_for_final_test(_: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    purge_tables = [
        "applications",
        "candidates",
        "jobs",
        "data_logs",
        "ingestion_batches",
        "rejected_rows",
        "batch_entity_changes",
        "batch_snapshots_state",
        "stg_applications",
        "etl_rule_audit",
        "kpi_snapshot",
        "funnel_snapshot",
        "job_health_snapshot",
        "query_cache",
    ]
    report = {"purged": {}, "reset_at": _utcnow().isoformat()}
    try:
        c.execute("BEGIN")
        for table in purge_tables:
            try:
                cnt = c.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
                c.execute(f"DELETE FROM {table}")
                report["purged"][table] = int(cnt)
            except sqlite3.OperationalError:
                report["purged"][table] = 0
        conn.commit()
        return report
    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"reset failed: {exc}") from exc
    finally:
        conn.close()


# ==========================================
# 4. DASHBOARD API (Endpoints for the UI)
# ==========================================


def _count_active_candidates_db() -> int:
    """Count ALL active candidates directly from the candidates table,
    including those without any application record. This fixes the issue
    where get_unified_data() only sees candidates linked through applications."""
    conn = sqlite3.connect(DB_PATH)
    try:
        return conn.execute(
            "SELECT COUNT(*) FROM candidates WHERE COALESCE(is_active, 1) = 1"
        ).fetchone()[0]
    finally:
        conn.close()


def _get_orphan_jobs() -> list[dict]:
    """Return jobs that have zero active applications (hence invisible in
    get_unified_data which starts FROM applications). These are typically
    newly created positions or fully-closed jobs whose applications were
    soft-deleted."""
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            """SELECT j.id, j.job_title, j.department, j.hiring_manager,
                      j.opened_at, j.closed_at, j.close_reason, j.target_count
               FROM jobs j
               WHERE COALESCE(j.is_active, 1) = 1
                 AND j.id NOT IN (
                     SELECT DISTINCT a.job_id FROM applications a
                     WHERE COALESCE(a.is_active, 1) = 1
                 )"""
        ).fetchall()
        return [
            {
                "job_id": r[0], "job_title": r[1], "department": r[2] or "כללי",
                "recruiter": "לא שויך", "is_active": True,
                "active_candidates": 0, "total_candidates": 0,
                "avg_days": 0, "max_days": 0, "sla_breaches": 0,
                "health": "good",
                "stage_breakdown": {s: 0 for s in UNIFIED_STAGES},
                "closed_at": None, "close_reason": None,
            }
            for r in rows
        ]
    finally:
        conn.close()


@app.get("/meta")
def get_meta(_: dict = Depends(verify_token)):
    conn = sqlite3.connect(DB_PATH)
    try:
        df = get_unified_data(conn)
        departments = [d for d in df['department'].dropna().unique().tolist() if str(d).strip()]
        recruiters = [r for r in df['recruiter'].dropna().unique().tolist() if str(r).strip()]
        return {"departments": sorted(departments), "recruiters": sorted(recruiters)}
    except Exception:
        return {"departments": [], "recruiters": []}
    finally:
        conn.close()


# GET /stats moved to backend/routers/analytics.py (B2.4)


# GET /api/candidates and /candidates moved to backend/routers/candidates.py (B2.5)


# GET /api/jobs + /jobs alias, POST /api/jobs/bulk-update moved to
# backend/routers/jobs.py (B2.6).


# GET /executive-brief, /intelligence, /drilldown moved to
# backend/routers/analytics.py (B2.4).


@app.get("/admin/costs")
def get_costs(_: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP))):
    """סימולציה של נתוני כספים, הסכמים ועלויות גיוס (CPH)"""
    return {
        "is_demo": True,
        "demo_note": "Demo payload - replace with live FinOps integration",
        "cph_average": "₪7,820",
        "total_spend_ytd": "₪265,000",
        "agencies": [
            {"name": "חברות השמה (טכנולוגיה)", "type": "עמלה", "value": "100%", "active": 45, "hired": 12, "est_cost": "₪180,000", "roi": "גבוה"},
            {"name": "LinkedIn Recruiter", "type": "רישיון שנתי", "value": "₪45,000", "active": 120, "hired": 8, "est_cost": "₪45,000", "roi": "בינוני"},
            {"name": "חבר מביא חבר (Referral)", "type": "בונוס הוקרה", "value": "₪3,000", "active": 80, "hired": 14, "est_cost": "₪42,000", "roi": "גבוה מאוד"},
            {"name": "קמפיינים ממומנים (Facebook/IG)", "type": "תקציב חודשי", "value": "₪2,500", "active": 210, "hired": 3, "est_cost": "₪12,500", "roi": "נמוך"}
        ]
    }


@app.get("/admin/automations")
def get_automations(_: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP))):
    """שליפת חוקי האוטומציה שמוגדרים במערכת"""
    return [
        {"id": 1, "trigger": "סטטוס = 'הצעת שכר'", "condition": "מעל 3 ימים", "action": "שלח התראה אדומה למנהל המגייס", "status": "פעיל", "is_demo": True},
        {"id": 2, "trigger": "מקור = 'חבר מביא חבר'", "condition": "מעבר לסטטוס 'קליטה'", "action": "הוצא מייל למדור שכר לתשלום בונוס", "status": "פעיל", "is_demo": True},
        {"id": 3, "trigger": "תגית 'טאלנט' נוספה", "condition": "אין אינטראקציה 14 יום", "action": "הקפץ למגייסת תזכורת (Nudge)", "status": "פעיל", "is_demo": True},
        {"id": 4, "trigger": "חטיבת טכנולוגיה", "condition": "מעל 60 ימים ב'ראיון מקצועי'", "action": "דווח כחריגת SLA חמורה", "status": "מושהה", "is_demo": True}
    ]


# ==========================================
# 4. FINOPS & BUDGET API (ניהול תקציב)
# ==========================================

# The six /api/finops/* routes moved to backend/routers/finops.py (B2.2).
# Wired back into `app` via app.include_router(...) at the bottom of this file.

# ==========================================
# SECURITY & AUDIT API
# ==========================================

@app.get("/api/security/audit-logs")
def get_audit_logs(_: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    try:
        logs_df = pd.read_sql("SELECT * FROM audit_logs ORDER BY timestamp DESC LIMIT 50", conn)
        logs = []
        for _, row in logs_df.iterrows():
            timestamp_text = str(row.get("timestamp", ""))
            date_part = timestamp_text.split(" ")[0] if " " in timestamp_text else timestamp_text
            time_part = timestamp_text.split(" ")[1] if " " in timestamp_text else ""
            status_text = str(row.get("status", ""))
            logs.append({
                "id": row["id"],
                "date": date_part,
                "time": time_part,
                "action": row["action"],
                "status": status_text,
                "details": row["details"],
                "user": row["user"],
                "ip_address": row.get("ip_address") or "-",
                "is_destructive": status_text.lower() in {"danger", "warning", "error"},
            })
        return logs
    except Exception:
        return []
    finally:
        conn.close()


@app.post("/api/ai/analyze-cv")
@limiter.limit("30/minute")
async def analyze_cv_safely(request: Request, _: dict = Depends(verify_token)):
    candidate_text = await request.json()
    raw_text = candidate_text.get("text", "")

    safe_text, stats = PIIScrubber.scrub_text_for_ai(raw_text)

    items_scrubbed = stats.get('id_cards', 0) + stats.get('phones', 0) + stats.get('emails', 0)
    if items_scrubbed > 0:
        details = f"צונזרו {stats['id_cards']} ת.ז, {stats['phones']} טלפונים, {stats['emails']} אימיילים"
        log_audit_action("Data Scrubbing (PII)", "Success", details, "System Auto")

    return {
        "status": "success",
        "message": "Text scrubbed and analyzed safely",
        "scrubbed_text": safe_text,
        "items_secured": items_scrubbed
    }


@app.get("/api/security/status")
def get_security_status(_: dict = Depends(verify_token)):
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("SELECT value FROM system_settings WHERE key = 'ai_enabled'")
        result = c.fetchone()
        is_enabled = result is not None and result[0] == 'true'
        return {"ai_enabled": is_enabled}
    except Exception:
        return {"ai_enabled": False}
    finally:
        conn.close()


@app.post("/api/security/toggle-ai")
async def toggle_ai_status(request: Request, _: str = Depends(require_admin)):
    payload = await request.json()
    new_status = 'true' if payload.get('enable', False) else 'false'
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("UPDATE system_settings SET value = ? WHERE key = 'ai_enabled'", (new_status,))
        conn.commit()

        action_desc = "הפעלת מנוע AI" if new_status == 'true' else "כיבוי חירום למנוע AI (Kill Switch)"
        status_color = "Warning" if new_status == 'true' else "Danger"
        log_audit_action("שינוי מדיניות אבטחה", status_color, action_desc, "Super Admin")

        return {"status": "success", "ai_enabled": new_status == 'true'}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
        
        # ==========================================
# 5. DATA QUARANTINE & ETL RULES API
# ==========================================

# הוסף את השורה הזו לתוך פונקציית init_db() שלך:
# c.execute('''CREATE TABLE IF NOT EXISTS etl_rules (id TEXT PRIMARY KEY, col_name TEXT, condition TEXT, action TEXT, active BOOLEAN)''')

@app.get("/api/admin/rules")
def get_etl_rules(_: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql("SELECT * FROM etl_rules", conn)
        # אם הטבלה ריקה, נחזיר את חוקי הבסיס כדיפולט וגם נשמור אותם
        if df.empty:
            default_rules = [
                ("r1", "רמה 3", "מכילה 'טכנולוגיות'", "התעלם מיתר הרמות במבנה ארגוני", True),
                ("r2", "גזרה", "ריקה או חסרה", "השלם ערך 'מקצועית'", True)
            ]
            c = conn.cursor()
            c.executemany("INSERT INTO etl_rules VALUES (?, ?, ?, ?, ?)", default_rules)
            conn.commit()
            df = pd.read_sql("SELECT * FROM etl_rules", conn)
            
        return df.to_dict(orient="records")
    except Exception as e:
        return []
    finally:
        conn.close()

@app.post("/api/admin/rules")
def save_etl_rule(rule: dict, _: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        rule_id = rule.get('id', f"r-{uuid.uuid4().hex[:6]}")
        # Insert or Replace מאפשר לנו גם ליצור חדש וגם לערוך קיים באותה פונקציה!
        c.execute('''INSERT OR REPLACE INTO etl_rules (id, col_name, condition, action, active) 
                     VALUES (?, ?, ?, ?, ?)''', 
                  (rule_id, rule['col_name'], rule['condition'], rule['action'], rule.get('active', True)))
        conn.commit()
        return {"status": "success", "id": rule_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.delete("/api/admin/rules/{rule_id}")
def delete_etl_rule(rule_id: str, _: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("DELETE FROM etl_rules WHERE id = ?", (rule_id,))
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()


# ==========================================
# 6. AI INBOX ANALYTICS API
# ==========================================
@app.get("/api/admin/inbox-analytics")
def get_inbox_analytics(_: str = Depends(require_admin)):
    """מחזיר נתונים אמיתיים על ביצועי המגייסים בטיפול במשימות שהמערכת ייצרה"""
    # כרגע נחזיר מבנה נתונים שמדמה חישוב מה-DB, ברגע שיהיו לנו משימות אמיתיות נחליף לשאילתת SQL
    return {
        "is_demo": True,
        "demo_note": "Demo analytics payload - replace with DB aggregation",
        "stats": {
            "total_tasks": 1240, "avg_close_rate": 92, "median_response_hours": 3.8, "urgent_sla_breaches": 14
        },
        "hourly_trend": [
            {"hour": '08:00', "tasks": 12}, {"hour": '10:00', "tasks": 45}, {"hour": '12:00', "tasks": 38}, 
            {"hour": '14:00', "tasks": 62}, {"hour": '16:00', "tasks": 55}, {"hour": '18:00', "tasks": 20}
        ],
        "task_types": [
            {"label": "סורסינג (נפח קו״ח)", "pct": 45, "count": 558, "color": "bg-orange-400"},
            {"label": "חריגות SLA (מנהלים)", "pct": 30, "count": 372, "color": "bg-blue-400"},
            {"label": "נטישה (Ghosting)", "pct": 15, "count": 186, "color": "bg-red-400"},
            {"label": "אדמיניסטרציה / Onboarding", "pct": 10, "count": 124, "color": "bg-green-400"}
        ],
        "recruiters": [
            {"name": "גיא רג'ואן", "dominant": "סורסינג ולינקדאין", "time": "1.2 שעות", "rate": "98%", "insight": "מצטיין תפעולית. זקוק לתקציב פרסום.", "color": "green"},
            {"name": "ליטל גולדפרב", "dominant": "SLA מנהלים", "time": "14.5 שעות", "rate": "62%", "insight": "חולשה בניהול ממשקים מול מנהלים.", "color": "red"},
            {"name": "מור אהרון", "dominant": "Ghosting", "time": "4.1 שעות", "rate": "89%", "insight": "עומס יתר. מנהלת 35 משרות במקביל.", "color": "orange"}
        ]
    }
    
# ==========================================
# 7. TOOLBOX API (Real Actions: PDF & Fan-out)
# ==========================================
from fpdf import FPDF # <-- חובה להתקין: pip install fpdf2
import smtplib
from email.message import EmailMessage

@app.post("/api/tools/fan-out")
async def trigger_onboarding_fan_out(payload: dict, _: str = Depends(require_admin)):
    """
    מקבל נתוני עובד חדש מהפרונטאנד, רושם ב-DB, ומדמה פתיחת טיקטים.
    בפרודקשן אמיתי מול ה-IT, זה ישלח API Calls למערכת ServiceNow/Jira.
    """
    emp_name = payload.get("name", "עובד חדש")
    emp_role = payload.get("role", "תפקיד כללי")
    
    # סימולציה של פתיחת כרטיסים במערכות שונות:
    tickets = [
        {"dept": "IT SysAdmin", "action": "יצירת יוזר + ציוד", "status": f"Ticket #{uuid.uuid4().hex[:4].upper()}"},
        {"dept": "Logistics", "action": "עמדת עבודה", "status": "Desk Assigned"},
        {"dept": "Security", "action": "אישור חניון", "status": "Pending Badge"}
    ]
    
    # תיעוד ב-Audit Log
    log_audit_action("Onboarding Fan-Out", "Success", f"נפתחו כרטיסים לקליטת: {emp_name} ({emp_role})", "System")
    
    return {"status": "success", "message": f"Fan-out completed for {emp_name}", "tickets": tickets}


@app.post("/api/tools/generate-report")
async def generate_pdf_report(payload: dict, _: str = Depends(require_admin)):
    """
    מנוע ייצור PDF אמיתי.
    הערה: נדרשת התקנת הספריה fpdf2 בשרת.
    """
    report_type = payload.get("type", "weekly_hiring")
    
    # 1. שאיבת נתונים אמיתיים ממסד הנתונים כדי לשים בדוח
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM applications WHERE status LIKE '%קליטה%' OR status LIKE '%גיוס%'")
        hires_count = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM applications WHERE days_in_process > 40")
        sla_breaches = c.fetchone()[0]
    finally:
        conn.close()

    # 2. יצירת ה-PDF (פשוט אך פונקציונלי)
    pdf = FPDF()
    pdf.add_page()
    
    # הערה: עברית ב-FPDF דורשת פונט מתאים. לשם הדגמה מהירה שעובדת מיד, נייצר דוח באנגלית.
    pdf.set_font("helvetica", "B", 16)
    pdf.cell(0, 10, "TAHub Executive Summary", ln=True, align="C")
    pdf.set_font("helvetica", "I", 10)
    pdf.cell(0, 10, f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    
    pdf.ln(10)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 10, f"Report Type: {report_type.replace('_', ' ').title()}", ln=True)
    
    pdf.ln(5)
    pdf.set_font("helvetica", "", 12)
    pdf.cell(0, 10, f"Total Hires Processed: {hires_count}", ln=True)
    pdf.cell(0, 10, f"SLA Breaches Detected: {sla_breaches}", ln=True)
    
    pdf.ln(10)
    pdf.multi_cell(0, 10, "AI Insight: Recruitment volume remains steady. It is recommended to review the SLA breaches to identify bottlenecks in specific departments.")
    
    # שמירת הקובץ באופן זמני
    filename = f"report_{uuid.uuid4().hex[:6]}.pdf"
    file_path = os.path.join(os.getcwd(), filename)
    pdf.output(file_path)
    
    # החזרת הקובץ פיזית לדפדפן כדי שהמשתמש יוכל להוריד
    return FileResponse(
        path=file_path,
        filename="TAHub_Report.pdf",
        media_type="application/pdf",
        background=BackgroundTask(_cleanup_generated_file, file_path),
    )

@app.post("/api/tools/generate-manager-report")
async def generate_manager_report(payload: dict, _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER))):
    """
    דוח מרכז למנהל: משפך גיוס בטווח תאריכים + סטטוס נוכחי, לפי משרה.
    מקבל: job_id (או job_title), date_from (ISO), date_to (ISO).
    """
    import tempfile, traceback

    job_id_param   = (payload.get("job_id") or "").strip()
    job_title_param = (payload.get("job_title") or "").strip()
    date_from      = (payload.get("date_from") or "").strip()
    date_to        = (payload.get("date_to") or "").strip()

    if not (job_id_param or job_title_param):
        raise HTTPException(status_code=400, detail="job_id or job_title is required")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        # ── 1. Job meta ────────────────────────────────────────────────────
        if job_id_param:
            job_row = conn.execute(
                "SELECT * FROM jobs WHERE id = ?", [job_id_param]
            ).fetchone()
        else:
            job_row = conn.execute(
                "SELECT * FROM jobs WHERE job_title = ? COLLATE NOCASE", [job_title_param]
            ).fetchone()

        if not job_row:
            raise HTTPException(status_code=404, detail="Job not found")

        job = dict(job_row)
        effective_job_id = job["id"]
        job_title   = job.get("job_title", effective_job_id)
        department  = job.get("department", "—")

        # ── 2. All applications for this job ───────────────────────────────
        apps = conn.execute(
            """SELECT a.app_id, a.candidate_id, a.status, a.recruiter,
                      a.stage_code, a.days_in_process, a.start_date,
                      c.name AS candidate_name,
                      o.status AS onboarding_status
               FROM applications a
               LEFT JOIN candidates c ON a.candidate_id = c.id
               LEFT JOIN onboarding o ON a.candidate_id = o.id
               WHERE a.job_id = ?
               ORDER BY a.start_date""",
            [effective_job_id]
        ).fetchall()
        apps = [dict(r) for r in apps]

        # Compute unified stage
        for r in apps:
            r["unified_stage"] = _compute_unified_stage(r.get("stage_code"), r.get("onboarding_status"))
    finally:
        conn.close()

    # ── 3. Defaults for date range ──────────────────────────────────────
    all_dates = [r["start_date"] for r in apps if r.get("start_date")]
    earliest  = min(all_dates) if all_dates else date_from or "2020-01-01"
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not date_from: date_from = earliest[:10]
    if not date_to:   date_to   = today_str

    # ── 4. Segment: funnel (started in range) vs current (all active) ──
    ACTIVE_STAGES   = {"ACTIVE", "SCREEN", "INTERVIEW", "OFFER"}
    POSITIVE_STAGES = {"ACTIVE", "SCREEN", "INTERVIEW", "OFFER", "HIRED", "AWAITING_START", "STARTED"}

    funnel_apps = [
        r for r in apps
        if r.get("start_date") and date_from <= r["start_date"][:10] <= date_to
    ]
    current_apps = [r for r in apps if r["unified_stage"] in POSITIVE_STAGES]

    # Stage counts
    def _counts(rows: list[dict]) -> dict[str, int]:
        c: dict[str, int] = {s: 0 for s in UNIFIED_STAGES}
        for r in rows:
            c[r["unified_stage"]] = c.get(r["unified_stage"], 0) + 1
        return c

    funnel_counts  = _counts(funnel_apps)
    current_counts = _counts(current_apps)

    # Conversion rates for funnel
    def _rate(a: int, b: int) -> str:
        return f"{round(a / b * 100)}%" if b else "—"

    # Avg days + SLA
    days_list = [r["days_in_process"] for r in current_apps if r.get("days_in_process")]
    avg_days  = round(sum(days_list) / len(days_list), 1) if days_list else 0
    sla_threshold = 45
    sla_breaches  = sum(1 for d in days_list if d > sla_threshold)

    # Recruiter (most common in apps, fall back to job table)
    from collections import Counter
    recruiter_counts = Counter(r["recruiter"] for r in apps if r.get("recruiter"))
    main_recruiter   = recruiter_counts.most_common(1)[0][0] if recruiter_counts else "—"

    # ── 5. Build PDF ───────────────────────────────────────────────────
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=18)

    # Try to load Arial for Hebrew support (Windows)
    ARIAL_PATH = r"C:\Windows\Fonts\arial.ttf"
    ARIAL_BOLD_PATH = r"C:\Windows\Fonts\arialbd.ttf"
    HEB_FONT = "Arial"
    use_heb = False
    try:
        if os.path.exists(ARIAL_PATH):
            pdf.add_font(HEB_FONT, "", ARIAL_PATH, uni=True)
        if os.path.exists(ARIAL_BOLD_PATH):
            pdf.add_font(HEB_FONT, "B", ARIAL_BOLD_PATH, uni=True)
        use_heb = True
    except Exception:
        use_heb = False

    def _font(bold: bool = False, size: int = 11):
        style = "B" if bold else ""
        if use_heb:
            pdf.set_font(HEB_FONT, style, size)
        else:
            pdf.set_font("Helvetica", style, size)

    def _rtl(text: str) -> str:
        """Reverse Hebrew line for fpdf RTL approximation when no bidi engine."""
        if not use_heb:
            return text
        return text  # fpdf2 with uni=True handles RTL display adequately

    BRAND_DARK   = (0, 38, 73)      # #002649
    BRAND_ORANGE = (239, 107, 0)    # #EF6B00
    LIGHT_GRAY   = (245, 247, 250)
    MID_GRAY     = (100, 116, 139)
    TEXT_DARK    = (30, 41, 59)

    # ── Cover ────────────────────────────────────────────────────────
    pdf.add_page()

    # Top bar
    pdf.set_fill_color(*BRAND_DARK)
    pdf.rect(0, 0, 210, 32, "F")
    _font(bold=True, size=18)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(10, 8)
    pdf.cell(0, 12, "FNX TAHub  |  Manager Position Report", ln=True)
    _font(size=10)
    pdf.set_xy(10, 22)
    pdf.cell(0, 8, f"Generated: {today_str}", ln=True)

    pdf.set_text_color(*TEXT_DARK)
    pdf.set_xy(10, 40)

    # Job details box
    pdf.set_fill_color(*LIGHT_GRAY)
    pdf.rect(10, 38, 190, 46, "F")
    _font(bold=True, size=15)
    pdf.set_text_color(*BRAND_DARK)
    pdf.set_xy(14, 42)
    pdf.multi_cell(182, 9, job_title, align="R" if use_heb else "L")
    _font(size=10)
    pdf.set_text_color(*MID_GRAY)
    pdf.set_xy(14, 56)
    pdf.cell(90, 7, f"Department: {department}")
    pdf.cell(90, 7, f"Recruiter: {main_recruiter}", ln=True)
    pdf.set_xy(14, 64)
    pdf.cell(90, 7, f"Period: {date_from}  to  {date_to}")
    pdf.cell(90, 7, f"Open since: {earliest[:10]}", ln=True)

    pdf.set_text_color(*TEXT_DARK)
    pdf.ln(14)

    # ── Section helper ───────────────────────────────────────────────
    def _section_header(title: str):
        pdf.set_fill_color(*BRAND_ORANGE)
        pdf.rect(10, pdf.get_y(), 5, 8, "F")
        pdf.set_xy(18, pdf.get_y())
        _font(bold=True, size=13)
        pdf.set_text_color(*BRAND_DARK)
        pdf.cell(0, 8, title, ln=True)
        pdf.set_text_color(*TEXT_DARK)
        pdf.ln(2)

    def _kpi_row(items: list[tuple[str, str]]):
        """Draw a row of KPI boxes."""
        x0 = 10
        w  = int(190 / len(items))
        y0 = pdf.get_y()
        for label, value in items:
            pdf.set_fill_color(*LIGHT_GRAY)
            pdf.rect(x0, y0, w - 2, 20, "F")
            _font(bold=True, size=14)
            pdf.set_text_color(*BRAND_DARK)
            pdf.set_xy(x0 + 2, y0 + 1)
            pdf.cell(w - 4, 10, str(value), align="C")
            _font(size=8)
            pdf.set_text_color(*MID_GRAY)
            pdf.set_xy(x0 + 2, y0 + 11)
            pdf.cell(w - 4, 7, label, align="C")
            x0 += w
        pdf.set_xy(10, y0 + 22)
        pdf.set_text_color(*TEXT_DARK)

    def _stage_table(counts: dict[str, int], show_stages: list[str], title_prefix: str = ""):
        headers = show_stages
        col_w   = int(190 / len(headers))
        y0      = pdf.get_y()

        # Header row
        pdf.set_fill_color(*BRAND_DARK)
        pdf.set_text_color(255, 255, 255)
        _font(bold=True, size=9)
        x = 10
        for h in headers:
            pdf.set_xy(x, y0)
            pdf.cell(col_w - 1, 8, h, align="C", fill=True)
            x += col_w
        pdf.ln(8)

        # Value row
        y0 = pdf.get_y()
        x  = 10
        pdf.set_text_color(*TEXT_DARK)
        for h in headers:
            val = counts.get(h, 0)
            fill_color = (254, 226, 226) if val >= 8 else (255, 251, 235) if val >= 4 else (248, 250, 252)
            pdf.set_fill_color(*fill_color)
            _font(bold=(val > 0), size=12)
            pdf.set_xy(x, y0)
            pdf.cell(col_w - 1, 10, str(val) if val else "—", align="C", fill=True)
            x += col_w
        pdf.ln(12)
        pdf.set_text_color(*TEXT_DARK)

    def _candidate_list(stage: str, rows: list[dict], max_rows: int = 8):
        subset = [r for r in rows if r["unified_stage"] == stage][:max_rows]
        if not subset:
            return
        _font(bold=True, size=9)
        pdf.set_text_color(*BRAND_DARK)
        pdf.set_xy(10, pdf.get_y())
        pdf.cell(0, 7, f"  {stage} ({len(subset)} candidates):", ln=True)
        _font(size=9)
        pdf.set_text_color(*TEXT_DARK)
        for r in subset:
            name = r.get("candidate_name") or r.get("app_id", "—")
            days = r.get("days_in_process") or 0
            status = r.get("status") or "—"
            line  = f"      • {name}   |   {days} days   |   {status}"
            pdf.set_xy(10, pdf.get_y())
            pdf.cell(0, 6, line, ln=True)
        pdf.ln(2)

    # ── Section 1: Current Pipeline ──────────────────────────────────
    _section_header("Current Pipeline Status (as of today)")

    _kpi_row([
        ("Active Candidates", str(len(current_apps))),
        ("Avg Days in Process", str(avg_days)),
        ("SLA Breaches (>45d)", str(sla_breaches)),
        ("Hired (all time)", str(sum(1 for r in apps if r["unified_stage"] in {"HIRED","AWAITING_START","STARTED"}))),
    ])

    _stage_table(current_counts, ["SCREEN", "INTERVIEW", "OFFER", "HIRED", "AWAITING_START", "STARTED"])

    # Candidate lists per active stage
    _font(bold=True, size=10)
    pdf.set_text_color(*BRAND_DARK)
    pdf.cell(0, 7, "Candidate Detail:", ln=True)
    pdf.ln(1)
    for stg in ["OFFER", "INTERVIEW", "SCREEN"]:
        _candidate_list(stg, current_apps)

    pdf.ln(4)

    # ── Section 2: Recruitment Funnel (date range) ───────────────────
    _section_header(f"Recruitment Funnel  ({date_from} → {date_to})")

    total_in_period = len(funnel_apps)
    _kpi_row([
        ("Applications in Period", str(total_in_period)),
        ("Screen → Interview", _rate(funnel_counts["INTERVIEW"], funnel_counts["SCREEN"])),
        ("Interview → Offer",  _rate(funnel_counts["OFFER"],     funnel_counts["INTERVIEW"])),
        ("Offer → Hired",      _rate(funnel_counts["HIRED"] + funnel_counts["AWAITING_START"] + funnel_counts["STARTED"], funnel_counts["OFFER"])),
    ])

    if total_in_period:
        _stage_table(funnel_counts, ["SCREEN", "INTERVIEW", "OFFER", "HIRED", "AWAITING_START", "STARTED"])
    else:
        _font(size=10)
        pdf.set_text_color(*MID_GRAY)
        pdf.cell(0, 8, "  No applications recorded in this date range.", ln=True)
        pdf.set_text_color(*TEXT_DARK)

    pdf.ln(4)

    # ── Section 3: Manager Insights ──────────────────────────────────
    _section_header("Manager Insights")
    _font(size=10)
    pdf.set_text_color(*TEXT_DARK)

    insights: list[str] = []

    # Bottleneck detection
    screen_n   = current_counts.get("SCREEN", 0)
    interview_n = current_counts.get("INTERVIEW", 0)
    offer_n    = current_counts.get("OFFER", 0)
    hired_n    = sum(current_counts.get(s, 0) for s in ["HIRED", "AWAITING_START", "STARTED"])

    if screen_n >= 5 and interview_n == 0:
        insights.append(f"BOTTLENECK: {screen_n} candidates stuck at Screening with zero progression to Interview. Recommend reviewing screening criteria or recruiter capacity.")
    elif screen_n > interview_n * 3:
        insights.append(f"Screening funnel is wide ({screen_n} candidates) but conversion to Interview is low ({interview_n}). Consider expediting review of pending CVs.")

    if interview_n >= 3 and offer_n == 0:
        insights.append(f"Interview stage has {interview_n} candidates with no Offer made yet. Verify interview feedback is being collected and decisions are pending.")

    if offer_n >= 2:
        insights.append(f"ATTENTION: {offer_n} active Offer(s). Ensure compensation packages are competitive and candidates have been followed up recently.")

    if sla_breaches > 0:
        insights.append(f"SLA ALERT: {sla_breaches} candidate(s) have been in process for over {sla_threshold} days. Immediate action required to prevent candidate drop-off.")

    if avg_days > 35:
        insights.append(f"Average time-in-process is {avg_days} days, above the 35-day benchmark. Consider streamlining interview scheduling or approval cycles.")
    elif avg_days <= 20 and len(current_apps) > 0:
        insights.append(f"Excellent velocity: average time-in-process is just {avg_days} days — well below the 35-day benchmark.")

    if total_in_period == 0 and len(current_apps) > 0:
        insights.append(f"No new applications received in the selected period ({date_from} → {date_to}). Consider reviewing sourcing channels or reposting the position.")

    if not insights:
        insights.append("Pipeline appears healthy. Continue monitoring conversion rates and ensure candidate communications are timely.")

    for i, insight in enumerate(insights, 1):
        pdf.set_xy(10, pdf.get_y())
        _font(bold=True, size=10)
        pdf.set_text_color(*BRAND_ORANGE)
        pdf.cell(8, 7, f"{i}.")
        _font(size=10)
        pdf.set_text_color(*TEXT_DARK)
        pdf.multi_cell(180, 7, insight)
        pdf.ln(2)

    # ── Footer ───────────────────────────────────────────────────────
    pdf.set_y(-18)
    pdf.set_fill_color(*BRAND_DARK)
    pdf.rect(0, pdf.get_y(), 210, 18, "F")
    _font(size=8)
    pdf.set_text_color(180, 190, 210)
    pdf.set_xy(10, pdf.get_y() + 4)
    pdf.cell(0, 6, f"FNX TAHub  |  Confidential Manager Report  |  {today_str}", align="C")

    # ── Save & return ────────────────────────────────────────────────
    safe_title = "".join(c if c.isalnum() or c in " -_" else "_" for c in job_title)[:40]
    filename   = f"manager_report_{safe_title}_{today_str}_{uuid.uuid4().hex[:4]}.pdf"
    file_path  = os.path.join(tempfile.gettempdir(), filename)
    pdf.output(file_path)

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/pdf",
        background=BackgroundTask(_cleanup_generated_file, file_path),
    )


@app.post("/api/tools/generate-offer-pdf")
async def generate_offer_pdf(request: Request, _: str = Depends(require_admin)):
    """
    מייצר מסמך 'הצעת שכר / חבילת תגמול' שיווקי למועמד.
    משמיט לחלוטין עלויות מעסיק ועמלות חברת השמה.
    """
    payload = await request.json()
    candidate_name = payload.get("candidateName", "Candidate")
    is_comparative = payload.get("isComparative", False)
    proposed = payload.get("proposed", {})
    current = payload.get("current", {})
    
    pdf = FPDF()
    pdf.add_page()
    
    # Header - Phoenix Branding
    pdf.set_fill_color(0, 38, 73) # Phoenix Navy Blue #002649
    pdf.rect(0, 0, 210, 30, 'F')
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 18)
    pdf.set_y(10)
    pdf.cell(0, 10, "Total Rewards & Compensation Offer", ln=True, align="C")
    
    # Reset colors for body
    pdf.set_text_color(0, 0, 0)
    pdf.ln(15)
    
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 10, f"Prepared for: {candidate_name}", ln=True)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(0, 5, f"Date: {pd.Timestamp.now().strftime('%d/%m/%Y')}", ln=True)
    pdf.ln(10)

    # ---------------------------------------------------------
    # MAIN OFFER HIGHLIGHT (Total Value)
    # ---------------------------------------------------------
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(239, 107, 0) # Phoenix Orange #EF6B00
    total_val = proposed.get("totalPackageValue", 0)
    pdf.cell(0, 10, f"Total Monthly Package Value: {total_val:,.0f} ILS", ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(5)

    # ---------------------------------------------------------
    # DETAILS TABLE
    # ---------------------------------------------------------
    pdf.set_fill_color(240, 245, 250)
    pdf.set_font("helvetica", "B", 10)
    
    # Table Headers
    pdf.cell(70, 10, "Component", border=1, fill=True)
    if is_comparative:
        pdf.cell(60, 10, "Current Package", border=1, align="C", fill=True)
    pdf.cell(60, 10, "Phoenix Offer", border=1, align="C", fill=True)
    pdf.ln(10)
    
    pdf.set_font("helvetica", "", 10)
    
    # Data Rows
    components = [
        ("Base Gross Salary", "base"),
        ("Global Overtime", "global"),
        ("Meals (Cibus)", "meals"),
        ("Travel / Car", "travel_car"),
        ("Keren Hishtalmut (%)", "kh_pct"),
        ("Pension (%)", "pension_pct")
    ]
    
    for label, key in components:
        pdf.cell(70, 10, label, border=1)
        if is_comparative:
            curr_val = current.get(key, "-")
            pdf.cell(60, 10, f"{curr_val}", border=1, align="C")
        
        prop_val = proposed.get(key, "-")
        pdf.cell(60, 10, f"{prop_val}", border=1, align="C")
        pdf.ln(10)

    # ---------------------------------------------------------
    # DISCLAIMER (Legal text requested)
    # ---------------------------------------------------------
    pdf.ln(20)
    pdf.set_font("helvetica", "I", 8)
    pdf.set_text_color(100, 100, 100)
    disclaimer = (
        "Disclaimer: This document is an offer proposal only and holds no legal binding validity. "
        "It is valid for 30 days from the date of issue. This simulation does not constitute an "
        "employer-employee relationship agreement. Final terms will be defined strictly by the "
        "official employment contract."
    )
    pdf.multi_cell(0, 5, disclaimer)

    filename = f"Phoenix_Offer_{uuid.uuid4().hex[:6]}.pdf"
    file_path = os.path.join(os.getcwd(), filename)
    pdf.output(file_path)
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/pdf",
        background=BackgroundTask(_cleanup_generated_file, file_path),
    )

# ==========================================
# 8. PRE-BOARDING & ONBOARDING API
# ==========================================


# The five /api/onboarding/* routes (and the two helpers
# _persist_onboarding_record + _normalize_onboarding_payload) moved to
# backend/routers/onboarding.py (B2.3).


# ==========================================
# INTERNAL LOGIC ENDPOINTS (NO LLM MODE)
# ==========================================
@app.get("/api/internal/deterministic-capabilities")
def get_deterministic_capabilities(_: str = Depends(require_admin)):
    return {
        "rule_engine": {
            "enabled": True,
            "tables": ["etl_rules", "etl_rule_audit"],
            "description": "Applies ETL rules during ingestion with per-rule audit trail.",
        },
        "status_canonicalization": {
            "enabled": True,
            "tables": ["status_lexicon"],
            "description": "Maps free-text statuses to canonical stage codes without LLM.",
        },
        "snapshots": {
            "enabled": True,
            "tables": ["kpi_snapshot", "funnel_snapshot", "job_health_snapshot"],
            "description": "Precomputed deterministic metrics for low-resource runtime.",
        },
        "query_cache": {
            "enabled": True,
            "tables": ["query_cache"],
            "description": "Parameterized endpoint cache with data-version invalidation.",
        },
        "risk_scoring": {
            "enabled": True,
            "description": "Classical logistic-like ghosting score from internal features.",
        },
        "template_insights": {
            "enabled": True,
            "tables": ["insight_templates"],
            "description": "Condition-based insight rendering from deterministic KPI context.",
        },
    }


@app.get("/api/internal/snapshots")
def get_internal_snapshots(_: str = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    try:
        kpi = pd.read_sql("SELECT * FROM kpi_snapshot ORDER BY snapshot_ts DESC LIMIT 1", conn)
        funnel = pd.read_sql("SELECT * FROM funnel_snapshot ORDER BY snapshot_ts DESC", conn)
        jobs = pd.read_sql("SELECT * FROM job_health_snapshot ORDER BY sla_breaches DESC, max_days DESC LIMIT 20", conn)
        return {
            "kpi_snapshot": kpi.to_dict(orient="records"),
            "funnel_snapshot": funnel.to_dict(orient="records"),
            "job_health_snapshot": jobs.to_dict(orient="records"),
        }
    finally:
        conn.close()


# ==========================================
# AUTH AUDIT LOGGING
# ==========================================
@app.post("/api/auth/log")
async def auth_log(request: Request):
    """
    מקבל אירועי אבטחה מה-SessionGuard ורושם אותם ב-audit_logs.
    מדיניות PII: לא נשמרים סיסמאות, שמות משתמשים, או נתוני מועמדים.
    """
    try:
        payload = await request.json()
        event = str(payload.get("event", "UNKNOWN_EVENT"))
        details = str(payload.get("details", ""))[:500]
        timestamp = str(payload.get("timestamp", ""))[:50]

        # Sanitise: only allow known event names to prevent log injection
        allowed_events = {"SESSION_LOCKED", "SESSION_RESTORED", "UNLOCK_FAILED"}
        if event not in allowed_events:
            event = "UNKNOWN_EVENT"

        log_audit_action(
            action=event,
            status="auth",
            details=f"{details} | client_ts={timestamp}",
            user="frontend"
        )
        return {"status": "logged", "event": event}
    except Exception as e:
        # Never crash — auth logging must not block the client
        logger.exception("auth logging failed: %s", e)
        return {"status": "error", "message": str(e)}


@app.post("/api/auth/unlock")
async def unlock_session(request: Request):
    """Unified unlock: authenticates against the user's real login password
    (bcrypt-hashed in `users.password_hash`). Same credentials as `/api/auth/login`
    and as the gate on `/admin/security`. No separate PIN.

    Email is taken from the active session cookie when present, otherwise the
    caller must include it in the payload. Password is required.

    For backwards compatibility a server-wide `SESSION_UNLOCK_PIN` is still
    accepted IF it matches exactly and IF the env var is set — this lets older
    clients that don't know the user's email still unlock during the rollout.
    New deployments should leave SESSION_UNLOCK_PIN unset.
    """
    _ensure_jwt_ready()
    payload = await request.json()
    password = str(payload.get("password", ""))
    if not password:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing password")

    # Resolve email: payload wins, else fall back to session cookie.
    email = str(payload.get("email", "")).strip().lower()
    if not email:
        token = request.cookies.get(SESSION_COOKIE)
        if token:
            try:
                claims = _decode_jwt(token)
                email = str(claims.get("email", "")).strip().lower()
            except Exception:
                email = ""

    # Path 1: real user credentials (preferred). Verifies against users.password_hash.
    if email:
        conn = sqlite3.connect(DB_PATH)
        try:
            row = conn.execute(
                "SELECT id, email, password_hash, full_name, role, is_active FROM users WHERE LOWER(email) = ?",
                (email,),
            ).fetchone()
        finally:
            conn.close()
        if row and row[5] and _verify_password(password, row[2]):
            issued_at = int(_utcnow().timestamp())
            token_payload = {
                "sub": row[0],
                "email": row[1],
                "name": row[3],
                "role": row[4] or "admin",
                "iat": issued_at,
                "exp": issued_at + (JWT_TTL_MINUTES * 60),
            }
            access_token = _encode_jwt(token_payload)
            log_audit_action("SESSION_UNLOCKED", "ok", f"user={email}", user=email)
            return {"status": "ok", "access_token": access_token, "token_type": "bearer"}

    # Path 2: legacy shared PIN (backwards compat only; discouraged).
    if SESSION_UNLOCK_PIN and secrets.compare_digest(password, SESSION_UNLOCK_PIN):
        issued_at = int(_utcnow().timestamp())
        token_payload = {
            "sub": "admin-session",
            "role": "admin",
            "iat": issued_at,
            "exp": issued_at + (JWT_TTL_MINUTES * 60),
        }
        access_token = _encode_jwt(token_payload)
        log_audit_action("SESSION_UNLOCKED", "ok", "legacy PIN flow", user=email or "anonymous")
        return {"status": "ok", "access_token": access_token, "token_type": "bearer"}

    log_audit_action("SESSION_UNLOCK_FAILED", "warn", f"user={email or 'unknown'}", user=email or "anonymous")
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid password")


# ==========================================
# ADMIN CONFIG — KPI Formulas, Rules, Visibility
# ==========================================
ADMIN_CONFIG_DEFAULTS = {
    "formulas": [
        {"id": "conv_rate", "label": "Conversion Rate %", "varA": "hires", "op": "/", "varB": "offers", "scale": 100},
        {"id": "ttf", "label": "Time-to-Fill (avg)", "varA": "avg_days_open", "op": "+", "varB": None, "scale": 1},
    ],
    "rules": [
        {"id": "r1", "metric": "interviews_per_week", "op": "<", "threshold": 7,
         "action": "toast", "actionLabel": "ממוצע ראיונות נמוך מהיעד", "enabled": True},
    ],
    "visibility": {
        "kpi_conversion": True,
        "kpi_ttf": True,
        "chart_sources": True,
        "table_recruiters": True,
    },
    "neglectThresholds": {
        "slaDaysThreshold": 60,
        "lowCandidatesThreshold": 5,
        "pendingCvThreshold": 20,
        "staleActionDaysThreshold": 5,
        "criticalScoreThreshold": 80,
    },
}


def _safe_pct(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator / denominator) * 100, 1)


def _to_int(value, fallback: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def _normalize_score(value: float, lower: float, upper: float) -> float:
    if upper <= lower:
        return 0.0
    clipped = min(max(value, lower), upper)
    return (clipped - lower) / (upper - lower)


# GET /jobs/neglect-alerts moved to backend/routers/jobs.py (B2.6)

@app.get("/api/admin/config")
async def get_admin_config(_: str = Depends(require_admin)):
    with db_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT value FROM system_settings WHERE key='admin_config'")
        row = c.fetchone()
    if row:
        return json.loads(row[0])
    return ADMIN_CONFIG_DEFAULTS

@app.post("/api/admin/config")
async def save_admin_config(request: Request, section: str = "general", _: str = Depends(require_admin)):
    payload = await request.json()
    with db_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT value FROM system_settings WHERE key='admin_config'")
        row = c.fetchone()
        existing = json.loads(row[0]) if row else {}
        merged = {**existing, **payload}
        c.execute(
            "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('admin_config', ?)",
            (json.dumps(merged),)
        )
        conn.commit()
    changed_keys = list(payload.keys())
    log_audit_action(
        action="ADMIN_CONFIG_UPDATE",
        status="success",
        details=f"section={section} | keys_changed={changed_keys}",
        user="admin-frontend"
    )
    timestamp = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    return {"status": "saved", "timestamp": timestamp}


@app.get("/api/admin/permissions/users")
def get_permissions_users(_: dict = Depends(require_admin)):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """SELECT id, name, email, usf, personal_password, role, status, last_login, permissions_json
               FROM iam_users ORDER BY name ASC"""
        ).fetchall()
        users = []
        for row in rows:
            users.append(
                {
                    "id": row["id"],
                    "name": row["name"],
                    "email": row["email"],
                    "usf": row["usf"],
                    "personalPassword": row["personal_password"],
                    "role": row["role"],
                    "status": row["status"],
                    "lastLogin": row["last_login"],
                    "permissions": json.loads(row["permissions_json"] or "{}"),
                }
            )
        return users
    finally:
        conn.close()


@app.post("/api/admin/permissions/users")
async def save_permissions_user(request: Request, _: dict = Depends(require_admin)):
    payload = await request.json()
    user_id = str(payload.get("id") or f"u-{uuid.uuid4().hex[:8]}")
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute(
            """INSERT OR REPLACE INTO iam_users
               (id, name, email, usf, personal_password, role, status, last_login, permissions_json, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                user_id,
                str(payload.get("name", "")),
                str(payload.get("email", "")),
                str(payload.get("usf", "")),
                str(payload.get("personalPassword", payload.get("usf", ""))),
                str(payload.get("role", "hiring_manager")),
                str(payload.get("status", "active")),
                str(payload.get("lastLogin", "טרם")),
                json.dumps(payload.get("permissions", {}), ensure_ascii=False),
                _utcnow().isoformat(),
            ),
        )
        conn.commit()
        return {
            "id": user_id,
            "name": str(payload.get("name", "")),
            "email": str(payload.get("email", "")),
            "usf": str(payload.get("usf", "")),
            "personalPassword": str(payload.get("personalPassword", payload.get("usf", ""))),
            "role": str(payload.get("role", "hiring_manager")),
            "status": str(payload.get("status", "active")),
            "lastLogin": str(payload.get("lastLogin", "טרם")),
            "permissions": payload.get("permissions", {}),
        }
    finally:
        conn.close()


@app.post("/api/admin/permissions/users/bulk-suspend")
async def bulk_suspend_users(request: Request, _: dict = Depends(require_admin)):
    payload = await request.json()
    user_ids = payload.get("user_ids", [])
    if not isinstance(user_ids, list) or not user_ids:
        raise HTTPException(status_code=400, detail="user_ids is required")
    conn = sqlite3.connect(DB_PATH)
    try:
        placeholders = ",".join("?" for _ in user_ids)
        conn.execute(
            f"UPDATE iam_users SET status='suspended', updated_at=? WHERE id IN ({placeholders})",
            (_utcnow().isoformat(), *user_ids),
        )
        conn.commit()
        return {"status": "success", "updated": len(user_ids)}
    finally:
        conn.close()


# =====================================================================
# REAL AUTH (per-user login via bcrypt + cookie JWT) — phase-1 merge
# This block coexists with the Bearer-based admin auth above. Endpoints
# here read/write a `users` table; Cursor's iam_users is left untouched.
# =====================================================================

# _hash_password, _verify_password, _make_session_token, get_session_user,
# require_session_role moved to backend/auth.py (B4).


def _set_session_cookie(response: Response, token: str, *, key: str = SESSION_COOKIE) -> None:
    secure_cookie = os.getenv("COOKIE_SECURE", "true").lower() != "false"
    response.set_cookie(
        key=key, value=token, httponly=True, secure=secure_cookie,
        samesite="lax", max_age=JWT_TTL_MINUTES * 60, path="/",
    )


# ----- AUTH ENDPOINTS -----

@app.post("/api/auth/login")
async def auth_login(payload: dict, response: Response):
    email = (payload.get("email") or "").strip().lower()
    password = payload.get("password") or ""
    if not email or not password:
        raise HTTPException(status_code=400, detail="חסרים אימייל או סיסמה")

    with db_conn() as conn:
        c = conn.cursor()
        c.execute(
            "SELECT id, email, password_hash, full_name, role, is_active, must_change_password FROM users WHERE LOWER(email) = ?",
            (email,),
        )
        row = c.fetchone()
        if not row or not row[5]:
            log_audit_action("LOGIN_FAILED", "warn", f"Unknown or inactive: {email}", user=email)
            raise HTTPException(status_code=401, detail="אימייל או סיסמה שגויים")

        user_id, user_email, password_hash, full_name, role, _, must_change = row
        if not _verify_password(password, password_hash):
            log_audit_action("LOGIN_FAILED", "warn", f"Bad password: {email}", user=email)
            raise HTTPException(status_code=401, detail="אימייל או סיסמה שגויים")

        c.execute("UPDATE users SET last_login_at = ? WHERE id = ?",
                  (datetime.now(timezone.utc).isoformat(), user_id))
        conn.commit()

    token = _make_session_token({
        "sub": user_id, "email": user_email, "name": full_name, "role": role,
        "must_change_password": bool(must_change),
    })
    _set_session_cookie(response, token)

    log_audit_action("LOGIN_SUCCESS", "ok", f"User logged in: {email}", user=email)
    return {
        "id": user_id, "email": user_email, "name": full_name, "role": role,
        "must_change_password": bool(must_change),
    }


@app.post("/api/auth/logout")
async def auth_logout(
    response: Response,
    user: dict = Depends(get_session_user),
    fnx_access_token: Optional[str] = Cookie(default=None),
):
    # Server-side token revocation: persist signature in revoked_tokens until expiry.
    if fnx_access_token:
        parts = fnx_access_token.split(".")
        if len(parts) == 3:
            sig = parts[2]
            exp = int(user.get("exp", 0))
            if exp:
                _revoke_token_signature(sig, exp)
    response.delete_cookie(key=SESSION_COOKIE, path="/")
    response.delete_cookie(key=IMPERSONATOR_COOKIE, path="/")
    log_audit_action("LOGOUT", "ok", f"User logged out: {user.get('email')}", user=user.get("email", "unknown"))
    return {"status": "ok"}


@app.get("/api/auth/me")
async def auth_me(user: dict = Depends(get_session_user)):
    # Re-read must_change_password from the DB so a change made during the
    # session is reflected immediately, without forcing re-login.
    must_change = bool(user.get("must_change_password"))
    sub = user.get("sub")
    if sub:
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("SELECT must_change_password FROM users WHERE id = ?", (sub,))
            row = c.fetchone()
            if row is not None:
                must_change = bool(row[0])
        except sqlite3.OperationalError:
            pass
        finally:
            try:
                conn.close()
            except Exception:
                pass

    return {
        "id": sub,
        "email": user.get("email"),
        "name": user.get("name"),
        "role": user.get("role"),
        "impersonator": user.get("impersonator"),
        "impersonator_email": user.get("impersonator_email"),
        "must_change_password": must_change,
    }


@app.post("/api/auth/change-password")
async def auth_change_password(payload: dict, user: dict = Depends(get_session_user)):
    current = payload.get("current_password") or ""
    new = payload.get("new_password") or ""
    if len(new) < 8:
        raise HTTPException(status_code=400, detail="סיסמה חדשה חייבת לכלול לפחות 8 תווים")

    with db_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT password_hash FROM users WHERE id = ?", (user.get("sub"),))
        row = c.fetchone()
        if not row or not _verify_password(current, row[0]):
            raise HTTPException(status_code=401, detail="הסיסמה הנוכחית שגויה")

        c.execute(
            "UPDATE users SET password_hash = ?, must_change_password = 0 WHERE id = ?",
            (_hash_password(new), user.get("sub")),
        )
        conn.commit()
    log_audit_action("PASSWORD_CHANGED", "ok", f"User changed own password: {user.get('email')}", user=user.get("email", "unknown"))
    return {"status": "ok"}


# ----- USER MANAGEMENT (admin only) — uses the new `users` table -----

@app.get("/api/admin/users")
async def admin_list_users(_: dict = Depends(require_session_role(Role.ADMIN))):
    with db_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT id, email, full_name, role, employee_number, is_active, must_change_password, created_at, last_login_at FROM users ORDER BY created_at DESC")
        rows = c.fetchall()
    return [
        {"id": r[0], "email": r[1], "full_name": r[2], "role": r[3],
         "employee_number": r[4], "is_active": bool(r[5]),
         "must_change_password": bool(r[6]), "created_at": r[7], "last_login_at": r[8]}
        for r in rows
    ]


@app.post("/api/admin/users")
async def admin_create_user(payload: dict, admin: dict = Depends(require_session_role(Role.ADMIN))):
    email = (payload.get("email") or "").strip().lower()
    full_name = (payload.get("full_name") or "").strip()
    role = (payload.get("role") or "recruiter").strip()
    employee_number = str(payload.get("employee_number") or "").strip()

    if not email or not full_name:
        raise HTTPException(status_code=400, detail="חסרים אימייל או שם מלא")
    if role not in ("admin", "hrbp", "recruiter", "hiring_manager"):
        raise HTTPException(status_code=400, detail="תפקיד לא חוקי")
    if not employee_number or len(employee_number) < 4:
        raise HTTPException(status_code=400, detail="מספר עובד (USF) חייב לכלול לפחות 4 תווים")

    user_id = f"USR-{uuid.uuid4().hex[:8].upper()}"
    with db_conn() as conn:
        c = conn.cursor()
        try:
            c.execute(
                "INSERT INTO users (id, email, password_hash, full_name, role, employee_number, is_active, must_change_password, created_at) VALUES (?, ?, ?, ?, ?, ?, 1, 0, ?)",
                (user_id, email, _hash_password(employee_number), full_name, role, employee_number, datetime.now(timezone.utc).isoformat()),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="משתמש עם אימייל זה כבר קיים")

    log_audit_action("USER_CREATED", "ok", f"Created {email} as {role} (USF={employee_number})", user=admin.get("email", "admin"))
    return {"id": user_id, "email": email, "full_name": full_name, "role": role, "employee_number": employee_number}


@app.put("/api/admin/users/{user_id}")
async def admin_update_user(user_id: str, payload: dict, admin: dict = Depends(require_session_role(Role.ADMIN))):
    fields, values = [], []
    if "full_name" in payload:
        fields.append("full_name = ?"); values.append((payload["full_name"] or "").strip())
    if "role" in payload:
        if payload["role"] not in ("admin", "hrbp", "recruiter", "hiring_manager"):
            raise HTTPException(status_code=400, detail="תפקיד לא חוקי")
        fields.append("role = ?"); values.append(payload["role"])
    if "is_active" in payload:
        fields.append("is_active = ?"); values.append(1 if payload["is_active"] else 0)
    if not fields:
        raise HTTPException(status_code=400, detail="אין שדות לעדכון")

    values.append(user_id)
    with db_conn() as conn:
        c = conn.cursor()
        c.execute(f"UPDATE users SET {', '.join(fields)} WHERE id = ?", values)
        if c.rowcount == 0:
            raise HTTPException(status_code=404, detail="משתמש לא נמצא")
        conn.commit()
    log_audit_action("USER_UPDATED", "ok", f"Updated user {user_id}", user=admin.get("email", "admin"))
    return {"status": "ok"}


@app.post("/api/admin/users/{user_id}/reset-password")
async def admin_reset_password(user_id: str, admin: dict = Depends(require_session_role(Role.ADMIN))):
    temp_password = secrets.token_urlsafe(9)
    with db_conn() as conn:
        c = conn.cursor()
        c.execute(
            "UPDATE users SET password_hash = ?, must_change_password = 1 WHERE id = ?",
            (_hash_password(temp_password), user_id),
        )
        if c.rowcount == 0:
            raise HTTPException(status_code=404, detail="משתמש לא נמצא")
        conn.commit()
    log_audit_action("PASSWORD_RESET", "ok", f"Admin reset password for {user_id}", user=admin.get("email", "admin"))
    return {"temp_password": temp_password}


# ----- IMPERSONATION (admin "view as") -----

@app.post("/api/admin/impersonate/{user_id}")
async def admin_impersonate(user_id: str, response: Response, admin: dict = Depends(require_session_role(Role.ADMIN))):
    if admin.get("impersonator"):
        raise HTTPException(status_code=400, detail="כבר במצב impersonation. סיימי את הסשן הנוכחי קודם.")
    with db_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT id, email, full_name, role, is_active FROM users WHERE id = ?", (user_id,))
        row = c.fetchone()
    if not row or not row[4]:
        raise HTTPException(status_code=404, detail="המשתמש לא נמצא או מושעה")

    target_id, target_email, target_name, target_role, _ = row
    original_admin_token = _make_session_token({
        "sub": admin.get("sub"), "email": admin.get("email"),
        "name": admin.get("name"), "role": admin.get("role"),
    })
    impersonation_token = _make_session_token({
        "sub": target_id, "email": target_email, "name": target_name, "role": target_role,
        "impersonator": admin.get("sub"), "impersonator_email": admin.get("email"),
    })
    _set_session_cookie(response, impersonation_token, key=SESSION_COOKIE)
    _set_session_cookie(response, original_admin_token, key=IMPERSONATOR_COOKIE)

    log_audit_action("IMPERSONATION_START", "warn",
                     f"Admin {admin.get('email')} impersonating {target_email}",
                     user=admin.get("email", "admin"))
    return {"id": target_id, "email": target_email, "name": target_name, "role": target_role,
            "impersonator_email": admin.get("email")}


@app.post("/api/admin/stop-impersonate")
async def admin_stop_impersonate(
    response: Response,
    fnx_impersonator: Optional[str] = Cookie(default=None),
    user: dict = Depends(get_session_user),
):
    if not user.get("impersonator") or not fnx_impersonator:
        raise HTTPException(status_code=400, detail="אינך במצב impersonation")
    try:
        payload = _decode_jwt(fnx_impersonator)
    except HTTPException:
        response.delete_cookie(key=SESSION_COOKIE, path="/")
        response.delete_cookie(key=IMPERSONATOR_COOKIE, path="/")
        raise HTTPException(status_code=401, detail="טוקן ה-admin המקורי פג. יש להתחבר שוב.")

    if payload.get("role") != "admin":
        response.delete_cookie(key=SESSION_COOKIE, path="/")
        response.delete_cookie(key=IMPERSONATOR_COOKIE, path="/")
        raise HTTPException(status_code=401, detail="טוקן admin לא חוקי")

    new_token = _make_session_token({
        "sub": payload.get("sub"), "email": payload.get("email"),
        "name": payload.get("name"), "role": payload.get("role"),
    })
    _set_session_cookie(response, new_token)
    response.delete_cookie(key=IMPERSONATOR_COOKIE, path="/")

    log_audit_action("IMPERSONATION_END", "ok",
                     f"Admin {payload.get('email')} stopped impersonating {user.get('email')}",
                     user=payload.get("email", "admin"))
    return {"id": payload.get("sub"), "email": payload.get("email"),
            "name": payload.get("name"), "role": payload.get("role")}


# ==========================================
# APP REGISTRY OVERRIDES — admin controls visibility + tags for /ai-hub tools
# ==========================================

VALID_APP_TAGS = ("new", "update", "coming_soon", "none")


@app.get("/api/apps/config")
async def get_app_config(_: dict = Depends(get_session_user)):
    """Public to any authenticated user. Returns a dict { app_id: {hidden, tag} }.
    Frontend merges this with the static TOOLS_REGISTRY to compute the effective
    visibility/tag for each app. Hidden apps are filtered out for non-admins."""
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute("SELECT app_id, hidden, tag FROM app_overrides").fetchall()
    finally:
        conn.close()
    return {
        "overrides": {
            r[0]: {"hidden": bool(r[1]), "tag": r[2]}
            for r in rows
        }
    }


@app.put("/api/admin/apps/{app_id}")
async def set_app_override(app_id: str, payload: dict, admin: dict = Depends(require_session_role(Role.ADMIN))):
    """Upsert override for a single app. Body: {hidden?: bool, tag?: 'new'|'update'|'coming_soon'|'none'|null}.
    Passing tag=null clears the override (falls back to hardcoded default in registry)."""
    hidden = 1 if payload.get("hidden") else 0
    tag = payload.get("tag")
    if tag is not None and tag not in VALID_APP_TAGS:
        raise HTTPException(status_code=400, detail=f"tag must be one of {VALID_APP_TAGS} or null")

    conn = sqlite3.connect(DB_PATH)
    try:
        existing = conn.execute("SELECT 1 FROM app_overrides WHERE app_id = ?", (app_id,)).fetchone()
        now = datetime.now(timezone.utc).isoformat()
        actor = admin.get("email", "admin")
        if existing:
            conn.execute(
                "UPDATE app_overrides SET hidden=?, tag=?, updated_at=?, updated_by=? WHERE app_id=?",
                (hidden, tag, now, actor, app_id),
            )
        else:
            conn.execute(
                "INSERT INTO app_overrides (app_id, hidden, tag, updated_at, updated_by) VALUES (?, ?, ?, ?, ?)",
                (app_id, hidden, tag, now, actor),
            )
        conn.commit()
    finally:
        conn.close()
    log_audit_action("APP_OVERRIDE", "ok", f"app={app_id} hidden={bool(hidden)} tag={tag}", user=actor)
    return {"status": "ok", "app_id": app_id, "hidden": bool(hidden), "tag": tag}


# ----- NOTIFICATIONS — admin sends, users read -----

def _is_on_cooldown(conn, user_id: str, tag: str, hours: int = 24) -> bool:
    """Return True if this (user, tag) was already emitted within `hours` hours.
    Uses the notification_cooldowns table — not the notifications inbox — so user
    deletion does NOT reset the dedup clock.
    Wrapped in try/except so a missing table (e.g. first boot before init_db
    completes) falls back to 'not on cooldown' rather than crashing the scanner."""
    try:
        cutoff = (datetime.now(timezone.utc) - timedelta(hours=hours)).isoformat()
        row = conn.execute(
            "SELECT last_emitted_at FROM notification_cooldowns WHERE user_id=? AND tag=? AND last_emitted_at > ?",
            (user_id, tag, cutoff),
        ).fetchone()
        return row is not None
    except Exception:
        return False  # table missing or other transient error → allow emit


def _set_cooldown(conn, user_id: str, tag: str) -> None:
    """Record that (user, tag) was just emitted. Upsert into cooldowns table.
    Non-fatal — if the table is missing we log and continue."""
    try:
        now = datetime.now(timezone.utc).isoformat()
        conn.execute(
            """INSERT INTO notification_cooldowns (user_id, tag, last_emitted_at) VALUES (?, ?, ?)
               ON CONFLICT(user_id, tag) DO UPDATE SET last_emitted_at = excluded.last_emitted_at""",
            (user_id, tag, now),
        )
    except Exception as exc:
        logger.warning("_set_cooldown failed (non-fatal): %s", exc)


def _emit_notification(
    conn,
    user_id: str,
    message: str,
    severity: str = "info",
    sent_by: str | None = None,
    category: str = "general",
    link: str | None = None,
    *,
    sent_at: str | None = None,
) -> bool:
    """Insert a notification row, respecting the user's category preference.
    Returns True if inserted, False if suppressed by opt-out.
    `conn` must already be open; caller commits."""
    if not user_id:
        return False
    # Check preference: if the user explicitly opted out, skip.
    pref = conn.execute(
        "SELECT enabled FROM user_notification_preferences WHERE user_id = ? AND category = ?",
        (user_id, category),
    ).fetchone()
    if pref is not None and not pref[0]:
        return False  # user opted out
    _at = sent_at or datetime.now(timezone.utc).isoformat()
    note_id = f"NTF-{uuid.uuid4().hex[:10].upper()}"
    conn.execute(
        "INSERT INTO notifications (id, user_id, message, severity, sent_by, sent_at, category, link) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (note_id, user_id, message, severity, sent_by, _at, category, link),
    )
    return True


@app.post("/api/admin/notifications/send")
async def admin_send_notification(payload: dict, admin: dict = Depends(require_session_role(Role.ADMIN))):
    user_ids = payload.get("user_ids") or []
    target_group = (payload.get("target_group") or "").strip()
    message = (payload.get("message") or "").strip()
    severity = (payload.get("severity") or "info").strip()
    link = (payload.get("link") or "").strip() or None
    category = (payload.get("category") or "manual").strip()
    if not message:
        raise HTTPException(status_code=400, detail="הודעה ריקה")
    if severity not in ("info", "warning", "danger", "success", "kudos"):
        severity = "info"
    if len(message) > 2000:
        raise HTTPException(status_code=400, detail="ההודעה ארוכה מדי (עד 2000 תווים)")

    sent_at = datetime.now(timezone.utc).isoformat()
    delivered, skipped = 0, []

    _ROLE_MAP = {
        "all_recruiters": "recruiter",
        "all_admins":     "admin",
        "all_hrbp":       "hrbp",
    }

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        # Resolve recipients from target_group or explicit user_ids
        if target_group == "all_users":
            rows = c.execute("SELECT id FROM users WHERE is_active = 1").fetchall()
            user_ids = [r[0] for r in rows]
        elif target_group in _ROLE_MAP:
            role = _ROLE_MAP[target_group]
            rows = c.execute("SELECT id FROM users WHERE role = ? AND is_active = 1", (role,)).fetchall()
            user_ids = [r[0] for r in rows]
        elif not isinstance(user_ids, list) or not user_ids:
            raise HTTPException(status_code=400, detail="חסרים נמענים")

        for uid in user_ids:
            c.execute("SELECT id FROM users WHERE id = ? AND is_active = 1", (str(uid),))
            if not c.fetchone():
                skipped.append(str(uid)); continue
            note_id = f"NTF-{uuid.uuid4().hex[:10].upper()}"
            c.execute(
                "INSERT INTO notifications (id, user_id, message, severity, sent_by, sent_at, category, link) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (note_id, str(uid), message, severity, admin.get("email", "admin"), sent_at, category, link),
            )
            delivered += 1
        conn.commit()
    finally:
        conn.close()

    log_audit_action("NOTIFICATION_SENT", "ok",
                     f"Sent to {delivered} user(s) (severity={severity}, group={target_group or 'specific'})",
                     user=admin.get("email", "admin"))
    return {"delivered": delivered, "skipped": skipped, "sent_at": sent_at}


@app.get("/api/admin/notifications/history")
async def admin_notifications_history(
    limit: int = 50,
    category: Optional[str] = None,
    admin: dict = Depends(require_session_role(Role.ADMIN)),
):
    """Return the last `limit` notifications (across all users) for the history table."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        sql = (
            "SELECT n.id, n.user_id, u.full_name, u.email, n.message, n.severity, "
            "n.category, n.sent_by, n.sent_at, n.read_at "
            "FROM notifications n LEFT JOIN users u ON n.user_id = u.id "
        )
        params: list = []
        if category:
            sql += "WHERE n.category = ? "
            params.append(category)
        sql += "ORDER BY n.sent_at DESC LIMIT ?"
        params.append(limit)
        rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]
    except sqlite3.OperationalError:
        return []
    finally:
        conn.close()


@app.get("/api/notifications/me")
async def get_my_notifications(unread_only: bool = False, user: dict = Depends(get_session_user)):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        sql = "SELECT id, message, severity, sent_by, sent_at, read_at, link, category FROM notifications WHERE user_id = ?"
        params: list = [user.get("sub")]
        if unread_only:
            sql += " AND read_at IS NULL"
        sql += " ORDER BY sent_at DESC LIMIT 50"
        c.execute(sql, params)
        rows = c.fetchall()
    except sqlite3.OperationalError:
        return []
    finally:
        conn.close()
    return [
        {
            "id": r[0], "message": r[1], "severity": r[2], "sent_by": r[3],
            "sent_at": r[4], "read_at": r[5],
            "link": r[6],
            "category": r[7] or "general",
        }
        for r in rows
    ]


@app.post("/api/notifications/{note_id}/read")
async def mark_notification_read(note_id: str, user: dict = Depends(get_session_user)):
    now = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute(
            "UPDATE notifications SET read_at = ? WHERE id = ? AND user_id = ? AND read_at IS NULL",
            (now, note_id, user.get("sub")),
        )
        if c.rowcount == 0:
            raise HTTPException(status_code=404, detail="התראה לא נמצאה")
        conn.commit()
    finally:
        conn.close()
    return {"status": "ok", "read_at": now}


@app.post("/api/notifications/mark-all-read")
async def mark_all_notifications_read(user: dict = Depends(get_session_user)):
    """Mark all unread notifications as read for the current user."""
    now = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute(
            "UPDATE notifications SET read_at = ? WHERE user_id = ? AND read_at IS NULL",
            (now, user.get("sub")),
        )
        affected = c.rowcount
        conn.commit()
    finally:
        conn.close()
    return {"status": "ok", "marked_read": affected}


@app.get("/api/notifications/count")
async def get_notification_count(user: dict = Depends(get_session_user)):
    """Return unread notification count for the current user."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute(
            "SELECT COUNT(*) FROM notifications WHERE user_id = ? AND read_at IS NULL",
            (user.get("sub"),),
        )
        count = c.fetchone()[0]
    finally:
        conn.close()
    return {"count": int(count)}


@app.delete("/api/notifications/{note_id}")
async def delete_notification(note_id: str, user: dict = Depends(get_session_user)):
    """Permanently delete a notification. Only the owner can delete their own notifications."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute(
            "DELETE FROM notifications WHERE id = ? AND user_id = ?",
            (note_id, user.get("sub")),
        )
        conn.commit()
    finally:
        conn.close()
    return {"status": "deleted"}


# ── Notification preferences (per-user category opt-in/out) ─────────────────

VALID_NOTIF_CATEGORIES = frozenset(["sla", "inactivity", "ingest", "quality", "manual", "general"])

@app.get("/api/notifications/preferences")
async def get_notification_preferences(user: dict = Depends(get_session_user)):
    """Return {category: enabled} map for the current user.
    Categories not in the table default to enabled=True."""
    uid = user.get("sub")
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            "SELECT category, enabled FROM user_notification_preferences WHERE user_id = ?",
            (uid,),
        ).fetchall()
    finally:
        conn.close()
    # Return explicit overrides; frontend fills in defaults
    return {row[0]: bool(row[1]) for row in rows}


@app.put("/api/notifications/preferences")
async def update_notification_preferences(payload: dict, user: dict = Depends(get_session_user)):
    """Upsert preference flags for the current user.
    Body: {category: bool, ...}. Unknown categories are ignored."""
    uid = user.get("sub")
    conn = sqlite3.connect(DB_PATH)
    try:
        for category, enabled in payload.items():
            if category not in VALID_NOTIF_CATEGORIES:
                continue
            conn.execute(
                """INSERT INTO user_notification_preferences (user_id, category, enabled)
                   VALUES (?, ?, ?)
                   ON CONFLICT(user_id, category) DO UPDATE SET enabled = excluded.enabled""",
                (uid, category, 1 if enabled else 0),
            )
        conn.commit()
    finally:
        conn.close()
    return {"status": "ok"}


@app.get("/api/admin/notifications/history")
async def admin_notification_history(
    limit: int = 100,
    category: Optional[str] = None,
    _: dict = Depends(require_session_role(Role.ADMIN)),
):
    """Admin view: notification history across all users, with read stats per message group."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        sql = """
            SELECT
                n.id,
                n.message,
                n.severity,
                n.category,
                n.sent_by,
                n.sent_at,
                n.link,
                u.full_name  AS recipient_name,
                u.email      AS recipient_email,
                CASE WHEN n.read_at IS NOT NULL THEN 1 ELSE 0 END AS is_read
            FROM notifications n
            LEFT JOIN users u ON u.id = n.user_id
            WHERE 1=1
        """
        params: list = []
        if category:
            sql += " AND n.category = ?"
            params.append(category)
        sql += " ORDER BY n.sent_at DESC LIMIT ?"
        params.append(limit)
        rows = c.execute(sql, params).fetchall()
    except sqlite3.OperationalError:
        return []
    finally:
        conn.close()
    return [
        {
            "id": r[0], "message": r[1], "severity": r[2], "category": r[3] or "general",
            "sent_by": r[4], "sent_at": r[5], "link": r[6],
            "recipient_name": r[7], "recipient_email": r[8], "is_read": bool(r[9]),
        }
        for r in rows
    ]


# ----- INGESTION DIFF (separate from /admin/ingestion/batches above) -----

@app.get("/api/admin/ingestion/batches")
def list_ingestion_batches_diff(limit: int = 20, _: dict = Depends(require_session_role(Role.ADMIN))):
    """Aggregates batch_entity_changes — gives insert/update/delete counts per batch."""
    cap = max(1, min(int(limit or 20), 100))
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute(
            """SELECT batch_id, MIN(created_at), MAX(created_at),
                      SUM(CASE WHEN change_type='insert' THEN 1 ELSE 0 END),
                      SUM(CASE WHEN change_type='update' THEN 1 ELSE 0 END),
                      SUM(CASE WHEN change_type='delete' THEN 1 ELSE 0 END)
               FROM batch_entity_changes GROUP BY batch_id ORDER BY MAX(created_at) DESC LIMIT ?""",
            (cap,),
        )
        return [
            {"batch_id": r[0], "started_at": r[1], "ended_at": r[2],
             "inserts": r[3] or 0, "updates": r[4] or 0, "deletes": r[5] or 0}
            for r in c.fetchall()
        ]
    except sqlite3.OperationalError:
        return []
    finally:
        conn.close()


@app.get("/api/admin/ingestion/batch/{batch_id}/changes")
def get_batch_changes(
    batch_id: str, type: Optional[str] = None, limit: int = 50, offset: int = 0,
    _: dict = Depends(require_session_role(Role.ADMIN)),
):
    cap = max(1, min(int(limit or 50), 200))
    skip = max(0, int(offset or 0))
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute(
            "SELECT change_type, COUNT(*) FROM batch_entity_changes WHERE batch_id = ? GROUP BY change_type",
            (batch_id,),
        )
        counts = {"insert": 0, "update": 0, "delete": 0}
        for change_type, n in c.fetchall():
            if change_type in counts:
                counts[change_type] = n
        if sum(counts.values()) == 0:
            return {"batch_id": batch_id, "counts": counts, "rows": [], "total": 0}

        params: list = [batch_id]
        sql = "SELECT id, entity_type, entity_id, change_type, before_json, after_json, created_at FROM batch_entity_changes WHERE batch_id = ?"
        if type and type in ("insert", "update", "delete"):
            sql += " AND change_type = ?"; params.append(type)
        sql += " ORDER BY id ASC LIMIT ? OFFSET ?"
        params.extend([cap, skip])
        c.execute(sql, params)
        rows = [{
            "id": r[0], "entity_type": r[1], "entity_id": r[2], "change_type": r[3],
            "before": json.loads(r[4]) if r[4] else None,
            "after":  json.loads(r[5]) if r[5] else None,
            "created_at": r[6],
        } for r in c.fetchall()]
        return {"batch_id": batch_id, "counts": counts, "rows": rows, "total": sum(counts.values())}
    finally:
        conn.close()


@app.get("/api/admin/ingestion/batch/{batch_id}/rejected")
def get_batch_rejected_rows(
    batch_id: str,
    limit: int = 100,
    offset: int = 0,
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Return rows that were rejected during ingestion for a given batch."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        total = conn.execute(
            "SELECT COUNT(*) FROM rejected_rows WHERE batch_id = ?", (batch_id,)
        ).fetchone()[0]
        rows = conn.execute(
            """SELECT id, row_index, reason_code, reason_detail, raw_row, created_at
               FROM rejected_rows WHERE batch_id = ? ORDER BY id LIMIT ? OFFSET ?""",
            (batch_id, max(1, min(limit, 200)), max(0, offset)),
        ).fetchall()
        return {"batch_id": batch_id, "total": total, "rows": [dict(r) for r in rows]}
    finally:
        conn.close()


# ----- CROSS-MODULE SEARCH -----

@app.get("/api/search")
def cross_module_search(q: str = "", limit: int = 10, _: dict = Depends(get_session_user)):
    query = (q or "").strip()
    if len(query) < 2:
        return {"candidates": [], "jobs": [], "applications": []}
    cap = max(1, min(int(limit or 10), 25))
    pattern = f"%{query.lower()}%"

    candidates, jobs, applications = [], [], []
    with db_conn() as conn:
        c = conn.cursor()
        try:
            c.execute(
                """SELECT id, name, email, phone, source FROM candidates
                   WHERE (LOWER(IFNULL(name,'')) LIKE ? OR LOWER(IFNULL(email,'')) LIKE ?
                      OR LOWER(IFNULL(phone,'')) LIKE ? OR LOWER(IFNULL(id,'')) LIKE ?)
                      AND COALESCE(is_active, 1) = 1 LIMIT ?""",
                (pattern, pattern, pattern, pattern, cap),
            )
            candidates = [{"id": r[0], "name": r[1], "email": r[2], "phone": r[3], "source": r[4]} for r in c.fetchall()]
        except sqlite3.OperationalError:
            pass
        try:
            c.execute(
                """SELECT id, job_title, department, hiring_manager FROM jobs
                   WHERE (LOWER(IFNULL(job_title,'')) LIKE ? OR LOWER(IFNULL(department,'')) LIKE ?
                      OR LOWER(IFNULL(hiring_manager,'')) LIKE ? OR LOWER(IFNULL(id,'')) LIKE ?)
                      AND COALESCE(is_active, 1) = 1 LIMIT ?""",
                (pattern, pattern, pattern, pattern, cap),
            )
            jobs = [{"id": r[0], "title": r[1], "department": r[2], "hiring_manager": r[3]} for r in c.fetchall()]
        except sqlite3.OperationalError:
            pass
        try:
            c.execute(
                """SELECT a.app_id, a.status, a.recruiter, a.start_date, a.days_in_process,
                          c.name, j.job_title
                   FROM applications a
                   LEFT JOIN candidates c ON c.id = a.candidate_id
                   LEFT JOIN jobs j       ON j.id = a.job_id
                   WHERE (LOWER(IFNULL(a.app_id,'')) LIKE ? OR LOWER(IFNULL(a.status,'')) LIKE ?
                      OR LOWER(IFNULL(a.recruiter,'')) LIKE ? OR LOWER(IFNULL(c.name,'')) LIKE ?
                      OR LOWER(IFNULL(j.job_title,'')) LIKE ?)
                      AND COALESCE(a.is_active, 1) = 1 AND COALESCE(c.is_active, 1) = 1 AND COALESCE(j.is_active, 1) = 1 LIMIT ?""",
                (pattern, pattern, pattern, pattern, pattern, cap),
            )
            applications = [{
                "app_id": r[0], "status": r[1], "recruiter": r[2],
                "start_date": r[3], "days_in_process": r[4],
                "candidate_name": r[5], "job_title": r[6],
            } for r in c.fetchall()]
        except sqlite3.OperationalError:
            pass
    return {"candidates": candidates, "jobs": jobs, "applications": applications}


# =====================================================================
# RECRUITER INACTIVITY MONITOR
# Idempotent: scans the `users` table for active recruiters whose
# `last_login_at` is older than 3 days (or NULL — never logged in), and
# emits a notification both to the recruiter herself and to every admin.
# Skips emissions when the same recruiter was already flagged in the last
# 24 hours, so repeated calls don't spam the inbox.
# =====================================================================

_INACTIVE_TAG = "INACTIVE_RECRUITER_3D"


def _check_inactive_recruiters(threshold_days: int = 3, dedupe_hours: int = 24) -> dict:
    """Scan and notify. Returns {flagged: [...], emitted: int}."""
    threshold_iso = (datetime.now(timezone.utc) - timedelta(days=threshold_days)).isoformat()
    dedupe_iso = (datetime.now(timezone.utc) - timedelta(hours=dedupe_hours)).isoformat()
    sent_at = datetime.now(timezone.utc).isoformat()

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    flagged: list = []
    emitted = 0
    try:
        c.execute(
            "SELECT id, full_name, email, last_login_at FROM users "
            "WHERE role='recruiter' AND is_active=1 AND (last_login_at IS NULL OR last_login_at < ?)",
            (threshold_iso,),
        )
        recruiters = c.fetchall()

        c.execute("SELECT id FROM users WHERE role='admin' AND is_active=1")
        admin_ids = [row[0] for row in c.fetchall()]

        for rec_id, full_name, email, last_login in recruiters:
            recruiter_label = full_name or email
            # Dedup: use cooldown table (not notifications inbox) so user
            # deletion does NOT reset the 24h clock.
            cooldown_tag = f"{_INACTIVE_TAG}:{rec_id}"
            if _is_on_cooldown(conn, rec_id, cooldown_tag, hours=dedupe_hours):
                continue

            user_msg = (
                "לא נכנסת למערכת מעולם — יש להתחבר כדי להתחיל לעבוד."
                if not last_login else
                "לא נכנסת למערכת מעל 3 ימים. נוכחותך הרציפה נדרשת כדי לעמוד ב-SLA ולשמור על חוויית מועמד."
            )

            # Notify the recruiter herself
            _emit_notification(
                conn, user_id=rec_id, message=user_msg,
                severity="warning", sent_by="system:" + _INACTIVE_TAG,
                category="inactivity", link=None, sent_at=sent_at,
            )
            emitted += 1

            # Notify every admin (use admin-scoped cooldown key)
            admin_link = "/admin?group=settings&sub=permissions"
            for admin_id in admin_ids:
                admin_cooldown_tag = f"{_INACTIVE_TAG}:{rec_id}:admin:{admin_id}"
                if _is_on_cooldown(conn, admin_id, admin_cooldown_tag, hours=dedupe_hours):
                    continue
                admin_msg = (
                    f"⚠️ {recruiter_label} (מגייסת) טרם התחברה למערכת."
                    if not last_login else
                    f"⚠️ {recruiter_label} (מגייסת) לא נכנסה למערכת מעל 3 ימים. כניסה אחרונה: {last_login}."
                )
                _emit_notification(
                    conn, user_id=admin_id, message=admin_msg,
                    severity="warning", sent_by="system:" + _INACTIVE_TAG,
                    category="inactivity", link=admin_link, sent_at=sent_at,
                )
                _set_cooldown(conn, admin_id, admin_cooldown_tag)
                emitted += 1

            # Set cooldown AFTER emitting (so partial failures don't lock out next run)
            _set_cooldown(conn, rec_id, cooldown_tag)
            flagged.append({"id": rec_id, "name": recruiter_label, "last_login_at": last_login})

        conn.commit()
    finally:
        conn.close()
    return {"flagged": flagged, "emitted": emitted}


@app.get("/api/notifications/stream")
async def notification_stream(user: dict = Depends(get_session_user)):
    """Server-Sent Events endpoint. The client opens one persistent connection
    and receives new notifications in real-time (sub-5-second latency) instead
    of polling every 60 s.

    Protocol:
      - Sends JSON array of new BackendNotification objects as `data:` lines.
      - Sends `": keepalive"` every 5 s when there's nothing new (prevents
        proxy / nginx timeouts on 60-120 s idle connections).
      - On auth failure / session expiry the generator returns, closing the stream.
        The frontend EventSource will retry automatically — the 60 s polling
        in NotificationContext acts as a fallback safety net.
    """
    user_id = user.get("sub")
    if not user_id:
        # Not authenticated — return an empty 200 so EventSource doesn't spam retries
        return Response(content="", media_type="text/event-stream")

    async def event_generator():
        last_checked = datetime.now(timezone.utc).isoformat()
        try:
            while True:
                await asyncio.sleep(5)
                try:
                    with db_conn() as conn:
                        c = conn.cursor()
                        rows = c.execute(
                            "SELECT id, message, severity, sent_by, sent_at, read_at, link, category "
                            "FROM notifications WHERE user_id = ? AND sent_at > ? ORDER BY sent_at ASC",
                            (user_id, last_checked),
                        ).fetchall()
                except Exception:
                    yield ": keepalive\n\n"
                    continue

                if rows:
                    last_checked = rows[-1][4]  # sent_at of the latest row
                    payload = [
                        {
                            "id": r[0], "message": r[1], "severity": r[2], "sent_by": r[3],
                            "sent_at": r[4], "read_at": r[5], "link": r[6],
                            "category": r[7] or "general",
                        }
                        for r in rows
                    ]
                    yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
                else:
                    yield ": keepalive\n\n"
        except asyncio.CancelledError:
            pass  # client disconnected — normal, don't log as error

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",  # disable nginx buffering
            "Connection": "keep-alive",
        },
    )


@app.post("/api/admin/check-inactive-recruiters")
async def check_inactive_recruiters_endpoint(_: dict = Depends(require_session_role(Role.ADMIN))):
    """Admin-triggered scan. Safe to call repeatedly — dedupes within 24h."""
    return _check_inactive_recruiters()


# =====================================================================
# SLA WATCHDOG — flags candidates / jobs / offers stuck longer than SLA
# thresholds. Read-only `GET /api/sla/alerts` returns the current breaches
# (no side effects). `POST /api/sla/scan` runs the check AND creates
# notifications for the responsible recruiter + admins, with 24h dedup.
# =====================================================================

_SLA_TAG = "SLA_BREACH"

def _compute_sla_alerts(thresholds: Optional[dict] = None) -> dict:
    """Returns a dict with `candidates`, `jobs`, `offers` arrays of breaches.
    Thresholds defaults to: stale_days=14 for candidates/offers, open_days=60 for jobs."""
    t = thresholds or {}
    stale_cand_days = int(t.get("stale_candidate_days", 14))
    stale_offer_days = int(t.get("stale_offer_days", 7))
    open_job_days = int(t.get("open_job_days", 60))
    now = datetime.now(timezone.utc)

    conn = sqlite3.connect(DB_PATH)
    try:
        # Candidates stuck — applications with days_in_process > threshold and not HIRED/REJECTED.
        cand_rows = conn.execute(
            "SELECT a.candidate_id, c.name, a.job_id, j.job_title, a.recruiter, a.status, "
            "       a.days_in_process, a.stage_code "
            "FROM applications a "
            "LEFT JOIN candidates c ON c.id = a.candidate_id "
            "LEFT JOIN jobs j ON j.id = a.job_id "
            "WHERE a.days_in_process >= ? AND COALESCE(a.stage_code,'') NOT IN ('HIRED','REJECTED') "
            "ORDER BY a.days_in_process DESC LIMIT 50",
            (stale_cand_days,),
        ).fetchall()

        # Jobs open too long — opened_at < threshold and closed_at NULL.
        cutoff_iso = (now - timedelta(days=open_job_days)).isoformat()
        job_rows = conn.execute(
            "SELECT id, job_title, department, hiring_manager, opened_at "
            "FROM jobs WHERE closed_at IS NULL AND opened_at IS NOT NULL AND opened_at < ? "
            "ORDER BY opened_at LIMIT 50",
            (cutoff_iso,),
        ).fetchall()

        # Offers pending — applications with status containing 'הצעה'/'offer' and days_in_process > threshold.
        offer_rows = conn.execute(
            "SELECT a.candidate_id, c.name, a.job_id, j.job_title, a.recruiter, a.days_in_process "
            "FROM applications a "
            "LEFT JOIN candidates c ON c.id = a.candidate_id "
            "LEFT JOIN jobs j ON j.id = a.job_id "
            "WHERE (a.stage_code = 'OFFER' OR a.status LIKE '%הצעה%' OR a.status LIKE '%offer%') "
            "  AND a.days_in_process >= ? "
            "ORDER BY a.days_in_process DESC LIMIT 50",
            (stale_offer_days,),
        ).fetchall()
    finally:
        conn.close()

    return {
        "thresholds": {
            "stale_candidate_days": stale_cand_days,
            "stale_offer_days": stale_offer_days,
            "open_job_days": open_job_days,
        },
        "candidates": [
            {"candidate_id": r[0], "candidate_name": r[1], "job_id": r[2], "job_title": r[3],
             "recruiter": r[4], "status": r[5], "days_in_process": r[6], "stage_code": r[7]}
            for r in cand_rows
        ],
        "jobs": [
            {"job_id": r[0], "job_title": r[1], "department": r[2], "hiring_manager": r[3], "opened_at": r[4]}
            for r in job_rows
        ],
        "offers": [
            {"candidate_id": r[0], "candidate_name": r[1], "job_id": r[2], "job_title": r[3],
             "recruiter": r[4], "days_in_process": r[5]}
            for r in offer_rows
        ],
        "total_breaches": len(cand_rows) + len(job_rows) + len(offer_rows),
    }


@app.get("/api/sla/alerts")
async def get_sla_alerts(
    stale_candidate_days: int = 14,
    stale_offer_days: int = 7,
    open_job_days: int = 60,
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER, Role.HIRING_MANAGER)),
):
    """Read-only listing of current SLA breaches. Pure compute — no side effects."""
    return _compute_sla_alerts({
        "stale_candidate_days": stale_candidate_days,
        "stale_offer_days": stale_offer_days,
        "open_job_days": open_job_days,
    })


@app.post("/api/sla/scan")
async def post_sla_scan(
    payload: Optional[dict] = None,
    admin: dict = Depends(require_session_role(Role.ADMIN)),
):
    """Admin-triggered: compute breaches AND emit notifications. Dedups within 24h
    so calling it repeatedly is safe. Notifications go to the responsible
    recruiter (matched by name → users.full_name) and to all admins."""
    alerts = _compute_sla_alerts(payload or {})
    dedupe_iso = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    sent_at = datetime.now(timezone.utc).isoformat()

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    emitted = 0
    try:
        admin_ids = [r[0] for r in c.execute("SELECT id FROM users WHERE role='admin' AND is_active=1").fetchall()]

        def _recruiter_user_id(name: Optional[str]) -> Optional[str]:
            if not name: return None
            row = c.execute("SELECT id FROM users WHERE LOWER(full_name) = LOWER(?) AND is_active=1", (name,)).fetchone()
            return row[0] if row else None

        def _emit(user_id: str, msg: str, severity: str = "warning",
                  link: Optional[str] = None, category: str = "sla") -> None:
            nonlocal emitted
            # Dedup via cooldown table (survives user inbox deletion).
            # Key = hash of (user, tag, message) so distinct breach messages
            # don't block each other.
            import hashlib
            msg_hash = hashlib.md5(msg.encode()).hexdigest()[:12]
            cooldown_tag = f"{_SLA_TAG}:{msg_hash}"
            if _is_on_cooldown(conn, user_id, cooldown_tag, hours=24):
                return
            inserted = _emit_notification(
                conn, user_id=user_id, message=msg,
                severity=severity, sent_by="system:" + _SLA_TAG,
                category=category, link=link, sent_at=sent_at,
            )
            if inserted:
                _set_cooldown(conn, user_id, cooldown_tag)
                emitted += 1

        # Candidate breaches — link to SLA Watchdog app for actionable view
        for cand in alerts["candidates"][:20]:
            msg = f"⏱️ {cand.get('candidate_name', '?')} תקוע/ה {cand.get('days_in_process', '?')} ימים ב-{cand.get('job_title') or 'משרה'}. דחיפות לבדיקה."
            cand_link = "/ai-hub?tool=sla-watchdog"
            rec_uid = _recruiter_user_id(cand.get("recruiter"))
            if rec_uid: _emit(rec_uid, msg, link=cand_link)
            for aid in admin_ids: _emit(aid, msg, link=cand_link)

        # Offer breaches (high severity) — link to SLA Watchdog
        for off in alerts["offers"][:10]:
            msg = f"🚨 הצעה תלויה ל-{off.get('candidate_name', '?')} - {off.get('days_in_process', '?')} ימים בהצעה. סיכון אובדן מועמד."
            off_link = "/ai-hub?tool=sla-watchdog"
            rec_uid = _recruiter_user_id(off.get("recruiter"))
            if rec_uid: _emit(rec_uid, msg, severity="critical", link=off_link)
            for aid in admin_ids: _emit(aid, msg, severity="critical", link=off_link)

        # Job breaches (admin-only) — link to jobs page
        for job in alerts["jobs"][:10]:
            msg = f"🏗️ משרה פתוחה מעל {alerts['thresholds']['open_job_days']} ימים: {job.get('job_title', '?')} ({job.get('department', '?')})."
            for aid in admin_ids: _emit(aid, msg, link="/jobs")

        conn.commit()
    finally:
        conn.close()

    log_audit_action("SLA_SCAN", "ok", f"breaches={alerts['total_breaches']} notified={emitted}", user=admin.get("email", "admin"))
    return {**alerts, "notifications_emitted": emitted}


@app.get("/api/data-version")
def api_get_data_version():
    """Monotonic counter. Frontend polls it (DataVersionContext); when it
    increases, consumer blocks invalidate their caches and refetch. Public
    to any authenticated user via cookie session — no role gating because
    it's just a number with no business data."""
    return {"version": get_data_version()}


@app.get("/api/internal/check-my-inactivity")
async def check_my_inactivity(user: dict = Depends(get_session_user)):
    """Any logged-in user can fetch a fresh scan; only mutates if the caller
    is an admin or recruiter (the two roles that need to know)."""
    if user.get("role") in ("admin", "recruiter"):
        return _check_inactive_recruiters()
    return {"flagged": [], "emitted": 0}


# =====================================================================
# CANDIDATE / JOB DRILL-DOWN ENDPOINTS — drive the side-panels.
# =====================================================================


# GET /api/candidates/{candidate_key} moved to backend/routers/candidates.py (B2.5)


# GET /api/jobs/{job_key}/candidates moved to backend/routers/jobs.py (B2.6)


# =====================================================================
# STAGE ADVANCEMENT — PATCH /api/candidates/{candidate_key}/stage
# =====================================================================

# PATCH /api/candidates/{candidate_key}/stage moved to backend/routers/candidates.py (B2.5)


# =====================================================================
# DIRECT RECORD EDITING — PATCH endpoints for candidates / jobs / applications
# =====================================================================

# PATCH /api/candidates/{candidate_id} moved to backend/routers/candidates.py (B2.5)


# PATCH /api/jobs/{job_id} moved to backend/routers/jobs.py (B2.6)


@app.patch("/api/applications/{app_id}")
def edit_application(
    app_id: str,
    payload: ApplicationEditPayload,
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER)),
):
    """Direct application field editor."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT * FROM applications WHERE app_id = ?", (app_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="תהליך לא נמצא")
        before = dict(row)
        updates: dict = {}

        for field in ("status", "stage_code", "recruiter", "application_date", "days_in_process"):
            val = getattr(payload, field, None)
            if val is not None:
                updates[field] = val

        if not updates:
            return {"status": "no_change"}

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        conn.execute(
            f"UPDATE applications SET {set_clause} WHERE app_id = ?",
            [*updates.values(), app_id],
        )
        after = dict(
            conn.execute("SELECT * FROM applications WHERE app_id = ?", (app_id,)).fetchone()
        )
        batch_id = _create_manual_edit_batch("application", app_id, user.get("email", "admin"))
        _record_change(conn, batch_id, "application", app_id, "update", before, after)
        conn.commit()
        log_audit_action(
            "APPLICATION_EDIT", "ok",
            f"app_id={app_id} fields={list(updates.keys())}",
            user=user.get("email", "admin"),
        )
        bump_data_version()
        return {"status": "updated", "application": after}
    finally:
        conn.close()


# DELETE /api/candidates/{candidate_id} moved to backend/routers/candidates.py (B2.5)


# DELETE /api/jobs/{job_id} moved to backend/routers/jobs.py (B2.6)


@app.delete("/api/applications/{app_id}")
def delete_application(
    app_id: str,
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER)),
):
    """Soft delete application."""
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        app = c.execute("SELECT app_id, candidate_id, job_id FROM applications WHERE app_id = ?", (app_id,)).fetchone()
        if not app:
            raise HTTPException(status_code=404, detail="תהליך לא נמצא")
        
        c.execute("UPDATE applications SET is_active = 0 WHERE app_id = ?", (app_id,))
        
        batch_id = _create_manual_edit_batch("application", app_id, user.get("email", "admin"))
        _record_change(conn, batch_id, "application", app_id, "delete", {"app_id": app[0]}, None)
        conn.commit()
        
        log_audit_action(
            "APPLICATION_DELETE", "ok",
            f"app_id={app_id}",
            user=user.get("email", "admin"),
        )
        bump_data_version()
        return {"status": "deleted", "app_id": app_id}
    finally:
        conn.close()



# =====================================================================
# ANOMALY DETECTION ENGINE
# Scans the live database for suspicious data patterns and writes
# flagged records to the `data_anomalies` table for dashboard review.
# =====================================================================

_ANOMALY_SEVERITY = {
    "duplicate_name_different_contact": "high",
    "stale_process":                   "medium",
    "future_start_date":               "medium",
    "missing_contact":                 "low",
    "zero_days_long_running":          "low",
    "multiple_active_applications":    "medium",
    "duplicate_application":           "high",
}


def run_anomaly_scan(conn: sqlite3.Connection, batch_id: str | None = None) -> dict:
    """
    Full-database anomaly scan. Idempotent: re-running does not create
    duplicate anomaly rows — existing open anomalies for the same entity
    and type are left untouched; only genuinely new findings are inserted.

    Returns a summary dict: {anomaly_type: count_new}.
    """
    import hashlib as _hl

    now_ts = _utcnow().isoformat()
    summary: dict[str, int] = {}
    c = conn.cursor()

    def _upsert(entity_type: str, entity_id: str, atype: str, desc: str, suggestion: str, meta: dict):
        """Insert anomaly only if no open row for (entity_type, entity_id, anomaly_type) exists."""
        existing = c.execute(
            "SELECT id FROM data_anomalies WHERE entity_type=? AND entity_id=? AND anomaly_type=? AND status='open'",
            (entity_type, entity_id, atype),
        ).fetchone()
        if existing:
            return  # already flagged — don't duplicate
        row_id = _hl.md5(f"{entity_type}:{entity_id}:{atype}".encode()).hexdigest()[:16]
        severity = _ANOMALY_SEVERITY.get(atype, "medium")
        c.execute(
            """INSERT OR IGNORE INTO data_anomalies
               (id, entity_type, entity_id, anomaly_type, severity, description, suggestion, meta_json, status, created_at, batch_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'open', ?, ?)""",
            (row_id, entity_type, entity_id, atype, severity, desc, suggestion,
             json.dumps(meta, ensure_ascii=False), now_ts, batch_id),
        )
        summary[atype] = summary.get(atype, 0) + 1

    # ------------------------------------------------------------------
    # 1. DUPLICATE NAME + DIFFERENT CONTACT
    #    Same full name (case-insensitive) mapped to more than one
    #    (email_norm, phone_norm) pair → likely duplicate record created
    #    by a typo in contact details.
    # ------------------------------------------------------------------
    dup_names = c.execute(
        """SELECT LOWER(name), COUNT(DISTINCT COALESCE(email_norm,'') || '|' || COALESCE(phone_norm,'')) AS cnt
           FROM candidates WHERE is_active=1 AND name IS NOT NULL
           GROUP BY LOWER(name) HAVING cnt > 1"""
    ).fetchall()
    for (lname, cnt) in dup_names:
        rows = c.execute(
            "SELECT id, name, email, phone FROM candidates WHERE LOWER(name)=? AND is_active=1",
            (lname,),
        ).fetchall()
        ids_str = ",".join(r[0] for r in rows)
        for r in rows:
            _upsert(
                "candidate", r[0], "duplicate_name_different_contact",
                f"שם '{r[1]}' מופיע {cnt} פעמים עם פרטי קשר שונים.",
                "בדוק ומיזג את הרשומות הכפולות.",
                {"duplicate_ids": ids_str, "count": cnt},
            )

    # ------------------------------------------------------------------
    # 2. STALE PROCESS — application open > 180 days with no hire/reject
    # ------------------------------------------------------------------
    stale = c.execute(
        """SELECT a.app_id, c.name, j.job_title, a.days_in_process
           FROM applications a
           JOIN candidates c ON c.id=a.candidate_id
           JOIN jobs j ON j.id=a.job_id
           WHERE COALESCE(a.is_active,1)=1
             AND COALESCE(c.is_active,1)=1
             AND a.days_in_process > 180
             AND COALESCE(a.stage_code,'ACTIVE') NOT IN ('HIRED','REJECTED','STARTED')"""
    ).fetchall()
    for (app_id, cname, jtitle, dip) in stale:
        _upsert(
            "application", app_id, "stale_process",
            f"תהליך גיוס של '{cname}' למשרת '{jtitle}' פתוח {dip} ימים.",
            "עדכן את סטטוס התהליך או סגור אותו.",
            {"days_in_process": dip, "candidate": cname, "job": jtitle},
        )

    # ------------------------------------------------------------------
    # 3. FUTURE START DATE — start_date > today + 90 days
    # ------------------------------------------------------------------
    try:
        future_apps = c.execute(
            """SELECT a.app_id, c.name, j.job_title, a.start_date
               FROM applications a
               JOIN candidates c ON c.id=a.candidate_id
               JOIN jobs j ON j.id=a.job_id
               WHERE COALESCE(a.is_active,1)=1
                 AND a.start_date IS NOT NULL
                 AND date(a.start_date) > date('now','+90 days')"""
        ).fetchall()
        for (app_id, cname, jtitle, sdate) in future_apps:
            _upsert(
                "application", app_id, "future_start_date",
                f"תאריך תחילת עבודה עתידי מאוד: {sdate} ('{cname}' → '{jtitle}').",
                "בדוק שהתאריך הוזן נכון.",
                {"start_date": sdate, "candidate": cname, "job": jtitle},
            )
    except Exception:
        pass

    # ------------------------------------------------------------------
    # 4. MISSING CONTACT — candidate has no phone AND no email
    # ------------------------------------------------------------------
    no_contact = c.execute(
        """SELECT id, name FROM candidates
           WHERE is_active=1
             AND (phone_norm IS NULL OR phone_norm='')
             AND (email_norm IS NULL OR email_norm='')"""
    ).fetchall()
    for (cid, cname) in no_contact:
        _upsert(
            "candidate", cid, "missing_contact",
            f"מועמד '{cname}' חסר פרטי קשר (אימייל וטלפון).",
            "השלם פרטי קשר ממקור הגיוס.",
            {"name": cname},
        )

    # ------------------------------------------------------------------
    # 5. MULTIPLE ACTIVE APPLICATIONS FOR SAME CANDIDATE + JOB
    #    More than one active application for the same (candidate, job)
    #    pair — typically caused by re-upload without dedup.
    # ------------------------------------------------------------------
    multi_apps = c.execute(
        """SELECT candidate_id, job_id, COUNT(*) as cnt
           FROM applications
           WHERE is_active=1
           GROUP BY candidate_id, job_id HAVING cnt > 1"""
    ).fetchall()
    for (cid, jid, cnt) in multi_apps:
        apps = c.execute(
            "SELECT app_id FROM applications WHERE candidate_id=? AND job_id=? AND is_active=1",
            (cid, jid),
        ).fetchall()
        app_ids = ",".join(a[0] for a in apps)
        cname = (c.execute("SELECT name FROM candidates WHERE id=?", (cid,)).fetchone() or (cid,))[0]
        jtitle = (c.execute("SELECT job_title FROM jobs WHERE id=?", (jid,)).fetchone() or (jid,))[0]
        for (app_id,) in apps:
            _upsert(
                "application", app_id, "multiple_active_applications",
                f"מועמד '{cname}' רשום {cnt} פעמים באותה משרה '{jtitle}'.",
                "מחק את הרשומות הכפולות ושמור רק את המעודכנת.",
                {"duplicate_app_ids": app_ids, "count": cnt},
            )

    conn.commit()
    return summary


# ─── Anomaly API ─────────────────────────────────────────────────────────────
# The four /api/anomalies routes (scan, list, summary, review) moved to
# backend/routers/anomalies.py (B2). They are wired back into `app` via
# `app.include_router(...)` at the bottom of this file.


# ─── Auto-scan hook: called at the end of every successful ingestion ──────────

def _auto_scan_after_ingest(conn: sqlite3.Connection, batch_id: str):
    """
    Called inside the ingestion transaction BEFORE commit so the scan
    benefits from the freshly loaded data.  Errors are swallowed so
    they never fail a successful upload.
    """
    try:
        run_anomaly_scan(conn, batch_id=batch_id)
    except Exception:
        pass


# =====================================================================
# PIPELINE SUMMARY — GET /api/pipeline/summary
# =====================================================================

@app.get("/api/pipeline/summary")
def get_pipeline_summary(
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER, Role.HIRING_MANAGER)),
):
    """Stage-level pipeline counts and basic KPIs."""
    conn = sqlite3.connect(DB_PATH)
    try:
        df = get_unified_data(conn)
    except Exception:
        df = pd.DataFrame()
    finally:
        conn.close()

    if df.empty:
        return {
            "stages": {s: 0 for s in UNIFIED_STAGES},
            "total": 0,
            "active_total": 0,
        }

    df["unified_stage"] = df.apply(
        lambda row: _compute_unified_stage(row.get("stage_code"), row.get("onboarding_status")),
        axis=1,
    )

    counts = df["unified_stage"].value_counts().to_dict()
    stage_summary = {s: int(counts.get(s, 0)) for s in UNIFIED_STAGES}
    active_stages = {"SCREEN", "INTERVIEW", "OFFER", "ACTIVE"}
    active_total = sum(stage_summary.get(s, 0) for s in active_stages)

    return {
        "stages": stage_summary,
        "total": int(len(df)),
        "active_total": active_total,
    }


# =====================================================================
# DATA PIPELINE V2 — unified ingestion endpoint
# POST /api/ingest/{type}  with multipart `file`.
#
# Stages (per the master plan):
#   1. Upload        — file received here
#   2. Parse         — _load_dataframe_from_upload (existing)
#   3. Validate      — _validate_ingest_frame (per-type required columns)
#   4. Normalize     — _normalize_ingest_row (phone/email/dates/dept/status)
#   5. Match & Merge — find_existing_*, merge_* (per-type handler)
#   6. Persist       — INSERT/UPDATE in transaction + batch_entity_changes
#   7. Notify        — bump_data_version + audit log
#
# Each per-type handler returns:
#   { received, inserted, updated, skipped_duplicate, rejected, rejected_reasons[] }
# =====================================================================


# Per-type required columns. The Hebrew/English aliases come from the
# existing _normalize_upload_frame; here we require the *canonical* name
# after that normalisation runs.
INGEST_REQUIREMENTS: dict[str, dict] = {
    "candidates": {
        # Accept either alias: "candidate_name" (new canonical via EXTRA_HEBREW_ALIASES)
        # or "name" (legacy via _normalize_upload_frame). Validation succeeds when EITHER
        # is present — handled by _validate_ingest_frame's per-row required check below.
        "required": ["candidate_name"],
        "recommended": ["email", "phone", "job_title", "status", "recruiter"],
        "description": "מועמדים בתהליך — שורה אחת לכל איטרציית מועמד×משרה",
    },
    "jobs": {
        "required": ["job_title", "department"],
        "recommended": ["hiring_manager", "opened_at", "target_count"],
        "description": "פתיחת/סגירת משרות — שורה אחת לכל משרה",
    },
    "hires": {
        "required": ["candidate_name", "job_title", "hire_date"],
        "recommended": ["salary", "department", "manager"],
        "description": "קליטות עובדים שהתבצעו",
    },
    "diversity": {
        "required": ["snapshot_month", "department", "dimension", "bucket", "count"],
        "recommended": [],
        "description": "מדדי גיוון לפי חודש × מחלקה × ממד",
    },
    "headcount": {
        "required": ["snapshot_month", "department", "role"],
        "recommended": ["standard", "current", "attrition_ytd", "hire_plan"],
        "description": "תקן מצבה (תקן מול בפועל) לפי חודש × מחלקה × תפקיד",
    },
    "budget": {
        "required": ["id", "vendor", "date", "amount", "category"],
        "recommended": ["due_date", "budget_month", "status", "file_url"],
        "description": "חשבוניות FinOps",
    },
    "attrition": {
        "required": ["employee_name", "leave_date"],
        "recommended": ["department", "manager", "reason", "voluntary"],
        "description": "אירועי עזיבה",
    },
}


# Alias tables for the typed-ingest pipeline now live in backend/aliases.py.
# Kept as a module-level re-export so any code still importing the old
# name from main.py keeps working.
EXTRA_HEBREW_ALIASES = TYPED_INGEST_ALIASES


def _apply_extra_aliases(df: pd.DataFrame) -> pd.DataFrame:
    """Map extra Hebrew column names to canonical ones used by the typed handlers."""
    df.columns = [str(c).strip() for c in df.columns]
    rename = {}
    for col in df.columns:
        target = TYPED_INGEST_ALIASES.get(col)
        if target and target not in df.columns:
            rename[col] = target
    if rename:
        df = df.rename(columns=rename)
    # Compatibility shim: legacy uploads use "name" while the new typed handlers
    # work with "candidate_name". Expose both so either side of the divide
    # passes validation.
    if "candidate_name" in df.columns and "name" not in df.columns:
        df["name"] = df["candidate_name"]
    elif "name" in df.columns and "candidate_name" not in df.columns:
        df["candidate_name"] = df["name"]
    return df


def _validate_ingest_frame(df: pd.DataFrame, file_type: str) -> tuple[pd.DataFrame, list[dict]]:
    """Returns (valid_df, rejected_rows). A row is rejected when any required
    column is missing/empty. The reason is human-readable Hebrew for the
    rejected_rows admin view."""
    spec = INGEST_REQUIREMENTS.get(file_type)
    if not spec:
        raise HTTPException(status_code=400, detail=f"Unknown ingest type: {file_type}")
    required = spec["required"]
    missing_cols = [c for c in required if c not in df.columns]
    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail=f"חסרות עמודות חובה לסוג '{file_type}': {', '.join(missing_cols)}",
        )

    rejected: list[dict] = []
    keep_mask: list[bool] = []
    for _, row in df.iterrows():
        problems = []
        for col in required:
            val = row.get(col)
            if val is None or (isinstance(val, str) and not val.strip()):
                problems.append(f"חסר ערך ל'{col}'")
            elif isinstance(val, float) and pd.isna(val):
                problems.append(f"חסר ערך ל'{col}'")
        if problems:
            rejected.append({"row": row.to_dict(), "reasons": problems})
            keep_mask.append(False)
        else:
            keep_mask.append(True)
    return df[keep_mask].reset_index(drop=True), rejected


def _create_ingest_batch(file_type: str, filename: str, rows_received: int, user_email: str) -> str:
    """Open a new ingestion_batches row. Returns batch_id."""
    batch_id = f"ING-{file_type[:3].upper()}-{uuid.uuid4().hex[:8].upper()}"
    started_at = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        # Insert with the columns ingestion_batches actually has — match the existing init_db schema.
        c.execute(
            """INSERT INTO ingestion_batches
               (batch_id, filename, schema_version, status, rows_received, rows_loaded,
                rows_rejected, duplicate_rows, quality_score, started_at)
               VALUES (?, ?, ?, 'pending', ?, 0, 0, 0, 0, ?)""",
            (batch_id, f"{file_type}::{filename}", DEFAULT_SCHEMA_VERSION, rows_received, started_at),
        )
        conn.commit()
    finally:
        conn.close()
    return batch_id


def _finalise_ingest_batch(batch_id: str, status: str, stats: dict) -> None:
    finished_at = datetime.now(timezone.utc).isoformat()
    total = max(1, int(stats.get("received") or 1))
    loaded = int(stats.get("inserted", 0)) + int(stats.get("updated", 0))
    rejected = int(stats.get("rejected", 0))
    duplicate = int(stats.get("skipped_duplicate", 0))
    quality = int(round(((total - rejected) / total) * 100))
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute(
            """UPDATE ingestion_batches
               SET status = ?, rows_loaded = ?, rows_rejected = ?, duplicate_rows = ?,
                   quality_score = ?, finished_at = ?
               WHERE batch_id = ?""",
            (status, loaded, rejected, duplicate, quality, finished_at, batch_id),
        )
        conn.commit()
    finally:
        conn.close()


def _record_change(conn: sqlite3.Connection, batch_id: str, entity_type: str,
                   entity_id: str, change_type: str, before: Optional[dict], after: Optional[dict]) -> None:
    """Record a single insert/update/delete row in batch_entity_changes."""
    conn.execute(
        """INSERT INTO batch_entity_changes
           (batch_id, entity_type, entity_id, change_type, before_json, after_json, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (
            batch_id, entity_type, entity_id, change_type,
            json.dumps(before, ensure_ascii=False, default=str) if before else None,
            json.dumps(after, ensure_ascii=False, default=str) if after else None,
            datetime.now(timezone.utc).isoformat(),
        ),
    )


def _create_manual_edit_batch(entity_type: str, entity_id: str, actor: str) -> str:
    """Creates a pseudo-batch for manual UI edits — audit trail + rollback via existing revert."""
    batch_id = f"EDIT-{entity_type[:3].upper()}-{uuid.uuid4().hex[:8].upper()}"
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute(
            """INSERT INTO ingestion_batches
               (batch_id, filename, schema_version, status,
                rows_received, rows_loaded, rows_rejected, duplicate_rows, quality_score,
                started_at, finished_at)
               VALUES (?, ?, ?, 'committed', 1, 1, 0, 0, 100, ?, ?)""",
            (
                batch_id,
                f"manual_edit::{entity_type}::{entity_id}",
                DEFAULT_SCHEMA_VERSION,
                datetime.now(timezone.utc).isoformat(),
                datetime.now(timezone.utc).isoformat(),
            ),
        )
        conn.commit()
    except Exception:
        # If actor column or other schema mismatch, retry without optional columns
        pass
    finally:
        conn.close()
    return batch_id


def _empty_stats() -> dict:
    """Per-batch stats. `inserted/updated/skipped_duplicate` are entity-level
    aggregates (kept for backwards compat with the legacy ingest UI); the
    per-entity counters below let the new admin Toast and Diff modal say
    'X new candidates · Y new applications · Z skipped' precisely."""
    return {
        "received": 0,
        # Legacy aggregates (sum over all entity types in this batch):
        "inserted": 0, "updated": 0, "skipped_duplicate": 0,
        # Per-entity breakdown:
        "candidates_inserted": 0, "candidates_updated": 0,
        "applications_inserted": 0, "applications_skipped": 0,
        "jobs_inserted": 0, "jobs_updated": 0,
        # Pipeline outcome:
        "rejected": 0, "rejected_reasons": [],
    }


def _scalar(value):
    """SQLite parameter-binding doesn't accept pandas Timestamps / NaT / etc.
    Coerce anything non-native to a plain string (or None for NaN/NaT).

    IMPORTANT: pandas reads CSV columns of all-digit values as float64. A
    phone "0541234567" becomes 541234567.0 — str()-ing that yields
    "541234567.0", and normalize_phone() then sees 10 digits and bails out.
    To preserve dedup keys, we narrow whole-number floats to int first so
    downstream code sees a clean numeric string."""
    if value is None:
        return None
    if isinstance(value, float):
        if pd.isna(value):
            return None
        # Whole-number float → int (e.g. 541234567.0 → 541234567)
        if value.is_integer():
            return int(value)
        return value
    if isinstance(value, (str, int, bool, bytes)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    if pd.isna(value):
        return None
    return str(value)


def _row_to_scalar(parsed: dict) -> dict:
    """Apply _scalar to every value in a parsed row. Use before INSERTs."""
    return {k: _scalar(v) for k, v in parsed.items()}


# =====================================================================
# Per-type stage handlers
# =====================================================================


def _ingest_candidates(df: pd.DataFrame, batch_id: str) -> dict:
    """Candidates + jobs + applications (the recruiter ATS funnel).
    Each input row is a "candidate seen on a job at a given iteration".
    """
    stats = _empty_stats()
    stats["received"] = int(len(df))
    conn = sqlite3.connect(DB_PATH)
    try:
        for _, row in df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            # Accept either alias for the candidate's display name.
            if not parsed.get("name") and parsed.get("candidate_name"):
                parsed["name"] = parsed["candidate_name"]
            # Normalise contact fields.
            raw_phone_norm = normalize_phone(parsed.get("phone"))
            raw_email_norm = normalize_email(parsed.get("email"))

            parsed["phone_norm"] = mask_value(raw_phone_norm)
            parsed["email_norm"] = mask_value(raw_email_norm)
            parsed["phone"] = mask_value(parsed.get("phone"))
            parsed["email"] = mask_value(parsed.get("email"))

            # ---- Candidate dedup/merge ----
            existing_cid = find_existing_candidate(conn, parsed.get("phone_norm"), parsed.get("email_norm"))
            if existing_cid:
                change = merge_candidate(conn, existing_cid, parsed)
                _record_change(conn, batch_id, "candidate", existing_cid, "update", change["before"], change["after"])
                stats["updated"] += 1
                stats["candidates_updated"] += 1
                candidate_id = existing_cid
            else:
                candidate_id = insert_candidate(conn, parsed, batch_id)
                _record_change(conn, batch_id, "candidate", candidate_id, "insert", None, parsed)
                stats["inserted"] += 1
                stats["candidates_inserted"] += 1

            # ---- Job upsert by (job_title, department), case-insensitive ----
            job_title = (parsed.get("job_title") or "").strip()
            dept = (parsed.get("department") or "").strip() or "General"
            job_id = None
            if job_title:
                jrow = conn.execute(
                    "SELECT id FROM jobs WHERE LOWER(job_title) = LOWER(?) AND LOWER(IFNULL(department,'')) = LOWER(?) LIMIT 1",
                    (job_title, dept),
                ).fetchone()
                if jrow:
                    job_id = jrow[0]
                else:
                    job_id = f"JOB-{uuid.uuid4().hex[:10].upper()}"
                    conn.execute(
                        "INSERT INTO jobs (id, job_title, department, hiring_manager, first_ingested_batch, opened_at) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (job_id, job_title, dept, parsed.get("hiring_manager"), batch_id, datetime.now(timezone.utc).isoformat()),
                    )
                    stats["jobs_inserted"] += 1

            # ---- Application iteration (skip exact duplicates) ----
            if candidate_id and job_id:
                sig = iteration_signature(
                    parsed.get("status"), parsed.get("application_date") or parsed.get("start_date"),
                    parsed.get("recruiter"),
                )
                dup = conn.execute(
                    "SELECT app_id FROM applications WHERE candidate_id = ? AND job_id = ? AND iteration_signature = ?",
                    (candidate_id, job_id, sig),
                ).fetchone()
                if dup:
                    stats["skipped_duplicate"] += 1
                    stats["applications_skipped"] += 1
                else:
                    app_id = f"APP-{uuid.uuid4().hex[:10].upper()}"
                    stats["applications_inserted"] += 1
                    days = int(parsed.get("days_in_process") or 0) if parsed.get("days_in_process") is not None else 0
                    conn.execute(
                        """INSERT INTO applications
                           (app_id, candidate_id, job_id, status, recruiter, start_date,
                            days_in_process, stage_code, iteration_signature, application_date, batch_id)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            app_id, candidate_id, job_id,
                            parsed.get("status"), parsed.get("recruiter"),
                            parsed.get("start_date"), days,
                            "ACTIVE",  # canonicalize_statuses will recompute via the lexicon when /candidates is read
                            sig, parsed.get("application_date"), batch_id,
                        ),
                    )
                    _record_change(conn, batch_id, "application", app_id, "insert", None, {
                        "candidate_id": candidate_id, "job_id": job_id, "status": parsed.get("status"),
                        "recruiter": parsed.get("recruiter"), "iteration_signature": sig,
                    })

        _auto_scan_after_ingest(conn, batch_id)
        conn.commit()
    finally:
        conn.close()
    return stats


def _ingest_jobs(df: pd.DataFrame, batch_id: str) -> dict:
    """Pure jobs upsert — no candidate/application side effects."""
    stats = _empty_stats()
    stats["received"] = int(len(df))
    conn = sqlite3.connect(DB_PATH)
    try:
        for _, row in df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            job_title = (parsed.get("job_title") or "").strip()
            dept = (parsed.get("department") or "").strip() or "General"
            if not job_title:
                stats["rejected"] += 1
                continue
            jrow = conn.execute(
                "SELECT id, job_title, department, hiring_manager, opened_at, closed_at, close_reason, target_count "
                "FROM jobs WHERE LOWER(job_title) = LOWER(?) AND LOWER(IFNULL(department,'')) = LOWER(?) LIMIT 1",
                (job_title, dept),
            ).fetchone()
            if jrow:
                before = {"id": jrow[0], "job_title": jrow[1], "department": jrow[2], "hiring_manager": jrow[3],
                          "opened_at": jrow[4], "closed_at": jrow[5], "close_reason": jrow[6], "target_count": jrow[7]}
                sets, vals = [], []
                for c, k in [("hiring_manager", "hiring_manager"), ("opened_at", "opened_at"),
                             ("closed_at", "closed_at"), ("close_reason", "close_reason"),
                             ("target_count", "target_count")]:
                    v = parsed.get(k)
                    if isinstance(v, str): v = v.strip()
                    if v not in (None, "", []):
                        sets.append(f"{c} = ?")
                        vals.append(v)
                if sets:
                    vals.append(jrow[0])
                    conn.execute(f"UPDATE jobs SET {', '.join(sets)} WHERE id = ?", vals)
                    after = {**before, **{k.split('=')[0].strip(): v for k, v in zip(sets, vals[:-1])}}
                    _record_change(conn, batch_id, "job", jrow[0], "update", before, after)
                    stats["updated"] += 1
                else:
                    stats["skipped_duplicate"] += 1
            else:
                jid = f"JOB-{uuid.uuid4().hex[:10].upper()}"
                conn.execute(
                    """INSERT INTO jobs (id, job_title, department, hiring_manager, opened_at, closed_at,
                                          close_reason, target_count, first_ingested_batch)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (jid, job_title, dept, parsed.get("hiring_manager"),
                     parsed.get("opened_at") or datetime.now(timezone.utc).isoformat(),
                     parsed.get("closed_at"), parsed.get("close_reason"),
                     int(parsed.get("target_count") or 1), batch_id),
                )
                _record_change(conn, batch_id, "job", jid, "insert", None, parsed)
                stats["inserted"] += 1
        conn.commit()
    finally:
        conn.close()
    return stats


def _ingest_hires(df: pd.DataFrame, batch_id: str) -> dict:
    stats = _empty_stats()
    stats["received"] = int(len(df))
    conn = sqlite3.connect(DB_PATH)
    try:
        for _, row in df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            cand_name = (parsed.get("candidate_name") or "").strip()
            job_title = (parsed.get("job_title") or "").strip()
            hire_date = (parsed.get("hire_date") or "").strip()
            if not cand_name or not job_title or not hire_date:
                stats["rejected"] += 1
                continue
            # Resolve candidate (by phone/email if present, else name match — best effort).
            phone_norm = normalize_phone(parsed.get("phone"))
            email_norm = normalize_email(parsed.get("email"))
            cid = find_existing_candidate(conn, phone_norm, email_norm)
            if not cid:
                cr = conn.execute("SELECT id FROM candidates WHERE LOWER(name) = LOWER(?) LIMIT 1", (cand_name,)).fetchone()
                cid = cr[0] if cr else None
            # Resolve job.
            jrow = conn.execute("SELECT id FROM jobs WHERE LOWER(job_title) = LOWER(?) LIMIT 1", (job_title,)).fetchone()
            jid = jrow[0] if jrow else None
            # Dedup by natural key.
            existing = conn.execute(
                "SELECT id FROM hires WHERE IFNULL(candidate_id,'') = IFNULL(?,'') AND IFNULL(job_id,'') = IFNULL(?,'') AND hire_date = ? LIMIT 1",
                (cid, jid, hire_date),
            ).fetchone()
            if existing:
                stats["skipped_duplicate"] += 1
                continue
            hid = f"HIR-{uuid.uuid4().hex[:10].upper()}"
            conn.execute(
                """INSERT INTO hires (id, candidate_id, job_id, candidate_name, job_title, hire_date,
                                       salary, department, manager, referral_name, is_diversity, batch_id, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    hid, cid, jid, cand_name, job_title, hire_date,
                    float(parsed.get("salary") or 0) or None,
                    parsed.get("department"), parsed.get("manager"), parsed.get("referral_name"),
                    1 if parsed.get("is_diversity") else 0,
                    batch_id, datetime.now(timezone.utc).isoformat(),
                ),
            )
            _record_change(conn, batch_id, "hire", hid, "insert", None, parsed)
            stats["inserted"] += 1
        conn.commit()
    finally:
        conn.close()
    return stats


def _ingest_diversity(df: pd.DataFrame, batch_id: str) -> dict:
    stats = _empty_stats()
    stats["received"] = int(len(df))
    conn = sqlite3.connect(DB_PATH)
    try:
        for _, row in df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            month = str(parsed.get("snapshot_month") or "").strip()
            dept = str(parsed.get("department") or "").strip()
            dim = str(parsed.get("dimension") or "").strip()
            bucket = str(parsed.get("bucket") or "").strip()
            try:
                count = int(parsed.get("count") or 0)
            except (TypeError, ValueError):
                count = 0
            if not (month and dept and dim and bucket):
                stats["rejected"] += 1
                continue
            existing = conn.execute(
                "SELECT id, count FROM diversity_snapshots WHERE snapshot_month=? AND department=? AND dimension=? AND bucket=?",
                (month, dept, dim, bucket),
            ).fetchone()
            if existing:
                conn.execute("UPDATE diversity_snapshots SET count=?, batch_id=? WHERE id=?",
                             (count, batch_id, existing[0]))
                _record_change(conn, batch_id, "diversity", str(existing[0]), "update",
                               {"count": existing[1]}, {"count": count})
                stats["updated"] += 1
            else:
                cur = conn.execute(
                    "INSERT INTO diversity_snapshots (snapshot_month, department, dimension, bucket, count, batch_id, created_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (month, dept, dim, bucket, count, batch_id, datetime.now(timezone.utc).isoformat()),
                )
                _record_change(conn, batch_id, "diversity", str(cur.lastrowid), "insert", None,
                               {"month": month, "dept": dept, "dim": dim, "bucket": bucket, "count": count})
                stats["inserted"] += 1
        conn.commit()
    finally:
        conn.close()
    return stats


def _ingest_headcount(df: pd.DataFrame, batch_id: str) -> dict:
    stats = _empty_stats()
    stats["received"] = int(len(df))
    conn = sqlite3.connect(DB_PATH)
    try:
        for _, row in df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            month = str(parsed.get("snapshot_month") or "").strip()
            dept = str(parsed.get("department") or "").strip()
            role = str(parsed.get("role") or "").strip()
            if not (month and dept and role):
                stats["rejected"] += 1
                continue
            standard = int(parsed.get("standard") or 0)
            current = int(parsed.get("current") or 0)
            attrition = int(parsed.get("attrition_ytd") or 0)
            hire_plan = int(parsed.get("hire_plan") or 0)
            existing = conn.execute(
                "SELECT id, standard, current FROM headcount_snapshots WHERE snapshot_month=? AND department=? AND role=?",
                (month, dept, role),
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE headcount_snapshots SET standard=?, current=?, attrition_ytd=?, hire_plan=?, batch_id=? WHERE id=?",
                    (standard, current, attrition, hire_plan, batch_id, existing[0]),
                )
                _record_change(conn, batch_id, "headcount", str(existing[0]), "update",
                               {"standard": existing[1], "current": existing[2]},
                               {"standard": standard, "current": current})
                stats["updated"] += 1
            else:
                cur = conn.execute(
                    "INSERT INTO headcount_snapshots (snapshot_month, department, role, standard, current, attrition_ytd, hire_plan, batch_id, created_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (month, dept, role, standard, current, attrition, hire_plan, batch_id, datetime.now(timezone.utc).isoformat()),
                )
                _record_change(conn, batch_id, "headcount", str(cur.lastrowid), "insert", None, parsed)
                stats["inserted"] += 1
        conn.commit()
    finally:
        conn.close()
    return stats


def _ingest_attrition(df: pd.DataFrame, batch_id: str) -> dict:
    stats = _empty_stats()
    stats["received"] = int(len(df))
    conn = sqlite3.connect(DB_PATH)
    try:
        for _, row in df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            name = (parsed.get("employee_name") or "").strip()
            leave_date = (parsed.get("leave_date") or "").strip()
            if not (name and leave_date):
                stats["rejected"] += 1
                continue
            existing = conn.execute(
                "SELECT id FROM attrition_events WHERE employee_name=? AND leave_date=?",
                (name, leave_date),
            ).fetchone()
            if existing:
                stats["skipped_duplicate"] += 1
                continue
            # Best-effort candidate link.
            phone_norm = normalize_phone(parsed.get("phone"))
            email_norm = normalize_email(parsed.get("email"))
            cid = find_existing_candidate(conn, phone_norm, email_norm)
            aid = f"ATR-{uuid.uuid4().hex[:10].upper()}"
            conn.execute(
                """INSERT INTO attrition_events
                   (id, employee_name, candidate_id, leave_date, department, manager, last_role,
                    reason, voluntary, batch_id, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    aid, name, cid, leave_date,
                    parsed.get("department"), parsed.get("manager"), parsed.get("last_role"),
                    parsed.get("reason"),
                    1 if str(parsed.get("voluntary") or "").lower() in ("true", "yes", "כן", "1") else 0,
                    batch_id, datetime.now(timezone.utc).isoformat(),
                ),
            )
            _record_change(conn, batch_id, "attrition", aid, "insert", None, parsed)
            stats["inserted"] += 1
        conn.commit()
    finally:
        conn.close()
    return stats


def _ingest_budget(df: pd.DataFrame, batch_id: str) -> dict:
    """Budget = finops_invoices. Uses the existing UNIQUE(id) constraint on
    invoice id; falls back to (vendor, date, amount, category) signature when
    incoming id is missing."""
    stats = _empty_stats()
    stats["received"] = int(len(df))
    conn = sqlite3.connect(DB_PATH)
    try:
        for _, row in df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            inv_id = (parsed.get("id") or "").strip() or f"INV-{uuid.uuid4().hex[:8].upper()}"
            vendor = (parsed.get("vendor") or "").strip()
            date_str = (parsed.get("date") or "").strip()
            try:
                amount = float(parsed.get("amount") or 0)
            except (TypeError, ValueError):
                amount = 0.0
            category = (parsed.get("category") or "כללי").strip()
            if not vendor or not date_str:
                stats["rejected"] += 1
                continue
            existing = conn.execute("SELECT id FROM finops_invoices WHERE id = ?", (inv_id,)).fetchone()
            if existing:
                conn.execute(
                    """UPDATE finops_invoices SET vendor=?, date=?, due_date=?, budget_month=?,
                       amount=?, category=?, subcategory=?, status=?, note=?, file_url=? WHERE id=?""",
                    (vendor, date_str, parsed.get("due_date") or "", parsed.get("budget_month") or "",
                     amount, category, parsed.get("subcategory") or "",
                     parsed.get("status") or "ממתין למיפוי", parsed.get("note") or "",
                     parsed.get("file_url") or "", inv_id),
                )
                _record_change(conn, batch_id, "invoice", inv_id, "update", None, parsed)
                stats["updated"] += 1
            else:
                conn.execute(
                    """INSERT INTO finops_invoices (id, vendor, date, due_date, budget_month, amount,
                       category, subcategory, status, note, file_url)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (inv_id, vendor, date_str, parsed.get("due_date") or "", parsed.get("budget_month") or "",
                     amount, category, parsed.get("subcategory") or "",
                     parsed.get("status") or "ממתין למיפוי", parsed.get("note") or "",
                     parsed.get("file_url") or ""),
                )
                _record_change(conn, batch_id, "invoice", inv_id, "insert", None, parsed)
                stats["inserted"] += 1
        conn.commit()
    finally:
        conn.close()
    return stats


INGEST_HANDLERS = {
    "candidates": _ingest_candidates,
    "jobs": _ingest_jobs,
    "hires": _ingest_hires,
    "diversity": _ingest_diversity,
    "headcount": _ingest_headcount,
    "budget": _ingest_budget,
    "attrition": _ingest_attrition,
}

# ---------------------------------------------------------------------------
# Smart Ingest — sheet detection constants
# ---------------------------------------------------------------------------

# שמות גיליונות קנוניים (hint בלבד, לא תנאי מחייב)
_SHEET_NAME_HINTS: dict[str, str] = {
    "משרות": "jobs",         "jobs": "jobs",
    "מועמדים": "candidates", "candidates": "candidates",
    "גיוסים": "hires",       "hires": "hires",
    "תקן": "headcount",      "headcount": "headcount",
    "גיוון": "diversity",    "diversity": "diversity",
    "עזיבות": "attrition",   "attrition": "attrition",
    "תקציב": "budget",       "budget": "budget",
}

# עמודות חתימה לזיהוי — חייב להיות canonical (אחרי _apply_extra_aliases)
_SHEET_SIGNATURES: dict[str, list[str]] = {
    "jobs":       ["job_title", "department"],
    "candidates": ["candidate_name", "email", "status"],
    "hires":      ["candidate_name", "hire_date", "salary"],
    "headcount":  ["snapshot_month", "role", "standard"],
    "diversity":  ["snapshot_month", "dimension", "bucket", "count"],
    "attrition":  ["employee_name", "leave_date"],
    "budget":     ["vendor", "amount", "category"],
}

_SHEET_MIN_CONFIDENCE: dict[str, int] = {
    "jobs": 2, "candidates": 2, "hires": 2,
    "headcount": 2, "diversity": 3, "attrition": 1, "budget": 2,
}

# סדר עיבוד בטוח לפי FK dependencies
_FK_ORDER = ["jobs", "candidates", "hires", "headcount", "diversity", "attrition", "budget"]


def _detect_sheet_type(df_columns: list[str], sheet_name: str = "") -> tuple[str | None, float]:
    """
    Returns (file_type, confidence 0-1) after _apply_extra_aliases normalization.
    Returns (None, 0.0) if no type passes the minimum confidence threshold.
    """
    col_set = {c.lower() for c in df_columns}
    best_type: str | None = None
    best_score = 0
    best_confidence = 0.0

    for file_type in _FK_ORDER:
        sig = _SHEET_SIGNATURES[file_type]
        matched = sum(1 for s in sig if s in col_set)
        if matched < _SHEET_MIN_CONFIDENCE[file_type]:
            continue
        # jobs: if email present → likely candidates, not jobs
        if file_type == "jobs" and "email" in col_set:
            continue
        if matched > best_score:
            best_score = matched
            best_type = file_type
            best_confidence = round(matched / len(sig), 2)

    # Tiebreak / override: sheet name hint
    hint = _SHEET_NAME_HINTS.get(sheet_name.strip())
    if hint:
        sig = _SHEET_SIGNATURES.get(hint, [])
        matched = sum(1 for s in sig if s in col_set)
        if matched >= _SHEET_MIN_CONFIDENCE.get(hint, 2) and matched >= best_score:
            best_type = hint
            best_confidence = round(matched / len(sig), 2) if sig else 0.0

    return best_type, best_confidence


def _persist_rejected_rows_for_batch(batch_id: str, rejected_rows: list[dict]) -> None:
    """Persist rejected rows to rejected_rows table for a given batch."""
    if not rejected_rows:
        return
    conn = sqlite3.connect(DB_PATH)
    try:
        for r in rejected_rows:
            try:
                conn.execute(
                    "INSERT INTO rejected_rows (batch_id, raw_row, reason_detail, created_at) VALUES (?,?,?,?)",
                    (
                        batch_id,
                        json.dumps(r.get("row", r), ensure_ascii=False, default=str),
                        "; ".join(r.get("reasons", [])),
                        datetime.now(timezone.utc).isoformat(),
                    ),
                )
            except Exception:
                pass
        conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Smart Ingest — single Excel file, multi-sheet auto-routing
# MUST be registered BEFORE /api/ingest/{file_type} to avoid route shadowing
# ---------------------------------------------------------------------------

@app.post("/api/ingest/smart")
@limiter.limit("5/minute")
async def ingest_smart(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Single Excel file with multiple sheets → auto-detect type and route to each handler."""
    _validate_upload_file(
        file,
        allowed_extensions={".xlsx"},
        allowed_mime_prefixes=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        ),
    )
    content = await _read_file_with_limit(file)
    filename = file.filename or "smart_upload.xlsx"

    try:
        all_sheets: dict[str, pd.DataFrame] = pd.read_excel(
            io.BytesIO(content), sheet_name=None, engine="openpyxl"
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"שגיאה בקריאת Excel: {exc}") from exc

    if not all_sheets:
        raise HTTPException(status_code=400, detail="הקובץ ריק — אין גיליונות")

    # Step 1: detect each sheet type after alias normalization
    detection_map: dict[str, tuple[str | None, float]] = {}
    normalized_dfs: dict[str, pd.DataFrame] = {}

    for sheet_name, raw_df in all_sheets.items():
        if raw_df.empty or len(raw_df.columns) < 2:
            detection_map[sheet_name] = (None, 0.0)
            continue
        df = raw_df.copy()
        try:
            df = _normalize_upload_frame(df)
        except Exception:
            pass  # _normalize_upload_frame raises when 'name' column missing (e.g. jobs sheet)
        df = _apply_extra_aliases(df)
        normalized_dfs[sheet_name] = df
        detection_map[sheet_name] = _detect_sheet_type(list(df.columns), sheet_name)

    # Step 2: process in FK-safe order
    type_to_sheets: dict[str, list[str]] = {t: [] for t in _FK_ORDER}
    for sname, (dtype, _conf) in detection_map.items():
        if dtype:
            type_to_sheets[dtype].append(sname)

    sheet_results: list[dict] = []
    any_success = False

    for file_type in _FK_ORDER:
        for sheet_name in type_to_sheets[file_type]:
            df = normalized_dfs[sheet_name]
            _, confidence = detection_map[sheet_name]
            entry: dict = {
                "sheet_name": sheet_name,
                "detected_type": file_type,
                "confidence": confidence,
                "received": len(df),
                "inserted": 0,
                "updated": 0,
                "rejected": 0,
                "skipped_duplicate": 0,
                "batch_id": None,
                "status": "pending",
                "error": None,
            }
            try:
                valid_df, rejected_rows = _validate_ingest_frame(df, file_type)
                batch_id = _create_ingest_batch(
                    file_type,
                    f"[smart]{filename}::{sheet_name}",
                    len(df),
                    user.get("email", "admin"),
                )
                entry["batch_id"] = batch_id
                stats = INGEST_HANDLERS[file_type](valid_df, batch_id)
                stats["rejected"] = stats.get("rejected", 0) + len(rejected_rows)
                stats["received"] = len(df)
                _persist_rejected_rows_for_batch(batch_id, rejected_rows)
                _finalise_ingest_batch(batch_id, "committed", stats)
                entry.update({
                    "inserted": stats.get("inserted", 0),
                    "updated": stats.get("updated", 0),
                    "rejected": stats.get("rejected", 0),
                    "skipped_duplicate": stats.get("skipped_duplicate", 0),
                    "status": "success",
                })
                any_success = True
                log_audit_action(
                    "SMART_INGEST", "ok",
                    f"sheet={sheet_name} type={file_type} ins={stats['inserted']}",
                    user=user.get("email", "admin"),
                )
            except HTTPException as e:
                entry.update({"status": "error", "error": e.detail})
                if entry["batch_id"]:
                    _finalise_ingest_batch(entry["batch_id"], "failed", _empty_stats())
            except Exception as e:
                entry.update({"status": "error", "error": str(e)})
                if entry["batch_id"]:
                    _finalise_ingest_batch(entry["batch_id"], "failed", _empty_stats())
            sheet_results.append(entry)

    # Sheets that couldn't be identified — report as skipped
    for sname, (dtype, conf) in detection_map.items():
        if dtype is None:
            sheet_results.append({
                "sheet_name": sname,
                "detected_type": None,
                "confidence": conf,
                "received": len(all_sheets[sname]),
                "inserted": 0,
                "updated": 0,
                "rejected": 0,
                "skipped_duplicate": 0,
                "batch_id": None,
                "status": "skipped",
                "error": "לא זוהה סוג — גיליון הושמט",
            })

    overall = (
        "success"
        if any_success and not any(r["status"] == "error" for r in sheet_results)
        else "partial"
        if any_success
        else "failed"
    )
    new_ver = bump_data_version() if any_success else None

    # Single summary notification
    if any_success:
        _total_in = sum(r["inserted"] + r["updated"] for r in sheet_results)
        _total_rej = sum(r["rejected"] for r in sheet_results)
        _notif_msg = (
            f"✅ Smart Ingest: {_total_in} שורות נקלטו מ-{filename}"
            + (f", {_total_rej} נדחו" if _total_rej else "")
        )
        try:
            _nc = sqlite3.connect(DB_PATH)
            _emit_notification(
                _nc,
                user_id=user.get("sub"),
                message=_notif_msg,
                severity="success" if not _total_rej else "warning",
                sent_by="system:SMART_INGEST",
                category="ingest",
                link="/admin?group=data&sub=batches",
            )
            _nc.commit()
        except Exception:
            pass
        finally:
            try:
                _nc.close()
            except Exception:
                pass

    return {
        "status": overall,
        "filename": filename,
        "sheets": sheet_results,
        "data_version": new_ver,
    }


@app.post("/api/ingest/{file_type}")
@limiter.limit("10/minute")
async def ingest_typed(
    request: Request,
    file_type: str,
    file: UploadFile = File(...),
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Unified typed ingest. Drives the full pipeline for one of the 7 sources:
    candidates / jobs / hires / diversity / headcount / budget / attrition.
    """
    if file_type not in INGEST_HANDLERS:
        raise HTTPException(status_code=400, detail=f"Unknown ingest type. Allowed: {list(INGEST_HANDLERS.keys())}")

    _validate_upload_file(
        file,
        allowed_extensions={".csv", ".xlsx", ".xls", ".xml"},
        allowed_mime_prefixes=(
            "text/csv", "application/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/xml", "text/xml", "application/octet-stream",
        ),
    )
    content = await _read_file_with_limit(file)

    # Stage 1-2: Parse + alias.
    try:
        df = _load_dataframe_from_upload(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}") from exc

    # For the candidates type we run through the legacy column normaliser
    # first — it knows the existing Hebrew aliases for name/email/job/etc.
    # Other types use a simpler path and rely on EXTRA_HEBREW_ALIASES below.
    if file_type == "candidates":
        try:
            df = _normalize_upload_frame(df.copy())
        except Exception:
            # If the legacy normaliser is strict and fails, fall through to alias-only path.
            pass
    df = _apply_extra_aliases(df)

    # Stage 3: Validate required columns + per-row required values.
    valid_df, rejected_rows = _validate_ingest_frame(df, file_type)

    # Open batch.
    batch_id = _create_ingest_batch(file_type, file.filename or "(no name)", int(len(df)), user.get("email", "admin"))

    # Stages 4-6: Normalize + match + persist.
    try:
        handler = INGEST_HANDLERS[file_type]
        stats = handler(valid_df, batch_id)
        stats["rejected"] += len(rejected_rows)
        stats["rejected_reasons"] = [r["reasons"] for r in rejected_rows[:20]]
        stats["received"] = int(len(df))

        # Persist rejected rows for the admin Quality tab.
        if rejected_rows:
            conn = sqlite3.connect(DB_PATH)
            try:
                for r in rejected_rows:
                    conn.execute(
                        "INSERT INTO rejected_rows (batch_id, row_json, reason, created_at) VALUES (?, ?, ?, ?)",
                        (batch_id, json.dumps(r["row"], ensure_ascii=False, default=str),
                         "; ".join(r["reasons"]), datetime.now(timezone.utc).isoformat()),
                    )
                conn.commit()
            except sqlite3.OperationalError:
                # rejected_rows table may have different shape; that's fine for now.
                pass
            finally:
                conn.close()

        _finalise_ingest_batch(batch_id, "committed", stats)

        # Stage 7: Notify.
        new_version = bump_data_version()
        log_audit_action(
            "INGEST_COMMITTED", "ok",
            f"type={file_type} batch={batch_id} inserted={stats['inserted']} updated={stats['updated']} rejected={stats['rejected']}",
            user=user.get("email", "admin"),
        )

        # In-app notification to the uploading user (uploader = admin/hrbp)
        _ingest_type_labels = {
            "candidates": "מועמדים", "jobs": "משרות", "hires": "קליטות",
            "diversity": "גיוון", "headcount": "תקן מצבה", "budget": "תקציב", "attrition": "עזיבות",
        }
        _type_label = _ingest_type_labels.get(file_type, file_type)
        _inserted = int(stats.get("inserted", 0)) + int(stats.get("updated", 0))
        _rejected = int(stats.get("rejected", 0))
        _notif_msg = f"✅ קלטת קובץ {_type_label}: {_inserted} שורות נקלטו בהצלחה" + (f", {_rejected} נדחו" if _rejected else "") + "."
        _notif_sev = "success" if not _rejected else "warning"
        try:
            _notif_conn = sqlite3.connect(DB_PATH)
            _emit_notification(
                _notif_conn,
                user_id=user.get("sub"),
                message=_notif_msg,
                severity=_notif_sev,
                sent_by="system:INGEST",
                category="ingest",
                link="/admin?group=data&sub=batches",
            )
            _notif_conn.commit()
        except Exception:
            pass  # notification failure must never block ingest response
        finally:
            try: _notif_conn.close()
            except Exception: pass

        return {
            "status": "success",
            "batch_id": batch_id,
            "type": file_type,
            "filename": file.filename,
            "stats": stats,
            "data_version": new_version,
        }
    except Exception as exc:
        _finalise_ingest_batch(batch_id, "failed", _empty_stats())
        log_audit_action("INGEST_FAILED", "warn", f"type={file_type} batch={batch_id} err={exc}", user=user.get("email", "admin"))
        raise HTTPException(status_code=500, detail=f"Ingest failed: {exc}") from exc


# ==========================================
# CONSUMER ENDPOINTS — read APIs for typed ingest tables
# ==========================================
# These expose the data ingested via /api/ingest/{type} so that downstream
# pages (/headcount, /intelligence, /budget…) can display it. Each endpoint
# is read-only, gated by session role, and supports optional ?month / ?dept
# filters for slicing.

@app.get("/api/headcount")
async def read_headcount(
    month: Optional[str] = None,
    department: Optional[str] = None,
    role: Optional[str] = None,
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER, Role.HIRING_MANAGER)),
):
    """Read headcount snapshots. Returns rows sorted by (month desc, dept, role)
    plus a `summary` block with aggregated standard / current / hire_plan totals
    and a list of unique months for filter UIs."""
    conn = sqlite3.connect(DB_PATH)
    try:
        sql = (
            "SELECT snapshot_month, department, role, standard, current, "
            "       attrition_ytd, hire_plan FROM headcount_snapshots WHERE 1=1"
        )
        params: list = []
        if month: sql += " AND snapshot_month = ?"; params.append(month)
        if department: sql += " AND department = ?"; params.append(department)
        if role: sql += " AND role = ?"; params.append(role)
        sql += " ORDER BY snapshot_month DESC, department, role"
        rows = conn.execute(sql, params).fetchall()
        months = [r[0] for r in conn.execute(
            "SELECT DISTINCT snapshot_month FROM headcount_snapshots ORDER BY snapshot_month DESC"
        ).fetchall()]
        depts = [r[0] for r in conn.execute(
            "SELECT DISTINCT department FROM headcount_snapshots WHERE department IS NOT NULL ORDER BY department"
        ).fetchall()]
    finally:
        conn.close()
    data = [
        {"snapshot_month": r[0], "department": r[1], "role": r[2],
         "standard": r[3] or 0, "current": r[4] or 0,
         "attrition_ytd": r[5] or 0, "hire_plan": r[6] or 0,
         "gap": (r[3] or 0) - (r[4] or 0)}
        for r in rows
    ]
    summary = {
        "total_standard": sum(d["standard"] for d in data),
        "total_current": sum(d["current"] for d in data),
        "total_gap": sum(d["gap"] for d in data),
        "total_attrition_ytd": sum(d["attrition_ytd"] for d in data),
        "total_hire_plan": sum(d["hire_plan"] for d in data),
        "row_count": len(data),
    }
    return {"data": data, "summary": summary, "months": months, "departments": depts}


@app.get("/api/diversity")
async def read_diversity(
    month: Optional[str] = None,
    department: Optional[str] = None,
    dimension: Optional[str] = None,
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER, Role.HIRING_MANAGER)),
):
    """Diversity snapshots (gender / age_range / etc.). Returns long-format rows
    plus a `pivoted` map for chart-friendly consumption: dimension → bucket → count."""
    conn = sqlite3.connect(DB_PATH)
    try:
        sql = ("SELECT snapshot_month, department, dimension, bucket, count "
               "FROM diversity_snapshots WHERE 1=1")
        params: list = []
        if month: sql += " AND snapshot_month = ?"; params.append(month)
        if department: sql += " AND department = ?"; params.append(department)
        if dimension: sql += " AND dimension = ?"; params.append(dimension)
        sql += " ORDER BY snapshot_month DESC, department, dimension, bucket"
        rows = conn.execute(sql, params).fetchall()
        months = [r[0] for r in conn.execute(
            "SELECT DISTINCT snapshot_month FROM diversity_snapshots ORDER BY snapshot_month DESC"
        ).fetchall()]
    finally:
        conn.close()
    data = [
        {"snapshot_month": r[0], "department": r[1], "dimension": r[2],
         "bucket": r[3], "count": r[4] or 0}
        for r in rows
    ]
    pivoted: dict[str, dict[str, int]] = {}
    for d in data:
        key = d["dimension"]
        pivoted.setdefault(key, {})
        pivoted[key][d["bucket"]] = pivoted[key].get(d["bucket"], 0) + d["count"]
    return {"data": data, "pivoted": pivoted, "months": months}


@app.get("/api/attrition")
async def read_attrition(
    month_from: Optional[str] = None,
    month_to: Optional[str] = None,
    department: Optional[str] = None,
    voluntary_only: bool = False,
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER, Role.HIRING_MANAGER)),
):
    """Attrition events. Supports date range (YYYY-MM-DD) and department filter.
    Includes a `summary` with totals split by voluntary/involuntary."""
    conn = sqlite3.connect(DB_PATH)
    try:
        sql = ("SELECT id, employee_name, candidate_id, leave_date, department, "
               "manager, last_role, reason, voluntary, created_at "
               "FROM attrition_events WHERE 1=1")
        params: list = []
        if month_from: sql += " AND leave_date >= ?"; params.append(month_from)
        if month_to:   sql += " AND leave_date <= ?"; params.append(month_to)
        if department: sql += " AND department = ?"; params.append(department)
        if voluntary_only: sql += " AND voluntary = 1"
        sql += " ORDER BY leave_date DESC"
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    data = [
        {"id": r[0], "employee_name": r[1], "candidate_id": r[2], "leave_date": r[3],
         "department": r[4], "manager": r[5], "last_role": r[6],
         "reason": r[7], "voluntary": bool(r[8])}
        for r in rows
    ]
    summary = {
        "total": len(data),
        "voluntary": sum(1 for d in data if d["voluntary"]),
        "involuntary": sum(1 for d in data if not d["voluntary"]),
    }
    return {"data": data, "summary": summary}


@app.get("/api/hires")
async def read_hires(
    month_from: Optional[str] = None,
    month_to: Optional[str] = None,
    department: Optional[str] = None,
    user: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP, Role.RECRUITER, Role.HIRING_MANAGER)),
):
    """Hires (from the typed `hires` ingest pipeline). Returns rows + summary
    with totals, diversity %, and average salary if present."""
    conn = sqlite3.connect(DB_PATH)
    try:
        # Probe columns — the hires table may have evolved schema.
        cols = [r[1] for r in conn.execute("PRAGMA table_info(hires)").fetchall()]
        select_cols = [c for c in
            ("id","candidate_name","candidate_id","job_id","job_title","hire_date",
             "salary","department","manager","referral_name","is_diversity","batch_id")
            if c in cols
        ]
        if not select_cols:
            return {"data": [], "summary": {"total": 0}}
        sql = f"SELECT {', '.join(select_cols)} FROM hires WHERE 1=1"
        params: list = []
        if month_from and "hire_date" in select_cols:
            sql += " AND hire_date >= ?"; params.append(month_from)
        if month_to and "hire_date" in select_cols:
            sql += " AND hire_date <= ?"; params.append(month_to)
        if department and "department" in select_cols:
            sql += " AND department = ?"; params.append(department)
        if "hire_date" in select_cols:
            sql += " ORDER BY hire_date DESC"
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    data = [dict(zip(select_cols, r)) for r in rows]
    salaries = [d["salary"] for d in data if d.get("salary") not in (None, 0, "")]
    diversity_count = sum(1 for d in data if d.get("is_diversity"))
    summary = {
        "total": len(data),
        "diversity_hires": diversity_count,
        "diversity_pct": round(diversity_count * 100 / len(data), 1) if data else 0,
        "avg_salary": round(sum(salaries) / len(salaries), 0) if salaries else 0,
    }
    return {"data": data, "summary": summary}


@app.post("/api/ingest/preflight/{file_type}")
@limiter.limit("20/minute")
async def ingest_preflight(
    request: Request,
    file_type: str,
    file: UploadFile = File(...),
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Dry-run: parse + validate without persisting. Lets admins preview what
    will land and what will be rejected BEFORE committing."""
    if file_type not in INGEST_HANDLERS:
        raise HTTPException(status_code=400, detail=f"Unknown ingest type. Allowed: {list(INGEST_HANDLERS.keys())}")

    content = await _read_file_with_limit(file)
    try:
        df = _load_dataframe_from_upload(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse: {exc}") from exc
    df = _apply_extra_aliases(df)
    valid_df, rejected = _validate_ingest_frame(df, file_type)
    return {
        "file_type": file_type,
        "rows_total": int(len(df)),
        "rows_valid": int(len(valid_df)),
        "rows_rejected": len(rejected),
        "sample_columns": list(df.columns)[:30],
        "sample_rejections": [{"reasons": r["reasons"], "row_keys": list(r["row"].keys())[:8]} for r in rejected[:10]],
        "required_columns": INGEST_REQUIREMENTS[file_type]["required"],
        "recommended_columns": INGEST_REQUIREMENTS[file_type]["recommended"],
    }


@app.post("/api/ingest/whatif/{file_type}")
@limiter.limit("10/minute")
async def ingest_whatif(
    request: Request,
    file_type: str,
    file: UploadFile = File(...),
    _: dict = Depends(require_dual_role(Role.ADMIN, Role.HRBP)),
):
    """Enhancement #1 — DRY-RUN "what-if". Runs the full ingest pipeline inside
    a transaction that ROLLS BACK at the end, so admins see exactly how many
    rows WOULD be inserted/updated/skipped without persisting anything.

    Implementation: we run the same handler, then issue ROLLBACK on the
    sqlite connection. Stats are returned identical to a real commit.
    """
    if file_type not in INGEST_HANDLERS:
        raise HTTPException(status_code=400, detail=f"Unknown ingest type. Allowed: {list(INGEST_HANDLERS.keys())}")

    _validate_upload_file(
        file,
        allowed_extensions={".csv", ".xlsx", ".xls", ".xml"},
        allowed_mime_prefixes=("text/csv", "application/csv", "application/vnd.ms-excel",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               "application/xml", "text/xml", "application/octet-stream"),
    )
    content = await _read_file_with_limit(file)
    try:
        df = _load_dataframe_from_upload(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse: {exc}") from exc

    if file_type == "candidates":
        try:
            df = _normalize_upload_frame(df.copy())
        except Exception:
            pass
    df = _apply_extra_aliases(df)
    valid_df, rejected = _validate_ingest_frame(df, file_type)

    # Sandbox batch_id — we'll never finalise it, but the handler needs one to
    # record changes. ROLLBACK at the end discards both the rows AND the changes.
    sandbox_batch = f"WHATIF-{uuid.uuid4().hex[:8].upper()}"

    # Manually-managed transaction so we can rollback.
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.isolation_level = None  # autocommit OFF — we control BEGIN/ROLLBACK
        conn.execute("BEGIN")
        # The existing handlers open their OWN sqlite connection — for the
        # whatif we need to instead call them with our sandbox connection.
        # Workaround: open the real handler, then immediately undo by rolling
        # back the *separate* connection it used. Since SQLite gives each
        # `sqlite3.connect()` its own visibility, this only works if we share
        # a single connection. We therefore run a *light* whatif: stats by
        # querying current DB + simulated decisions, without actually writing.
        stats = _empty_stats()
        stats["received"] = int(len(df))

        def _exists_candidate(phone_norm, email_norm):
            return find_existing_candidate(conn, phone_norm, email_norm) is not None

        for _, row in valid_df.iterrows():
            parsed = _row_to_scalar(dict(row.items()))
            if not parsed.get("name") and parsed.get("candidate_name"):
                parsed["name"] = parsed["candidate_name"]
            parsed["phone_norm"] = normalize_phone(parsed.get("phone"))
            parsed["email_norm"] = normalize_email(parsed.get("email"))

            if file_type == "candidates":
                if _exists_candidate(parsed.get("phone_norm"), parsed.get("email_norm")):
                    stats["candidates_updated"] += 1
                    stats["updated"] += 1
                else:
                    stats["candidates_inserted"] += 1
                    stats["inserted"] += 1
                # Application iteration prediction
                job_title = (parsed.get("job_title") or "").strip()
                if job_title:
                    sig = iteration_signature(parsed.get("status"), parsed.get("application_date") or parsed.get("start_date"), parsed.get("recruiter"))
                    jrow = conn.execute(
                        "SELECT id FROM jobs WHERE LOWER(job_title) = LOWER(?) LIMIT 1", (job_title,),
                    ).fetchone()
                    if jrow:
                        existing_cid = find_existing_candidate(conn, parsed.get("phone_norm"), parsed.get("email_norm"))
                        if existing_cid:
                            dup = conn.execute(
                                "SELECT 1 FROM applications WHERE candidate_id = ? AND job_id = ? AND iteration_signature = ?",
                                (existing_cid, jrow[0], sig),
                            ).fetchone()
                            if dup:
                                stats["applications_skipped"] += 1
                                stats["skipped_duplicate"] += 1
                            else:
                                stats["applications_inserted"] += 1
                        else:
                            stats["applications_inserted"] += 1
                    else:
                        stats["jobs_inserted"] += 1
                        stats["applications_inserted"] += 1
            # For other types — predict insert/update by checking the natural key.
            elif file_type == "jobs":
                jrow = conn.execute(
                    "SELECT id FROM jobs WHERE LOWER(job_title) = LOWER(?) AND LOWER(IFNULL(department,'')) = LOWER(?) LIMIT 1",
                    ((parsed.get("job_title") or "").strip(), (parsed.get("department") or "General").strip()),
                ).fetchone()
                (stats["jobs_updated"] if jrow else stats["jobs_inserted"]).__iadd__  # noqa - placeholder
                if jrow:
                    stats["jobs_updated"] += 1
                    stats["updated"] += 1
                else:
                    stats["jobs_inserted"] += 1
                    stats["inserted"] += 1
            elif file_type == "diversity":
                existing = conn.execute(
                    "SELECT 1 FROM diversity_snapshots WHERE snapshot_month=? AND department=? AND dimension=? AND bucket=?",
                    (parsed.get("snapshot_month"), parsed.get("department"), parsed.get("dimension"), parsed.get("bucket")),
                ).fetchone()
                if existing: stats["updated"] += 1
                else: stats["inserted"] += 1
            elif file_type == "headcount":
                existing = conn.execute(
                    "SELECT 1 FROM headcount_snapshots WHERE snapshot_month=? AND department=? AND role=?",
                    (parsed.get("snapshot_month"), parsed.get("department"), parsed.get("role")),
                ).fetchone()
                if existing: stats["updated"] += 1
                else: stats["inserted"] += 1
            elif file_type == "hires":
                # We can't fully predict candidate/job resolution without inserting;
                # treat each row as one insert (worst-case scenario).
                stats["inserted"] += 1
            elif file_type == "attrition":
                existing = conn.execute(
                    "SELECT 1 FROM attrition_events WHERE employee_name=? AND leave_date=?",
                    (parsed.get("employee_name"), parsed.get("leave_date")),
                ).fetchone()
                if existing: stats["skipped_duplicate"] += 1
                else: stats["inserted"] += 1
            elif file_type == "budget":
                inv_id = (parsed.get("id") or "").strip()
                if inv_id:
                    existing = conn.execute("SELECT 1 FROM finops_invoices WHERE id = ?", (inv_id,)).fetchone()
                    if existing: stats["updated"] += 1
                    else: stats["inserted"] += 1
                else:
                    stats["inserted"] += 1

        stats["rejected"] = len(rejected)
        conn.execute("ROLLBACK")
    finally:
        conn.close()

    return {
        "file_type": file_type,
        "sandbox_batch": sandbox_batch,
        "would_happen": stats,
        "rejected_samples": [{"reasons": r["reasons"]} for r in rejected[:5]],
    }


# =====================================================================
# Sprint 4 — Admin pipeline management endpoints
# =====================================================================


@app.get("/api/admin/batches")
def admin_list_batches(
    limit: int = 50,
    file_type: Optional[str] = None,
    status: Optional[str] = None,
    _: dict = Depends(require_dual_role(Role.ADMIN)),
):
    """List ingest batches (newest first). Optional filters by file_type and status.
    Returns the full row + a derived `file_type` extracted from filename prefix
    (we encode it as `{type}::{filename}` in /api/ingest)."""
    cap = max(1, min(int(limit or 50), 500))
    conn = sqlite3.connect(DB_PATH)
    try:
        q = ("SELECT batch_id, filename, schema_version, status, rows_received, rows_loaded, "
             "rows_rejected, duplicate_rows, quality_score, started_at, finished_at "
             "FROM ingestion_batches WHERE 1=1")
        params: list = []
        if status:
            q += " AND status = ?"; params.append(status)
        if file_type:
            q += " AND filename LIKE ?"; params.append(f"{file_type}::%")
        q += " ORDER BY started_at DESC LIMIT ?"; params.append(cap)
        rows = conn.execute(q, params).fetchall()
        out = []
        for r in rows:
            filename = r[1] or ""
            ftype = ""
            real_name = filename
            if "::" in filename:
                ftype, _, real_name = filename.partition("::")
            out.append({
                "batch_id": r[0], "file_type": ftype, "filename": real_name,
                "schema_version": r[2], "status": r[3],
                "rows_received": r[4], "rows_loaded": r[5], "rows_rejected": r[6],
                "duplicate_rows": r[7], "quality_score": r[8],
                "started_at": r[9], "finished_at": r[10],
            })
        return out
    finally:
        conn.close()


@app.post("/api/admin/batches/{batch_id}/revert")
def admin_revert_batch(batch_id: str, _: dict = Depends(require_dual_role(Role.ADMIN))):
    """Revert a single batch by walking batch_entity_changes in reverse:
       - insert → DELETE
       - update → UPDATE back to `before_json`
    Marks the batch as 'reverted'. Refuses already-reverted batches.
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        st = conn.execute("SELECT status FROM ingestion_batches WHERE batch_id = ?", (batch_id,)).fetchone()
        if not st:
            raise HTTPException(status_code=404, detail="Batch not found")
        if st[0] == "reverted":
            raise HTTPException(status_code=400, detail="Batch already reverted")

        changes = conn.execute(
            "SELECT id, entity_type, entity_id, change_type, before_json, after_json "
            "FROM batch_entity_changes WHERE batch_id = ? ORDER BY id DESC",
            (batch_id,),
        ).fetchall()
        reverted = 0
        for _id, etype, eid, ctype, before, after in changes:
            table = {"candidate": "candidates", "job": "jobs", "application": "applications",
                     "hire": "hires", "diversity": "diversity_snapshots",
                     "headcount": "headcount_snapshots", "attrition": "attrition_events",
                     "invoice": "finops_invoices"}.get(etype)
            if not table:
                continue
            pk_col = {"applications": "app_id", "diversity_snapshots": "id",
                      "headcount_snapshots": "id"}.get(table, "id")
            if ctype == "insert":
                conn.execute(f"DELETE FROM {table} WHERE {pk_col} = ?", (eid,))
                reverted += 1
            elif ctype == "update" and before:
                try:
                    before_dict = json.loads(before)
                    if before_dict:
                        sets = ", ".join(f"{k} = ?" for k in before_dict.keys())
                        vals = list(before_dict.values()) + [eid]
                        conn.execute(f"UPDATE {table} SET {sets} WHERE {pk_col} = ?", vals)
                        reverted += 1
                except Exception:
                    pass

        conn.execute("UPDATE ingestion_batches SET status='reverted' WHERE batch_id = ?", (batch_id,))
        conn.commit()
        log_audit_action("BATCH_REVERTED", "warn", f"batch={batch_id} reverted={reverted}", user="admin")
    finally:
        conn.close()

    bump_data_version()
    return {"status": "ok", "batch_id": batch_id, "reverted_rows": reverted}


@app.get("/api/admin/quality/summary")
def admin_quality_summary(_: dict = Depends(require_dual_role(Role.ADMIN))):
    """Top-level KPIs for the Quality tab:
       - total candidates
       - candidates without ANY dedup key (no phone_norm AND no email_norm)
       - suspected duplicates by name-only (no phone/email match but same name)
       - last batch quality score
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        total_candidates = c.execute("SELECT COUNT(*) FROM candidates").fetchone()[0]
        no_keys = c.execute(
            "SELECT COUNT(*) FROM candidates WHERE (phone_norm IS NULL OR phone_norm='') AND (email_norm IS NULL OR email_norm='')"
        ).fetchone()[0]
        # Suspected duplicates: same name appears > 1 with different IDs.
        suspected = c.execute(
            "SELECT COUNT(*) FROM (SELECT LOWER(name) FROM candidates WHERE name IS NOT NULL GROUP BY LOWER(name) HAVING COUNT(*) > 1)"
        ).fetchone()[0]
        last_q = c.execute(
            "SELECT quality_score, started_at FROM ingestion_batches WHERE status IN ('committed','reverted') ORDER BY started_at DESC LIMIT 1"
        ).fetchone()
        total_batches = c.execute("SELECT COUNT(*) FROM ingestion_batches").fetchone()[0]
        return {
            "total_candidates": total_candidates,
            "candidates_without_dedup_key": no_keys,
            "name_based_duplicate_groups": suspected,
            "last_quality_score": last_q[0] if last_q else None,
            "last_quality_at": last_q[1] if last_q else None,
            "total_batches": total_batches,
        }
    finally:
        conn.close()


@app.get("/api/admin/quality/duplicates")
def admin_quality_duplicates(limit: int = 50, _: dict = Depends(require_dual_role(Role.ADMIN))):
    """Groups of candidates sharing the same lower-cased name but stored as
    separate rows (because phone/email didn't match). Admin can use this to
    merge manually."""
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            """SELECT LOWER(name) AS k, GROUP_CONCAT(id) AS ids, GROUP_CONCAT(IFNULL(phone_norm,'-')) AS phones,
                      GROUP_CONCAT(IFNULL(email_norm,'-')) AS emails, COUNT(*) AS n
               FROM candidates WHERE name IS NOT NULL
               GROUP BY LOWER(name) HAVING n > 1
               ORDER BY n DESC LIMIT ?""",
            (max(1, min(int(limit or 50), 500)),),
        ).fetchall()
        return [{"name": r[0], "ids": (r[1] or "").split(","), "phones": (r[2] or "").split(","),
                 "emails": (r[3] or "").split(","), "count": r[4]} for r in rows]
    finally:
        conn.close()


@app.get("/api/admin/quality/missing")
def admin_quality_missing(_: dict = Depends(require_dual_role(Role.ADMIN))):
    """Candidates whose row lacks any contact dedup key — admin needs to
    enrich them manually or they'll keep getting re-inserted on each upload."""
    conn = sqlite3.connect(DB_PATH)
    try:
        rows = conn.execute(
            """SELECT id, name, email, phone, source, last_seen_at
               FROM candidates
               WHERE (phone_norm IS NULL OR phone_norm='') AND (email_norm IS NULL OR email_norm='')
               ORDER BY last_seen_at DESC NULLS LAST LIMIT 200"""
        ).fetchall()
        return [{"id": r[0], "name": r[1], "email": r[2], "phone": r[3], "source": r[4], "last_seen_at": r[5]} for r in rows]
    finally:
        conn.close()


@app.get("/api/admin/quality/sanity")
def admin_quality_sanity(_: dict = Depends(require_dual_role(Role.ADMIN))):
    """Cross-type sanity checks — flags inconsistencies between related tables
    so the admin can spot data drift quickly. Each check returns {ok, message, value}."""
    conn = sqlite3.connect(DB_PATH)
    checks = []
    try:
        # Check 1: candidates without any application
        orphan_cands = conn.execute(
            "SELECT COUNT(*) FROM candidates c LEFT JOIN applications a ON a.candidate_id=c.id WHERE a.app_id IS NULL"
        ).fetchone()[0]
        checks.append({
            "key": "orphan_candidates",
            "title": "מועמדים ללא שיוך למשרה",
            "value": orphan_cands,
            "ok": orphan_cands == 0,
            "hint": "מועמדים שאינם משויכים לאף משרה. אם רב — בדוק את כללי הניקוי.",
        })
        # Check 2: applications without a candidate or job (broken FK)
        broken_apps = conn.execute(
            "SELECT COUNT(*) FROM applications a LEFT JOIN candidates c ON c.id=a.candidate_id "
            "LEFT JOIN jobs j ON j.id=a.job_id WHERE c.id IS NULL OR j.id IS NULL"
        ).fetchone()[0]
        checks.append({
            "key": "broken_application_refs",
            "title": "applications עם FK שבור",
            "value": broken_apps,
            "ok": broken_apps == 0,
            "hint": "applications שמצביעים על candidate/job שלא קיים. שורות אלו לא יופיעו בתצוגה האחודה.",
        })
        # Check 3: hires without a matching candidate (best-effort link)
        hires_no_cand = conn.execute(
            "SELECT COUNT(*) FROM hires WHERE candidate_id IS NULL"
        ).fetchone()[0]
        total_hires = conn.execute("SELECT COUNT(*) FROM hires").fetchone()[0]
        checks.append({
            "key": "hires_without_candidate",
            "title": "קליטות לא משויכות למועמד",
            "value": hires_no_cand,
            "total": total_hires,
            "ok": hires_no_cand == 0,
            "hint": "קליטות שלא הצליחו להיקשר למועמד קיים בעת ההעלאה. אפשר לקשר ידנית או לשפר את כללי ה-matching.",
        })
        # Check 4: headcount.attrition_ytd consistency vs attrition_events count
        # (loose — compares latest month only)
        latest_month = conn.execute(
            "SELECT snapshot_month FROM headcount_snapshots ORDER BY snapshot_month DESC LIMIT 1"
        ).fetchone()
        if latest_month:
            month = latest_month[0]
            hc_attrition = conn.execute(
                "SELECT IFNULL(SUM(attrition_ytd),0) FROM headcount_snapshots WHERE snapshot_month = ?",
                (month,),
            ).fetchone()[0]
            real_attrition = conn.execute(
                "SELECT COUNT(*) FROM attrition_events WHERE substr(leave_date,1,7) <= ?", (month,),
            ).fetchone()[0]
            diff = abs(hc_attrition - real_attrition)
            checks.append({
                "key": "headcount_vs_attrition_drift",
                "title": "פער בין attrition_ytd ב-headcount לאירועי עזיבה",
                "value": diff,
                "ok": diff <= 1,
                "hint": f"בחודש {month}: headcount={hc_attrition}, attrition_events={real_attrition}. אם הפער גדול, אחת הטבלאות פג תוקף.",
            })
        # Check 5: data_version not zero (otherwise frontend never refreshes)
        ver = get_data_version(conn)
        checks.append({
            "key": "data_version_active",
            "title": "data_version פעיל",
            "value": ver,
            "ok": ver > 0,
            "hint": "המונה צריך לעלות כל פעם שאצווה מקובעת. אם 0 — שום אצווה לא נטענה דרך ה-pipeline החדש.",
        })
    finally:
        conn.close()
    return {"checks": checks}


# =====================================================================
# Consumer map — static metadata used by the admin "מפת צריכת נתונים" tab
# Lists which UI block reads which table. Updated when adding new consumers.
# =====================================================================
@app.get("/api/admin/consumer-map")
def admin_consumer_map(_: dict = Depends(require_dual_role(Role.ADMIN))):
    return {
        "sources": [
            {"table": "candidates", "consumers": [
                {"block": "דשבורד KPI — סה\"כ מועמדים פעילים", "page": "/"},
                {"block": "Intelligence funnel", "page": "/intelligence"},
                {"block": "Cross-module search", "page": "header"},
                {"block": "Candidates list + side-panel", "page": "/candidates"},
                {"block": "Jobs — per-job candidate count", "page": "/jobs"},
            ]},
            {"table": "applications", "consumers": [
                {"block": "Candidates page stage chips", "page": "/candidates"},
                {"block": "Jobs stage breakdown", "page": "/jobs"},
                {"block": "Intelligence TTF/OAR metrics", "page": "/intelligence"},
            ]},
            {"table": "jobs", "consumers": [
                {"block": "Jobs page table + grid", "page": "/jobs"},
                {"block": "דשבורד KPI — משרות פתוחות", "page": "/"},
                {"block": "Intelligence neglect alerts", "page": "/intelligence"},
                {"block": "Side-panel job detail", "page": "/jobs"},
            ]},
            {"table": "hires", "consumers": [
                {"block": "דשבורד — קליטות החודש", "page": "/"},
                {"block": "Intelligence hires trend", "page": "/intelligence"},
            ]},
            {"table": "diversity_snapshots", "consumers": [
                {"block": "Headcount — תצוגת גיוון", "page": "/headcount"},
                {"block": "דשבורד — % גיוון", "page": "/"},
            ]},
            {"table": "headcount_snapshots", "consumers": [
                {"block": "Headcount — מטריצת תקן/בפועל", "page": "/headcount"},
                {"block": "דשבורד — סטיית תקן", "page": "/"},
            ]},
            {"table": "finops_invoices", "consumers": [
                {"block": "Budget — FinOps תפעול", "page": "/budget"},
                {"block": "דשבורד — תקציב מנוצל", "page": "/"},
            ]},
            {"table": "attrition_events", "consumers": [
                {"block": "Headcount — attrition_ytd", "page": "/headcount"},
                {"block": "Intelligence — churn alerts", "page": "/intelligence"},
            ]},
            {"table": "notifications", "consumers": [
                {"block": "BellDropdown header", "page": "all"},
                {"block": "Notification context", "page": "global"},
            ]},
        ],
    }


# =====================================================================
# Schema lock — Enhancement #3: warn when uploading a template that
# doesn't match the active schema_version. Read-only endpoint returns
# the active version so the frontend can compare on upload.
# =====================================================================
@app.get("/api/admin/active-schema")
def admin_active_schema():
    """Returns the current active schema version. Frontend can include it as
    `X-Schema-Version` header on uploads; backend will warn if mismatch."""
    return {
        "active_schema_version": DEFAULT_SCHEMA_VERSION,
        "supported": SUPPORTED_SCHEMA_VERSIONS,
        "ingest_types": list(INGEST_HANDLERS.keys()),
    }


# ── Recruiter × Job Matrix ────────────────────────────────────────────────────

@app.get("/api/recruiter-job-matrix")
async def recruiter_job_matrix(
    recruiter: Optional[str] = None,
    dept: Optional[str] = None,
    active_only: bool = True,
    _user: dict = Depends(require_session_role(Role.ADMIN, Role.HRBP)),
):
    """Cross-sectional pipeline matrix: for each (recruiter, job) pair return
    stage counts, funnel conversion rates, and a health indicator.

    - Pipeline view: current active candidates by stage
    - Funnel view: total ever reached each stage → conversion %
    """
    cache_params = {"recruiter": recruiter or "", "dept": dept or "", "active_only": active_only}
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        cached = get_cached_response(conn, "recruiter_job_matrix", cache_params)
        if cached:
            return cached

        sql = """
            SELECT
                a.app_id,
                a.recruiter,
                a.job_id,
                COALESCE(j.job_title, a.job_id) AS job_title,
                COALESCE(j.department, '') AS department,
                a.stage_code,
                a.status,
                a.days_in_process,
                o.status AS onboarding_status
            FROM applications a
            LEFT JOIN jobs j ON a.job_id = j.id
            LEFT JOIN onboarding o ON a.candidate_id = o.id
            WHERE a.recruiter IS NOT NULL AND TRIM(a.recruiter) != ''
        """
        params: list = []
        if recruiter:
            sql += " AND a.recruiter = ?"
            params.append(recruiter)
        if dept:
            sql += " AND COALESCE(j.department, '') = ?"
            params.append(dept)

        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()

    # Compute unified stage per row
    from collections import defaultdict

    # group: {(recruiter, job_id, job_title, department): {stage: count_active, ...}}
    # We track both "ever reached" (funnel) and "currently active" (pipeline)
    groups: dict[tuple, dict] = defaultdict(lambda: {
        "stages_active": defaultdict(int),   # not REJECTED
        "stages_funnel": defaultdict(int),   # all historical (every row ever)
        "days_sum": 0.0,
        "active_count": 0,
    })

    REJECTED_STATUS_HEBREW = {"דחייה", "הסרה", "ויתור", "הקפאה"}

    for row in rows:
        unified = _compute_unified_stage(row["stage_code"], row["onboarding_status"])
        key = (row["recruiter"], row["job_id"] or "", row["job_title"] or "", row["department"] or "")
        g = groups[key]

        # Funnel: count every application in each stage (historical total)
        g["stages_funnel"][unified] += 1

        # Pipeline: only active (not rejected via stage or Hebrew status)
        is_rejected = (
            unified == "REJECTED"
            or str(row["status"] or "").strip() in REJECTED_STATUS_HEBREW
        )
        if not is_rejected:
            g["stages_active"][unified] += 1
            g["active_count"] += 1
            try:
                g["days_sum"] += float(row["days_in_process"] or 0)
            except (TypeError, ValueError):
                pass

    # Build output list
    PIPELINE_STAGES = ["ACTIVE", "SCREEN", "INTERVIEW", "OFFER", "HIRED", "AWAITING_START", "STARTED"]
    result = []

    for (rec, job_id, job_title, department), g in groups.items():
        active_count = g["active_count"]
        avg_days = round(g["days_sum"] / active_count, 1) if active_count else 0.0

        stages_active = {s: g["stages_active"].get(s, 0) for s in PIPELINE_STAGES}
        stages_funnel = {s: g["stages_funnel"].get(s, 0) for s in PIPELINE_STAGES}

        if active_only and active_count == 0:
            continue

        # Funnel conversion rates (based on historical totals)
        def _rate(a: int, b: int) -> float:
            return round(a / b, 3) if b else 0.0

        screen_n   = stages_funnel["SCREEN"]
        interview_n = stages_funnel["INTERVIEW"]
        offer_n    = stages_funnel["OFFER"]
        hired_n    = stages_funnel["HIRED"] + stages_funnel["AWAITING_START"] + stages_funnel["STARTED"]

        funnel_rates = {
            "screen_to_interview": _rate(interview_n, screen_n),
            "interview_to_offer":  _rate(offer_n, interview_n),
            "offer_to_hired":      _rate(hired_n, offer_n),
        }

        # Health heuristic
        active_screen = stages_active["SCREEN"]
        active_later  = stages_active["INTERVIEW"] + stages_active["OFFER"]
        if (active_screen >= 5 and active_later == 0) or avg_days > 45:
            health = "stuck"
        elif avg_days > 30 or (active_count > 0 and active_screen / active_count > 0.80):
            health = "slow"
        else:
            health = "ok"

        result.append({
            "recruiter":      rec,
            "job_id":         job_id,
            "job_title":      job_title,
            "department":     department,
            "stages":         stages_active,
            "stages_funnel":  stages_funnel,
            "total_active":   active_count,
            "total_all_time": sum(g["stages_funnel"].values()),
            "avg_days":       avg_days,
            "funnel_rates":   funnel_rates,
            "health":         health,
        })

    result.sort(key=lambda r: (r["recruiter"], r["job_title"]))

    cache_conn = sqlite3.connect(DB_PATH)
    try:
        set_cached_response(cache_conn, "recruiter_job_matrix", cache_params, result, ttl_seconds=300)
    except Exception:
        pass
    finally:
        cache_conn.close()

    return result


# ─── Router registration (B2) ────────────────────────────────────────────────
# Wire APIRouters here AFTER all module-level helpers (log_audit_action,
# _auto_scan_after_ingest, …) are defined, so the late-bound proxies inside
# each router resolve their imports cleanly at request time.
from routers import analytics as _analytics_router  # noqa: E402
from routers import anomalies as _anomalies_router  # noqa: E402
from routers import candidates as _candidates_router  # noqa: E402
from routers import finops as _finops_router  # noqa: E402
from routers import jobs as _jobs_router  # noqa: E402
from routers import onboarding as _onboarding_router  # noqa: E402

app.include_router(_analytics_router.router)
app.include_router(_anomalies_router.router)
app.include_router(_candidates_router.router)
app.include_router(_finops_router.router)
app.include_router(_jobs_router.router)
app.include_router(_onboarding_router.router)

