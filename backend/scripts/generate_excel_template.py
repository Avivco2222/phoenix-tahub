from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "name",
    "email",
    "job_title",
    "status",
    "recruiter",
    "start_date",
    "department",
    "source",
]

SAMPLE_ROWS = [
    ["Dana Cohen", "dana.cohen@example.com", "Backend Engineer", "חדש", "מור", "2026-05-10", "R&D", "Excel Upload"],
    ["Avi Levi", "avi.levi@example.com", "Data Analyst", "בראיון", "גיא", "2026-05-11", "Finance", "Excel Upload"],
    ["Noa Mizrahi", "noa.mizrahi@example.com", "Customer Success", "התקבל", "ליטל", "2026-05-12", "Service", "Excel Upload"],
]


def main() -> None:
    target = Path(__file__).resolve().parents[2] / "docs" / "release" / "excel-upload-template-v1.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "DATA"

    header_fill = PatternFill(start_color="002649", end_color="002649", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = data_ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            data_ws.cell(row=row_idx, column=col_idx, value=value)

    col_widths = {
        "A": 24,
        "B": 30,
        "C": 26,
        "D": 18,
        "E": 18,
        "F": 16,
        "G": 18,
        "H": 22,
    }
    for col, width in col_widths.items():
        data_ws.column_dimensions[col].width = width

    data_ws.freeze_panes = "A2"

    status_validation = DataValidation(
        type="list",
        formula1='"חדש,בסינון,בראיון,הצעה,התקבל,נדחה"',
        allow_blank=False,
    )
    data_ws.add_data_validation(status_validation)
    status_validation.add("D2:D5000")

    date_validation = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2100,12,31)")
    data_ws.add_data_validation(date_validation)
    date_validation.add("F2:F5000")

    guide_ws = wb.create_sheet("INSTRUCTIONS")
    guide_ws["A1"] = "Excel Upload Template - Best Practice"
    guide_ws["A1"].font = Font(bold=True, size=14)
    guide_ws["A3"] = "Use DATA sheet only for upload."
    guide_ws["A4"] = "Required columns (do not rename): name,email,job_title,status,recruiter,start_date,department,source"
    guide_ws["A5"] = "start_date format: YYYY-MM-DD"
    guide_ws["A6"] = "Avoid duplicates by checking unique pair: email + job_title"
    guide_ws["A7"] = "Recommended upload headers: X-Schema-Version=1.0 and a unique X-Idempotency-Key"
    guide_ws.column_dimensions["A"].width = 120

    wb.save(target)
    print(target)


if __name__ == "__main__":
    main()
